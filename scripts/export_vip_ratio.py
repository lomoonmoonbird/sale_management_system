import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook


server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)


wb = load_workbook("./VIP_ratio.xlsx")

sheet_names = wb.sheetnames
sheet = wb[sheet_names[0]]

school_id = []
for row in list(sheet.iter_rows())[1:]:
    school_id.append(row[2].value)
print(school_id)
sql = "select id, school_id, student_vip_expire from sigma_account_us_user where available = 1 and role_id = 2 and school_id in (%s)" % (",".join([str(id) for id in school_id]))

cursor = connection.cursor()
cursor.execute(sql)
users = cursor.fetchall()

school_users_defaultdict = defaultdict(lambda : defaultdict(list))

for user in users:
    print(user)
    if school_users_defaultdict[user['school_id']]['total']:
        school_users_defaultdict[user['school_id']]['total'].append(1)
    else:
        school_users_defaultdict[user['school_id']]['total'] = [1]
    if user['student_vip_expire'] !=0 and user['student_vip_expire'] >= time.time():
        if school_users_defaultdict[user['school_id']]['vips']:
            school_users_defaultdict[user['school_id']]['vips'].append(1)
        else:
            school_users_defaultdict[user['school_id']]['vips'] = [1]

for row in list(sheet.iter_rows())[1:]:
    vips = len(school_users_defaultdict.get(int(row[2].value), {}).get("vips")) if school_users_defaultdict.get(int(row[2].value), {}).get("vips") else 0
    total = len(school_users_defaultdict.get(int(row[2].value), {}).get("total")) if school_users_defaultdict.get(int(row[2].value), {}).get("total") else 0
    row[3].value = "{0:0.2f}%".format((vips / total if total else 0)*100)

wb.save('./bbb.xlsx')

server_mongo.stop()
server.stop()
cursor.close()
connection.close()