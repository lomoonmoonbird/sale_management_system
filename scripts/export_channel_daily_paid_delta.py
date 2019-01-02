import pymongo
import re
from collections import defaultdict
import json

from celery import Task
from enumconstant import Roles
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

LEVEL = {"primary": "小学", "junior":"初中", "senior": "高中"}


def all_day_between_2_date(begin, end):
    """
    获取两个日期之间的所有日期
    :param begin:
    :param end:
    :return:
    """

    end = datetime.datetime.strptime(end, "%Y-%m-%d")
    begin = datetime.datetime.strptime(begin, "%Y-%m-%d")
    delta = end - begin
    for i in range(delta.days + 1):
        yield (begin + timedelta(i)).strftime("%Y-%m-%d")

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

instance_db = mongo.instance



def daily_new_pay(mongo, channel_ids=[] ,begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                        end_time=datetime.datetime.now().strftime("%Y-%m-%d")):
    """
    每日新增付费
    :param request:
    :param channel_ids:
    :return:
    """
    items = []
    coll = mongo.channel_per_day
    if not begin_time:
        begin_time = begin_time
    if not end_time:
        end_time = end_time
    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids},
                    "day": {"$gte": begin_time, "$lte": end_time}
                }
            },
            {
                "$project": {
                    "channel": 1,
                    "day": 1,
                    "pay_amount": 1
                }
            },

            {"$group": {"_id": {"day": "$day", "channel": "$channel"},
                        "total_pay_amount_day": {"$sum": "$pay_amount"},
                        }
             },
            {
                "$project": {
                    "total_pay_amount_day": 1,

                }

            },
            {
                "$sort": {"_id.day": 1}
            }

        ])

    for item in item_count:
        items.append(item)

    return items


def channel_daily_sheet(items=[], area_map={}, channel_map={}, begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                            end_time=datetime.datetime.now().strftime("%Y-%m-%d")):
    """
    渠道每日付费excel
    :param items:
    :param area_map:
    :param channel_map:
    :return:
    """
    if not begin_time:
        begin_time =begin_time
    if not end_time:
        end_time = end_time

    # print("area_map", area_map)
    # print("channel_map", channel_map)
    wb = Workbook("所有渠道|2018-12-01--2018-12-31每日付费.xlsx")
    default = {
        "total_pay_amount": 0,
    }
    sheet_name = "sheet"
    if items:
        sheet_name = "渠道新增数据"
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]


        item_map_defaultdict = defaultdict(list)
        for item in items:
            item_map_defaultdict[item['_id']['channel']].append(item)

        channel_map_only_data = []
        area_channel_stat_defaultdict = defaultdict(list)
        for channel_id, data in channel_map.items():
            # if not item_map.get(channel_id):
            #     default['_id'] = channel_id
            #     items.append(default)
            data['stat'] = item_map_defaultdict.get(channel_id, [])
            channel_map_only_data.append(data['stat'])
            area_channel_stat_defaultdict[data.get("area_id", "")].append(data)


        all_days = list(all_day_between_2_date(begin_time, end_time))
        # ws.append(['大区名称', '渠道名称'] + all_days)

        # for row in range(0, len(channel_map_only_data)):
        #     stat_list = channel_map_only_data[row]
        #     area_name = channel_map.get(stat_list[0]['_id']['channel']).get("area_name")
        #     channel_name = channel_map.get(stat_list[0]['_id']['channel']).get("name")
        #     one_row = [area_name, channel_name]
        #
        #     for s in stat_list:
        #         if s['_id']['day'] != all_days[2+row]:
        #             one_row.append(0)
        #         else:
        #             one_row.append(s['total_pay_amount_day'])
        #     ws.append(one_row)
        # area_channel_stat_defaultdict["aaaaaaaaaaaaaaaaa"] = area_channel_stat_defaultdict["5c0539dbc6453208d23a8c86"]
        # print(json.dumps(area_channel_stat_defaultdict, indent=4))
        # print(len(area_channel_stat_defaultdict))
        for area_id, data in area_channel_stat_defaultdict.items():

            # print(data)
            for subdata in data:
                index = 0
                ws.append([subdata['area_name'] + "@" + subdata['name'], "新增付费金额"])
                stat_list = subdata['stat']
                stat_list_day_map = {}
                for s in stat_list:
                    stat_list_day_map[s['_id']['day']] = s
                stat_list_len = len(stat_list)
                all_days_len = len(all_days)
                # print(stat_list_len, all_days_len)
                for row in range(index, len(all_days)):
                    index+=1
                    ws.append([all_days[row], stat_list_day_map.get(all_days[row], {}).get("total_pay_amount_day", 0)])
                index += 2
    wb.save("所有渠道|2018-12-01--2018-12-31每日付费.xlsx")



area_name = ''

begin_time = "2018-12-01"
end_time = '2018-12-31'

# 渠道 begin
query = {"role": Roles.AREA.value, "status": 1}
if area_name:
    query.update({"name": {"$in": area_name.split(',')}})

areas = instance_db.find(query)
areas = list(areas)
area_ids = [str(item['_id']) for item in areas]
# print(areas)
channels = instance_db.find({"parent_id": {"$in": area_ids},"role": Roles.CHANNEL.value,"status": 1})
channels = list(channels)
# print(channels)
old_ids = [item['old_id'] for item in channels]

area_map = {}
for area in areas:
    area_map[str(area['_id'])] = area

sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
      ','.join([str(id) for id in old_ids])
print(old_ids)
cursor = connection.cursor()
cursor.execute(sql)
real_channels = cursor.fetchall()

channels_map = {}
for channel in channels:
    channels_map[channel['old_id']] = channel
channel_mongo_map = {}
for channel in real_channels:
    channel.update({"name": channel.get("name", ""),
                    "area_id": channels_map.get(channel['id'], {}).get("parent_id", ""),
                    "area_name": area_map.get(channels_map.get(channel['id'], {}).get("parent_id", "")).get(
                        "name", "")})
    channel_mongo_map[channel['id']] = channel
items = daily_new_pay(mongo, old_ids, begin_time, end_time)
print(json.dumps(items,indent=4))
channel_daily_sheet(items,area_map,channel_mongo_map,begin_time,end_time)


server_mongo.stop()
server.stop()
connection.close()
