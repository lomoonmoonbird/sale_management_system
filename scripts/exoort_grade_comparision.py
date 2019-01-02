import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

LEVEL = {"primary": "小学", "junior":"初中", "senior": "高中"}

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)


exclude_channel_sql = "select id from sigma_account_us_user WHERE id in (select channel_id from sigma_account_re_channel_group where group_id = 9);"
exlude_student_sql = "select id from sigma_account_us_user where school_id in ( select id from sigma_account_ob_school where owner_id in (%s) )"
student_that_has_upload_exercise_sql = "select distinct student_id from sigma_pool_as_hermes where available = 1  and time_create>='2018-12-28' and time_create <= '2019-01-01'"
school_group_sql = "select s.id, s.full_name, s.owner_id, g.name, g.stage, g.grade from sigma_account_ob_school as s join sigma_account_ob_group as g on s.available = 1 and s.id = g.school_id join sigma_account_re_groupuser as gu on g.id = gu.group_id and gu.user_id in (%s) group by g.school_id, g.grade,g.name"
all_channel_sql = "select id, username, name from sigma_account_us_user where available = 1 and role_id = 6"

cursor = connection.cursor()
cursor.execute(all_channel_sql)
all_channels = cursor.fetchall()

all_channels_map = {}
for channel in all_channels:
    all_channels_map[channel['id']] = channel


cursor = connection.cursor()
cursor.execute(exclude_channel_sql)
exclude_channel = cursor.fetchall()
exclude_channel_ids = [item['id'] for item in exclude_channel] + [8, 81]

print("exclude_channel_ids", exclude_channel_ids)

cursor = connection.cursor()
cursor.execute(exlude_student_sql % ','.join([str(id) for id in exclude_channel_ids]))
exclude_userid = cursor.fetchall()
exclude_userids = [item['id'] for item in exclude_userid]
print("exclude_userids", len(exclude_userids))
cursor = connection.cursor()
cursor.execute(student_that_has_upload_exercise_sql)
student_that_has_upload_exercise = cursor.fetchall()
student_that_has_upload_exercise_student_ids = [item['student_id'] for item in student_that_has_upload_exercise]

print("student_that_has_upload_exercise_student_ids", len(student_that_has_upload_exercise_student_ids))

student_ids = list( set(student_that_has_upload_exercise_student_ids).difference(set(exclude_userids)))


cursor = connection.cursor()
cursor.execute(school_group_sql % ','.join([str(id) for id in student_ids]))
student_info = cursor.fetchall()

channel_school_defaultdict = defaultdict(lambda: defaultdict(list))


for studentinfo in student_info:
    channel_key = str( all_channels_map.get(studentinfo['owner_id']).get("id") ) + "@" + all_channels_map.get(studentinfo['owner_id']).get("username") + "@" + all_channels_map.get(studentinfo['owner_id']).get("name")
    school_key = str(studentinfo['id']) + "@" + studentinfo['full_name']
    channel_school_defaultdict[  channel_key ][school_key].append({
        "channel_id": studentinfo['owner_id'],
        "school_id": studentinfo['id'],
        "stage": studentinfo['stage'],
        "class": studentinfo['name'],
        "grade": studentinfo['grade']
    })



# print(type(content_json))
for key, content in channel_school_defaultdict.items():
    wb = Workbook("./sheet/" + key.split('@')[2])
    # print(content)
    for subkey, subcontent in content.items():
        wb.create_sheet(title=subkey.split("@")[1])
        ws = wb[subkey.split("@")[1]]
        # print(ws)
        print(subcontent)
        ws.append(['渠道id', '学校id', '学段','班级','年级'])
        for row in range(0,len(subcontent)):

            ws.append([subcontent[row]['channel_id'], subcontent[row]['school_id'], LEVEL.get(subcontent[row]['stage'], "没找到学段"), subcontent[row]['class'], subcontent[row]['grade']])
            # ws.cell(row=row, column=1).value = subcontent[row]['stage']
            # ws.cell(row=row, column=2).value = subcontent[row]['class']
            # ws.cell(row=row, column=3).value = subcontent[row]['grade']
    wb.save("./sheet/" + key.split('@')[2]+'.xlsx')


with open("./check.json", mode='a') as f:
    f.write(json.dumps(channel_school_defaultdict,indent=4))
    f.flush()
print(json.dumps(channel_school_defaultdict,indent=4))
server_mongo.stop()
server.stop()
cursor.close()
connection.close()


