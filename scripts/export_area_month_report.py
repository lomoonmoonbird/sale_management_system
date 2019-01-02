import pymongo
import re
from collections import defaultdict
import json
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill,Alignment
from datetime import datetime, timedelta
from basehandler import BaseHandler
from celery import Task
from enumconstant import Roles
import pymysql
import pymongo
from bson import ObjectId
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
import os
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

LEVEL = {"primary": "小学", "junior":"初中", "senior": "高中"}


def _red_font():
    font = Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                color='FFFF0000')
    return font


def _black_font():
    font = Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                color='FF000000')
    return font


def _white_font():
    font = Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                color='FFFFFFFF')
    return font


def _alignment():
    return Alignment(horizontal='center', wrapText=True)


def _background_header_color():
    return PatternFill(bgColor="FFC7CE", fgColor="000000", fill_type="solid")


def _white_background_color():
    return PatternFill(fgColor="FFFFFF", fill_type="solid")


def _border():
    # border = Border(left=Side(border_style=None,color = 'FFDDDDDD'),right = Side(border_style=None,color = 'FFDDDDDD'),
    #        top = Side(border_style=None,color = 'FFDDDDDD'),bottom = Side(border_style=None,color = 'FFDDDDDD'),
    #        diagonal = Side(border_style=None,color = 'FFDDDDDD'),diagonal_direction = 0,
    #        outline = Side(border_style=None,color = 'FFDDDDDD'),vertical = Side(border_style=None,color = 'FFDDDDDD'),
    #        horizontal = Side(border_style=None,color = 'FFDDDDDD')
    # )
    thin = Side(border_style="thin", color="FF000000")
    double = Side(border_style="double", color="FF000000")

    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="DDDDDD")
    fill = GradientFill(stop=("000000", "FFFFFF"))
    font = Font(b=True, color="FF0000")
    al = Alignment(horizontal="center", vertical="center")
    return border



server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

instance_db = mongo.instance


def _list_month(mongo, channel_ids: list):
    """
    学校维度
    :param request:
    :param channel_ids:
    :return:
    """
    coll = mongo.channel_per_day
    items = []

    last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day, \
    curr_month_first_day, curr_month_last_day = BaseHandler()._curr_and_last_and_last_last_month()

    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids}
                }
            },
            {
                "$project": {
                    "channel": 1,
                    "school_number": 1,
                    "teacher_number": 1,
                    "student_number": 1,
                    "guardian_count": 1,
                    "guardian_unique_count": 1,
                    "pay_number": 1,
                    "pay_amount": 1,
                    "valid_reading_count": 1,
                    "valid_exercise_count" :1,
                    "valid_word_count": 1,
                    "e_image_c": 1,
                    "w_image_c": 1,
                    "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                    "total_pay_number": {
                        "$cond": [{"$and": [{"$gte": ["$day", "2016-01-01"]}]},
                                  "$pay_number", 0]},
                    "total_pay_amount": {
                        "$cond": [{"$and": [{"$gte": ["$day", "2016-01-01"]}]},
                                  "$pay_amount", 0]},

                    "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$school_number", 0]},
                    "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$school_number", 0]},

                    "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},
                    "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$teacher_number", 0]},

                    "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},
                    "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$student_number", 0]},

                    "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$guardian_count", 0]},
                    "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_unique_count", 0]},

                    "guardian_unique_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$guardian_unique_count", 0]},
                    "guardian_unique_number_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_count", 0]},

                    "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},
                    "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$pay_number", 0]},

                    "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},
                    "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$pay_amount", 0]},

                    "valid_reading_count_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_reading_count", 0]},
                    "valid_reading_count_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_reading_count", 0]},

                    "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},
                    "valid_exercise_count_last_month":
                        {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_exercise_count", 0]},

                    "valid_word_count_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},
                    "valid_word_count_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_word_count", 0]},

                    "e_image_c_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},
                    "e_image_c_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$e_image_c", 0]},

                    "w_image_c_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},
                    "w_image_c_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$w_image_c", 0]},

                    "day": 1
                }
            },

            {"$group": {"_id": "$channel",
                        "valid_reading_count": {"$sum": "$valid_reading_count"},
                        "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                        "valid_word_count": {"$sum": "$valid_word_count"},
                        "e_image_c": {"$sum": "$e_image_c"},
                        "w_image_c": {"$sum": "$w_image_c"},
                        "total_images": {"$sum": "$total_images"},
                        "total_school_number": {"$sum": "$school_number"},
                        "total_teacher_number": {"$sum": "$teacher_number"},
                        "total_student_number": {"$sum": "$student_number"},
                        "total_guardian_number": {"$sum": "$guardian_count"},
                        "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                        "total_pay_number": {"$sum": "$total_pay_number"},
                        "total_pay_amount": {"$sum": "$total_pay_amount"},
                        "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                        "school_number_last_month": {"$sum": "$school_number_last_month"},
                        "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                        "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                        "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                        "student_number_last_month": {"$sum": "$student_number_last_month"},
                        "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                        "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                        "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                        "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                        "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                        "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                        "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                        "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                        "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                        "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                        "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                        "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                        "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                        "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                        "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                        "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                        "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                        "w_image_c_last_month": {"$sum": "$w_image_c_last_month"},
                        "total_images_curr_month": {
                            "$sum": {"$add": ["$e_image_c_curr_month", "$w_image_c_curr_month"]}},
                        "total_images_last_month": {
                            "$sum": {"$add": ["$e_image_c_last_month", "$w_image_c_last_month"]}},
                        }
             },
            {
                "$project": {
                    "_id": 1,
                    "total_school_number": 1,
                    "total_teacher_number": 1,
                    "total_student_number": 1,
                    "total_guardian_number": 1,
                    "total_unique_guardian_number": 1,
                    "total_pay_number": 1,
                    "total_pay_amount": 1,
                    "valid_reading_count": 1,
                    "valid_exercise_count": 1,
                    "valid_word_count": 1,
                    "e_image_c": 1,
                    "w_image_c": 1,
                    "total_images": 1,
                    "total_images_curr_month": 1,
                    "total_images_last_month": 1,
                    "school_number_curr_month": 1,
                    "school_number_last_month": 1,
                    "teacher_number_curr_month": 1,
                    "teacher_number_last_month": 1,
                    "student_number_curr_month": 1,
                    "student_number_last_month": 1,
                    "guardian_number_curr_month": 1,
                    "guardian_number_last_month": 1,
                    "guardian_unique_number_curr_month": 1,
                    "guardian_unique_number_last_month": 1,
                    "pay_number_curr_month": 1,
                    "pay_number_last_month": 1,
                    "pay_amount_curr_month": 1,
                    "pay_amount_last_month": 1,
                    "valid_reading_count_curr_month": 1,
                    "valid_reading_count_last_month": 1,
                    "valid_exercise_count_curr_month": 1,
                    "valid_exercise_count_last_month": 1,
                    "valid_word_count_curr_month": 1,
                    "valid_word_count_last_month": 1,
                    "e_image_c_curr_month": 1,
                    "e_image_c_last_month": 1,
                    "w_image_c_curr_month": 1,
                    "w_image_c_last_month": 1,

                    # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                    # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                    # "school_MoM":
                }

            }

        ])

    for item in item_count:
        items.append(item)

    return items


def _list_month_channel_school(mongo, channel_ids ):
    """
    学校维度
    :param request:
    :param channel_ids:
    :return:
            """
    coll = mongo.school_per_day
    items = []

    last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day, \
    curr_month_first_day, curr_month_last_day = BaseHandler()._curr_and_last_and_last_last_month()

    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids}
                }
            },
            {
                "$project": {
                    "school_id": 1,
                    "channel": 1,
                    "teacher_number": 1,
                    "student_number": 1,
                    "guardian_count": 1,
                    "guardian_unique_count": 1,
                    "pay_number": 1,
                    "pay_amount": 1,
                    "valid_reading_count": 1,
                    "valid_exercise_count": 1,
                    "valid_word_count": 1,
                    "e_image_c": 1,
                    "w_image_c": 1,
                    "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                    "total_pay_number": {
                        "$cond": [{"$and": [{"$gte": ["$day", "2016-01-01"]}]},
                                  "$pay_number", 0]},
                    "total_pay_amount": {
                        "$cond": [{"$and": [{"$gte": ["$day", "2016-01-01"]}]},
                                  "$pay_amount", 0]},

                    "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$school_number", 0]},
                    "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$school_number", 0]},

                    "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},
                    "teacher_number_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$teacher_number", 0]},

                    "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},
                    "student_number_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$student_number", 0]},

                    "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$guardian_count", 0]},
                    "guardian_number_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_unique_count", 0]},

                    "guardian_unique_number_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_unique_count", 0]},
                    "guardian_unique_number_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_count", 0]},

                    "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},
                    "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$pay_number", 0]},

                    "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                        "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},
                    "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                        "$gte": ["$day", last_last_month_first_day]}]}, "$pay_amount", 0]},

                    "valid_reading_count_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_reading_count", 0]},
                    "valid_reading_count_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_reading_count", 0]},

                    "valid_exercise_count_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},
                    "valid_exercise_count_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_exercise_count", 0]},

                    "valid_word_count_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},
                    "valid_word_count_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_word_count", 0]},

                    "e_image_c_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},
                    "e_image_c_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$e_image_c", 0]},

                    "w_image_c_curr_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},
                    "w_image_c_last_month": {
                        "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$w_image_c", 0]},

                    "day": 1
                }
            },

            {"$group": {"_id": {"channel": "$channel", "school_id": "$school_id"},
                        "valid_reading_count": {"$sum": "$valid_reading_count"},
                        "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                        "valid_word_count": {"$sum": "$valid_word_count"},
                        "e_image_c": {"$sum": "$e_image_c"},
                        "w_image_c": {"$sum": "$w_image_c"},
                        "total_images": {"$sum": "$total_images"},
                        "total_teacher_number": {"$sum": "$teacher_number"},
                        "total_student_number": {"$sum": "$student_number"},
                        "total_guardian_number": {"$sum": "$guardian_count"},
                        "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                        "total_pay_number": {"$sum": "$total_pay_number"},
                        "total_pay_amount": {"$sum": "$total_pay_amount"},
                        "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                        "school_number_last_month": {"$sum": "$school_number_last_month"},
                        "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                        "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                        "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                        "student_number_last_month": {"$sum": "$student_number_last_month"},
                        "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                        "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                        "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                        "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                        "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                        "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                        "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                        "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                        "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                        "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                        "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                        "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                        "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                        "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                        "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                        "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                        "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                        "w_image_c_last_month": {"$sum": "$w_image_c_last_month"},
                        "total_images_curr_month": {
                            "$sum": {"$add": ["$e_image_c_curr_month", "$w_image_c_curr_month"]}},
                        "total_images_last_month": {
                            "$sum": {"$add": ["$e_image_c_last_month", "$w_image_c_last_month"]}},
                        }
             },
            {
                "$project": {
                    "_id": 1,
                    "total_teacher_number": 1,
                    "total_student_number": 1,
                    "total_guardian_number": 1,
                    "total_unique_guardian_number": 1,
                    "total_pay_number": 1,
                    "total_pay_amount": 1,
                    "valid_reading_count": 1,
                    "valid_exercise_count": 1,
                    "valid_word_count": 1,
                    "e_image_c": 1,
                    "w_image_c": 1,
                    "total_images": 1,
                    "total_images_curr_month": 1,
                    "total_images_last_month": 1,
                    "school_number_curr_month": 1,
                    "school_number_last_month": 1,
                    "teacher_number_curr_month": 1,
                    "teacher_number_last_month": 1,
                    "student_number_curr_month": 1,
                    "student_number_last_month": 1,
                    "guardian_number_curr_month": 1,
                    "guardian_number_last_month": 1,
                    "guardian_unique_number_curr_month": 1,
                    "guardian_unique_number_last_month": 1,
                    "pay_number_curr_month": 1,
                    "pay_number_last_month": 1,
                    "pay_amount_curr_month": 1,
                    "pay_amount_last_month": 1,
                    "valid_reading_count_curr_month": 1,
                    "valid_reading_count_last_month": 1,
                    "valid_exercise_count_curr_month": 1,
                    "valid_exercise_count_last_month": 1,
                    "valid_word_count_curr_month": 1,
                    "valid_word_count_last_month": 1,
                    "e_image_c_curr_month": 1,
                    "e_image_c_last_month": 1,
                    "w_image_c_curr_month": 1,
                    "w_image_c_last_month": 1,

                    # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                    # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                    # "school_MoM":
                }

            }

        ])

    for item in item_count:
        items.append(item)

    return items


def sheet(template, items, school_items, channel_map, real_schools, users, report_type, area_name):
    file = load_workbook(template)
    sheet_names = file.sheetnames
    sheet = file[sheet_names[0]]
    default = {
        "valid_reading_count": 0,
        "valid_exercise_count": 0,
        "valid_word_count": 0,
        "e_image_c": 0,
        "w_image_c": 0,
        "total_images": 0,
        "total_school_number": 0,
        "total_teacher_number": 0,
        "total_student_number": 0,
        "total_guardian_number": 0,
        "total_unique_guardian_number": 0,
        "total_pay_number": 0,
        "total_pay_amount": 0,
        "school_number_curr_month": 0,
        "school_number_last_month": 0,
        "teacher_number_curr_month": 0,
        "teacher_number_last_month": 0,
        "student_number_curr_month": 0,
        "student_number_last_month": 0,
        "guardian_number_curr_month": 0,
        "guardian_number_last_month": 0,
        "guardian_unique_number_curr_month": 0,
        "guardian_unique_number_last_month": 0,
        "pay_number_curr_month": 0,
        "pay_number_last_month": 0,
        "pay_amount_curr_month": 0,
        "pay_amount_last_month": 0,
        "valid_reading_count_curr_month": 0,
        "valid_reading_count_last_month": 0,
        "valid_exercise_count_curr_month": 0,
        "valid_exercise_count_last_month": 0,
        "valid_word_count_curr_month": 0,
        "valid_word_count_last_month": 0,
        "e_image_c_curr_month": 0,
        "e_image_c_last_month": 0,
        "w_image_c_curr_month": 0,
        "w_image_c_last_month": 0,
        "total_images_curr_month": 0,
        "total_images_last_month": 0
    }
    area_dimesion_items = {}
    items_map = {}
    for item in items:
        items_map[item["_id"]] = item

    summary_map = defaultdict(list)
    row1 = sheet[1]
    items = []
    for id, channel in channel_map.items():
        default['_id'] = id
        channel.update(items_map.get(int(id), default))
        items.append(channel)

    if report_type == 'week':
        last_week = BaseHandler().last_week()
        row1[0].value =area_name+ "大区_" + last_week[0] + "-" + last_week[6] + "周报数据"
        row1[0].border = _border()
    elif report_type == 'month':
        _, _, last_month, _, _, _ = BaseHandler()._curr_and_last_and_last_last_month()
        month = datetime.datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
        row1[0].value = area_name+" 大区" + str(month) + "月报数据"
        row1[0].border = _border()

    for one in list(sheet[2:3]):
        for cell in one:
            cell.font = _white_font()
            cell.fill = _background_header_color()
            cell.border = _border()
    for index, item in enumerate(items):
        row = sheet[index+4]

        for cell in row:
            cell.font = _black_font()
            cell.alignment = _alignment()
            cell.border = _border()
        #渠道名字
        row[0].value = channel_map.get(item['_id'], {}).get('name')
        # 新增学校
        mom = (item['school_number_curr_month'] - item['school_number_last_month'])/item['school_number_last_month'] \
            if item['school_number_last_month'] else 0
        row[1].value = item['total_school_number']
        row[2].value = str(item['school_number_last_month']) + '/' + str(item['school_number_curr_month'])
        summary_map[1].append(row[1].value)
        summary_map[2].append(row[2].value)

        # 新增教师数量
        mom = (item['teacher_number_curr_month'] - item['teacher_number_last_month']) / item[
            'teacher_number_last_month'] \
            if item['teacher_number_last_month'] else 0
        row[3].value = str(item['teacher_number_last_month']) + '/' + str(item['teacher_number_curr_month'])
        summary_map[3].append(row[3].value)

        # 新增学生数量
        mom = (item['student_number_curr_month'] - item['student_number_last_month']) / item[
            'student_number_last_month'] \
            if item['student_number_last_month'] else 0
        row[4].value = str(item['student_number_last_month']) + '/' + str(item['student_number_curr_month'])
        summary_map[4].append(row[4].value)
        # 新增考试数量
        mom = (item['valid_exercise_count_curr_month'] - item['valid_exercise_count_last_month']) / item[
            'valid_exercise_count_last_month'] \
            if item['valid_exercise_count_last_month'] else 0
        row[5].value = item['valid_exercise_count']
        row[6].value = str(item['valid_exercise_count_last_month']) + '/' + str(item['valid_exercise_count_curr_month'])
        summary_map[5].append(row[5].value)
        summary_map[6].append(row[6].value)
        # 单词
        mom = (item['e_image_c_curr_month'] - item['e_image_c_last_month']) / item[
            'e_image_c_last_month'] \
            if item['e_image_c_last_month'] else 0
        row[7].value = item['valid_word_count']
        row[8].value = str(item['valid_word_count_last_month']) + '/' + str(item['valid_word_count_curr_month'])
        summary_map[7].append(row[7].value)
        summary_map[8].append(row[8].value)
        # 新增阅读
        mom = (item['valid_word_count_curr_month'] - item['valid_word_count_last_month']) / item[
            'valid_word_count_last_month'] \
            if item['valid_word_count_last_month'] else 0
        row[9].value = item['valid_reading_count']
        row[10].value = str(item['valid_reading_count_last_month']) + "/" + str(item['valid_reading_count_curr_month'])
        summary_map[9].append(row[9].value)
        summary_map[10].append(row[10].value)

        # 图像总数
        mom = (item['w_image_c_curr_month'] - item['w_image_c_last_month']) / item[
            'w_image_c_last_month'] \
            if item['w_image_c_last_month'] else 0
        # row[24].value = sum(channel_data['w_image_c'])
        row[11].value = item['total_images']
        row[12].value = str(item['total_images_last_month']) + '/' + str(item['total_images_curr_month'])
        summary_map[11].append(row[11].value)
        summary_map[12].append(row[12].value)

        # 绑定
        mom = (item['valid_reading_count_curr_month'] - item['valid_reading_count_last_month']) / item[
            'valid_reading_count_last_month'] \
            if item['valid_reading_count_last_month'] else 0
        avg = item['total_unique_guardian_number'] / item['total_student_number'] if item[
                                                                                         'total_student_number'] > 0 else 0
        row[13].value = item['total_unique_guardian_number']
        row[14].value = BaseHandler().percentage(avg)
        row[15].value = str(item['guardian_unique_number_last_month']) + '/' + str(item['guardian_unique_number_curr_month'])
        summary_map[13].append(row[13].value)
        summary_map[14].append(str(item['total_unique_guardian_number']) + "/" + str(item['total_student_number']))
        summary_map[15].append(row[15].value)

        # 新增付费
        mom = (item['pay_amount_curr_month'] - item['pay_amount_last_month']) / item['pay_amount_last_month']\
            if item['pay_amount_last_month'] else 0
        row[16].value = item['total_pay_amount']
        row[17].value = str(item['pay_amount_last_month']) + '/' + str(item['pay_amount_curr_month'])
        summary_map[16].append(row[16].value)
        summary_map[17].append(row[17].value)
        for one in row:
            if isinstance(one.value, (int, float)):
                if one.value == 0:
                    one.font = _red_font()

            if isinstance(one.value, (str)):
                if ("/" in one.value and one.value.split('/')[0] == "0") or\
                        ("/" in one.value and one.value.split('/')[1] == "0") or \
                        ("%" in one.value and one.value.split("%")[0] == '0.00'):
                    one.font = _red_font()

    total_offset = len(items) + 4
    divider = len(items)
    for index, cell in enumerate(sheet[total_offset]):
        cell.border =_border()
        if index == 0:
            cell.value = "总计"
            continue
        if index in (2, 3, 4, 6, 8, 10, 12, 15, 17): #总和带/
            cell.value = str(
                sum([float(item.split('/')[0]) for item in summary_map.get(index, "0/0")])) + "/" + str(
                sum([float(item.split('/')[1]) for item in summary_map.get(index, "0/0")]))
            cell.alignment = _alignment()
        elif index in (14,): #平均
            total_guardian = sum([float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
            total_student = sum([float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
            cell.value = BaseHandler().percentage(total_guardian / total_student if total_student else 0)
        else: #求和

            cell.value = BaseHandler().rounding(sum(summary_map.get(index,[0])))
            cell.alignment = _alignment()
        if isinstance(cell.value, (int, float)):
            if cell.value == 0:
                cell.font = _red_font()
        if isinstance(cell.value, (str)):
            if ("/" in cell.value and cell.value.split('/')[0] == "0.0") or \
                    ("/" in cell.value and cell.value.split('/')[1] == "0.0") or \
                    ("%" in cell.value and cell.value.split("%")[0] == '0.00') or cell.value == '0.00':
                cell.font = _red_font()
    # 环比计算
    total_offset = len(items) + 4 + 1
    for index, cell in enumerate(sheet[total_offset]):
        cell.font = _black_font()
        cell.font = _black_font()
        cell.alignment = _alignment()
        cell.border = _border()
        if index == 0:
            cell.value = "环比新增率"
            continue
        cell.alignment = _alignment()
        if index in (1, 5, 7, 9, 11, 13, 14, 16 ):  # 不带/last_summary
            cell.value = "/"
        elif index in (2, 3, 4, 6, 8, 10, 12, 15, 17):  # 带/
            last_summary = sum([float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
            curr_summary = sum([float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
            cell.value = BaseHandler().percentage(
                (curr_summary - last_summary) / last_summary if last_summary else 0)
        else:
            pass
        if isinstance(cell.value, (int, float)):
            if cell.value == 0:
                cell.font =_red_font()
        if isinstance(cell.value, (str)):
            if ("/" in cell.value and cell.value.split('/')[0] == "0") or \
                    ("/" in cell.value and cell.value.split('/')[1] == "0") or \
                    ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                cell.font = _red_font()

    #学校表格
    default = {
        "valid_reading_count": 0,
        "valid_exercise_count": 0,
        "valid_word_count": 0,
        "e_image_c": 0,
        "w_image_c": 0,
        "total_images": 0,
        "total_teacher_number": 0,
        "total_student_number": 0,
        "total_guardian_number": 0,
        "total_unique_guardian_number": 0,
        "total_pay_number": 0,
        "total_pay_amount": 0,
        "school_number_curr_month": 0,
        "school_number_last_month": 0,
        "teacher_number_curr_month": 0,
        "teacher_number_last_month": 0,
        "student_number_curr_month": 0,
        "student_number_last_month": 0,
        "guardian_number_curr_month": 0,
        "guardian_number_last_month": 0,
        "guardian_unique_number_curr_month": 0,
        "guardian_unique_number_last_month": 0,
        "pay_number_curr_month": 0,
        "pay_number_last_month": 0,
        "pay_amount_curr_month": 0,
        "pay_amount_last_month": 0,
        "valid_reading_count_curr_month": 0,
        "valid_reading_count_last_month": 0,
        "valid_exercise_count_curr_month": 0,
        "valid_exercise_count_last_month": 0,
        "valid_word_count_curr_month": 0,
        "valid_word_count_last_month": 0,
        "e_image_c_curr_month": 0,
        "e_image_c_last_month": 0,
        "w_image_c_curr_month": 0,
        "w_image_c_last_month": 0,
        "total_images_curr_month": 0,
        "total_images_last_month": 0
    }
    start_point = total_offset + 2
    school_item_channel_school_map = {}
    for s_item in school_items:
        school_item_channel_school_map[str(s_item['_id']['channel']) + "@" + str(s_item['_id']['school_id'])] = s_item

    channel_defauldict = defaultdict(list)
    for school in real_schools:
        key = str(school['owner_id']) + "@" + str(school['id'])

        one = school_item_channel_school_map.get(key, {})
        if not one:
            one.update(default)
        one.update(school)

        channel_defauldict[key.split("@")[0]].append(one)
    # channel_defauldict['2']=channel_defauldict["1"]
    # print(json.dumps(channel_defauldict, indent=4))
    delta = 0
    for channel, data in channel_defauldict.items():
        summary_channel_map = defaultdict(list)
        row = sheet[start_point + delta + 2]
        sheet.merge_cells(start_row=start_point + delta + 2, start_column=1, end_row=start_point + delta + 3,
                          end_column=21)
        if report_type == 'week':
            last_week = BaseHandler().last_week()
            row[0].value = channel_map.get(int(channel), {}).get('name', '') + last_week[0] + "-" + last_week[6] + "周报数据"
            row[0].border = _border()
            row[0].font = _black_font()
            row[0].alignment = _alignment()
        elif report_type == 'month':
            _, _, last_month, _, _, _ = BaseHandler()._curr_and_last_and_last_last_month()
            month = datetime.datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
            row[0].value = channel_map.get(int(channel), {}).get('name', '') + "渠道" + str(month) + "月报数据"
            row[0].border = _border()
            row[0].font = _black_font()
            row[0].alignment = _alignment()

        title_row1 = sheet[start_point + delta + 4]
        title_row2 = sheet[start_point + delta + 5]

        for cell in title_row1:
            cell.fill = _background_header_color()
            cell.font = _white_font()
            cell.border = _border()
            cell.alignment = _alignment()
        for cell in title_row2:
            cell.fill = _background_header_color()
            cell.font = _white_font()
            cell.border = _border()
            cell.alignment = _alignment()

        title_row1[0].value = "学校名称"
        title_row1[1].value = "教师总数"
        title_row1[2].value = "新增教师"
        title_row1[3].value = "学生总数"
        title_row1[4].value = "新增学生"
        title_row1[5].value = "考试总数"
        title_row1[6].value = "新增考试"
        title_row1[7].value = "单词总数"
        title_row1[8].value = "新增单词"
        title_row1[9].value = "阅读总数"
        title_row1[10].value = "新增阅读"
        title_row1[11].value = "绑定总数"
        title_row1[12].value = "绑定率"
        title_row1[13].value = "新增绑定"
        title_row1[14].value = "付费总数"
        title_row1[15].value = "新增付费"
        if report_type == 'week':
            title_row2[2].value = "上周/本周"
            title_row2[4].value = "上周/本周"
            title_row2[6].value = "上周/本周"
            title_row2[8].value = "上周/本周"
            title_row2[10].value = "上周/本周"
            title_row2[13].value = "上周/本周"
            title_row2[15].value = "上周/本周"
        else:
            title_row2[2].value = "上月/本月"
            title_row2[4].value = "上月/本月"
            title_row2[6].value = "上月/本月"
            title_row2[8].value = "上月/本月"
            title_row2[10].value = "上月/本月"
            title_row2[13].value = "上月/本月"
            title_row2[15].value = "上月/本月"

        index = 0
        for item in data:
            row = sheet[start_point + delta + index + 6]
            for cell in row:
                cell.font = _black_font()
                cell.alignment = _alignment()
                cell.border = _border()

            #学校名
            row[0].value = item['full_name']
            #教师
            row[1].value = item['total_teacher_number']
            row[2].value = str(item['teacher_number_last_month']) + '/' + str(item['teacher_number_curr_month'])
            summary_channel_map[1].append(row[1].value)
            summary_channel_map[2].append(row[2].value)
            #学生
            row[3].value = item['total_student_number']
            row[4].value = str(item['student_number_last_month']) + '/' + str(item['student_number_curr_month'])
            summary_channel_map[3].append(row[3].value)
            summary_channel_map[4].append(row[4].value)
            #考试
            row[5].value = item['valid_exercise_count']
            row[6].value = str(item['valid_exercise_count_last_month']) + '/' + str(item['valid_exercise_count_curr_month'])
            summary_channel_map[5].append(row[5].value)
            summary_channel_map[6].append(row[6].value)
            #单词
            row[7].value = item['valid_word_count']
            row[8].value = str(item['valid_word_count_last_month']) + '/' + str(
                item['valid_word_count_curr_month'])
            summary_channel_map[7].append(row[7].value)
            summary_channel_map[8].append(row[8].value)
            #阅读
            row[9].value = item['valid_reading_count']
            row[10].value = str(item['valid_reading_count_last_month']) + '/' + str(
                item['valid_reading_count_curr_month'])
            summary_channel_map[9].append(row[9].value)
            summary_channel_map[10].append(row[10].value)
            #绑定
            row[11].value = item['total_unique_guardian_number']
            row[12].value = BaseHandler().percentage( item['guardian_unique_number_last_month'] / item['total_student_number'] if item['total_student_number'] >0 else 0 )
            row[13].value =  str(item['guardian_unique_number_last_month']) + '/' + str(
                item['guardian_unique_number_curr_month'])
            summary_channel_map[11].append(row[11].value)
            summary_channel_map[12].append( str(item['guardian_unique_number_last_month']) + '/' + str(item['total_student_number']) )
            summary_channel_map[13].append(row[13].value)
            #付费
            row[14].value = item['total_pay_amount']
            row[15].value = str(item['pay_amount_last_month']) + '/' + str(
                item['pay_amount_curr_month'])
            summary_channel_map[14].append(row[14].value)
            summary_channel_map[15].append(row[15].value)
            index += 1

        delta += len(data) + 6
        divider = len(data)
        for index, cell in enumerate(sheet[start_point + delta]):

            cell.font =_black_font()
            cell.alignment = _alignment()
            cell.border = _border()

            cell.font = _black_font()
            if index == 0:
                cell.value = "总计"
                continue
            cell.alignment = _alignment()
            if index in (1, 3, 5, 7, 9, 11, 14):  # 不带/ 总和
                cell.value = BaseHandler().rounding(sum((summary_channel_map.get(index, [0]))))
            elif index in (2, 4, 6, 8, 10, 13, 15):  # 带/
                cell.value = str(
                    sum([float(item.split('/')[0]) for item in summary_channel_map.get(index, "0/0")])) + "/" + str(
                    sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, "0/0")]))
            elif index in (12,):
                total_guardian = sum(
                    [float(item.split('/')[0]) for item in summary_channel_map.get(index, ["0/0"])])
                total_student = sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, ["0/0"])])
                cell.value = BaseHandler().percentage(total_guardian / total_student if total_student else 0)
            else:
                pass
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.font = _red_font()
            if isinstance(cell.value, (str)):
                if ("/" in cell.value and cell.value.split('/')[0] == "0.0") or \
                        ("/" in cell.value and cell.value.split('/')[1] == "0.0") or \
                        ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                    cell.font = _red_font()

        delta += 1
        for index, cell in enumerate(sheet[start_point + delta]):
            cell.font = _black_font()
            cell.font = _black_font()
            cell.alignment = _alignment()
            cell.border = _border()
            if index == 0:
                cell.value = "环比新增率"
                continue
            cell.alignment = _alignment()
            if index in (1,3,5,7,9,11,12,14):  # 不带/last_summary
                cell.value = "/"
            elif index in (2,4,6,8,10,13,15):  # 带/
                last_summary = sum([float(item.split('/')[0]) for item in summary_channel_map.get(index, ["0/0"])])
                curr_summary = sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, ["0/0"])])
                cell.value = BaseHandler().percentage((curr_summary - last_summary) / last_summary if last_summary else 0)
            else:
                pass
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.font = _red_font()
            if isinstance(cell.value, (str)):
                if ("/" in cell.value and cell.value.split('/')[0] == "0") or \
                        ("/" in cell.value and cell.value.split('/')[1] == "0") or \
                        ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                    cell.font = _red_font()

        delta += 2
        file.save("自营大区所有渠道2018-12-01--2018-12-31的月报.xlsx")




area_id = "5c0b72145dea9ddfec8bcc29"

channels = instance_db.find({"parent_id": area_id,
                                                                     "role": Roles.CHANNEL.value,
                                                                     "status": 1})

channels = list(channels)
channel_ids = [str(item['_id'] for item in channels)]
channel_users = instance_db.find({"channel_id": {"$in": channel_ids},
                                                                   "instance_role_id": Roles.CHANNEL.value,
                                                                  "status": 1})
channel_users = list(channel_users)
# sheet = b''
old_ids = [item['old_id'] for item in channels]
area_name = ""
if old_ids:
    sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
          ','.join([str(id) for id in old_ids])

    cursor = connection.cursor()
    cursor.execute(sql)
    real_channels = cursor.fetchall()

    real_channels_map = {}
    for real in real_channels:
        real_channels_map[real['id']] = real
    channel_id_map = {}
    channel_oid_map = {}
    for channel in channels:
        channel_id_map[str(channel['_id'])] = real_channels_map.get(channel['old_id'])
        channel_oid_map[channel['old_id']] = real_channels_map.get(channel['old_id'])

    for user in channel_users:
        user['channel_info'] = channel_id_map.get(user['channel_id'], {})

    area_info = instance_db.find_one({"_id": ObjectId(area_id), "status": 1})
    area_name = area_info.get("name", "")

    school_sql = "select id, full_name,owner_id " \
                 "from sigma_account_ob_school " \
                 "where available = 1 " \
                 "and owner_id in (%s)" % (','.join([str(id) for id in old_ids]))

    cursor = connection.cursor()
    cursor.execute(school_sql)
    real_schools = cursor.fetchall()


    items = _list_month(mongo, old_ids)
    school_items = _list_month_channel_school(mongo, old_ids)
    template_path ="./templates/area_month_template.xlsx"


    sheet(template_path,items,school_items,real_channels_map,real_schools,channel_users,"month",area_name)

server_mongo.stop()
server.stop()
connection.close()

    # sheet = await request.app.loop.run_in_executor(self.thread_pool,
    #                                                self.sheet,
    #                                                template_path,
    #                                                items,
    #                                                school_items,
    #                                                real_channels_map,
    #                                                real_schools,
    #                                                channel_users,
    #                                                "month",
    #                                                area_name)



# return await self.replay_stream(sheet, area_name+"大区月报-"+datetime.now().strftime("%Y-%m-%d"), request)