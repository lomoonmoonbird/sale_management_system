#! python3.6
# --*-- coding: utf-8 --*--

"""
celery任务队列
同步不同角色的用户信息，信息包括学校，年级，班级
"""
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook


server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)






route = {1: "合心", 2: "外研"}


sheet = load_workbook("./channel_merchant.xlsx")

sheet_names = sheet.sheetnames
sheet = sheet[sheet_names[0]]






channel_users = []
for row in list(sheet.iter_rows())[1:]:
    # print (row[2].value)
    channel_users.append(row[2].value)
print(channel_users)
print(['"' + str(id) + '"' for id in channel_users])
print (",".join(channel_users))
sql_merchant = "select * from sigma_pay_ob_merchant"
sql_user = "select id,username from sigma_account_us_user where username in (%s)" % (",".join([ "'"+str(id)+"'" for id in channel_users]))
print(sql_user)
cursor = connection.cursor()
cursor.execute(sql_user)
ret = cursor.fetchall()
cursor.execute(sql_merchant)
merchants = cursor.fetchall()

channels_map = {}
for r in ret:
    channels_map[r['username']] = r['id']

merchant_map = {}
for merchant in merchants:
    merchant_map[merchant['channel_id']] = merchant['merchant']

channel_user_map = {}
channel_users = []
reverse_merchant = {1: 2, 2: 1}
update_sql_file = "./update_merchant_file.sql"
for row in list(sheet.iter_rows())[1:]:
    # print (row[2].value)
    channel_users.append(row[2].value)
    username = row[2].value
    merchant = row[3].value

    test_merchant = route.get(int(merchant_map.get(channels_map[username])))
    with open(update_sql_file, 'a') as f:
        if merchant!=test_merchant:
            print (username, channels_map[username], merchant, test_merchant, int(merchant_map.get(channels_map[username])))
            f.write("UPDATE `sigma_centauri_new`.`sigma_pay_ob_merchant` SET  `merchant` = %s WHERE `channel_id` = %s;" % (reverse_merchant.get(int(merchant_map.get(channels_map[username]))),channels_map[username]) + "\n\r")
            f.flush()
# print(ret)
cursor.close()
connection.close()