import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

sheet = load_workbook("./channel_id.xlsx")

sheet_names = sheet.sheetnames
sheet = sheet[sheet_names[0]]


channel_username = []
for row in list(sheet.iter_rows())[:]:
    # print (row[2].value)
    channel_username.append(row[0].value)
# print(channel_username)
# sql_merchant = "select * from sigma_pay_ob_merchant"
sql_user = "select id,username from " \
           "sigma_account_us_user " \
           "where username in (%s)" % (",".join([ "'"+str(id)+"'" for id in channel_username]))
# print(sql_user)

cursor = connection.cursor()
cursor.execute(sql_user)
ret = cursor.fetchall()

channel_map = {}
for r in ret :
    channel_map[r['id']] = r

channel_ids = [item['id'] for item in ret]

school_sql = "select id,owner_id, full_name " \
             "from sigma_account_ob_school" \
             " where available =1 and owner_id in (%s)" % (','.join([str(id) for id in channel_ids]))
cursor = connection.cursor()
cursor.execute(school_sql)
schools = cursor.fetchall()

school_map = {}

for school in schools:
    school_map[school['id']] = school

school_ids = [item['id'] for item in schools]
grade_sql = "select grade, school_id " \
            "from sigma_account_ob_group " \
            "where available = 1 " \
            "and school_id in (%s) group by school_id, grade " % (','.join(str(id) for id in school_ids))
# print(grade_sql)
cursor = connection.cursor()
cursor.execute(grade_sql)
grades = cursor.fetchall()

grade_ids = [item['grade'] for item in grades]
# print(json.dumps(grades, indent=4))
items = mongo.grade_per_day.aggregate([
    {
        "$match": {"grade": {"$in": grade_ids}}
    },
    {

        "$project": {
            "school_id": 1,
            "grade": 1,
            "channel": 1,
            "day": 1,
            "guardian_unique_count": 1,

            "student_number": 1,
            "valid_exercise_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", "2018-12-26"]}, {
                                "$gte": ["$day", "2018-09-01"]}]}, "$valid_exercise_count", 0]},
            "valid_word_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", "2018-12-26"]}, {
                                "$gte": ["$day", "2018-09-01"]}]}, "$valid_word_count", 0]},

            "valid_exercise_count_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", "2018-12-26"]}, {
                                "$gte": ["$day", "2018-12-01"]}]}, "$valid_exercise_count", 0]},
        }

    },
    {
        "$group": {
            "_id": {"school_id": "$school_id", "grade": "$grade", "channel": "$channel"},
            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
            "valid_word_count": {"$sum": "$valid_word_count"},
            "valid_exercise_count_month": {"$sum": "$valid_exercise_count_month"},
            "total_guardian": {'$sum': "$guardian_unique_count"},
            "total_student_number": {"$sum": "$student_number"}
        }

    },
    {
        "$project": {
            "channel": 1,
            "valid_exercise_count": 1,
            "valid_word_count": 1,
            "valid_exercise_count_month": 1,
            "total_guardian": 1,
            "total_student_number": 1,
            "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_guardian", "$total_student_number"]}]},
        }}

])

data = list(items)

# print(json.dumps(data, indent=4))
final_data = []

data_map = {}

for d in data:
    data_map[str(d['_id']['school_id']) + '@' + str(d['_id']['grade'])] = d

for grade in grades:
    one = {
        'channel_id': school_map.get(grade['school_id'], {}).get("owner_id"),
        'channel_name': channel_map.get(school_map.get(grade['school_id'], {}).get("owner_id"), {}).get('username', ''),
        'school_name': school_map.get(grade['school_id'], {}).get('full_name', ""),
        'school_id': grade['school_id'],
        'grade': grade['grade'],
        'total_bind_student': 0,
        'bind_ratio': 0,
        'total_exercise_number': 0,
        'total_word_number': 0,
        'total_exercise_number_month': 0
    }

    if data_map.get(str(grade['school_id']) + '@' + grade['grade']):
        item = data_map.get(str(grade['school_id']) + '@' + grade['grade'])
        one = {
            'channel_id': item['_id']['channel'],
            'channel_name': channel_map.get(item['_id']['channel'], {}).get('username', ''),
            'school_name': school_map.get(grade['school_id'], {}).get('full_name', ""),
            'school_id': item['_id']['school_id'],
            'grade': grade['grade'],
            'total_bind_student': item['total_guardian'],
            'bind_ratio': item['bind_ratio'],
            'total_exercise_number': item['valid_exercise_count'],
            'total_word_number': item['valid_word_count'],
            'total_exercise_number_month': item['valid_exercise_count_month']
        }

    final_data.append(one)


print(json.dumps(data, indent=4))
print(len(final_data))
print(len(grades))
print(len(data))
# for school in schools:
#     for d in data :
#         if school[id] == d['_id']['school_id']:
#             one = {
#                 'channel_id': d['_id']['channel'],
#                 'school_name': school['full_name'],
#                 'school_id': d['_id']
#             }

server_mongo.stop()
server.stop()
cursor.close()
connection.close()