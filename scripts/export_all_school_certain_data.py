import pymongo
import re
from collections import defaultdict
import json

from celery import Task
from enumconstant import Roles
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

LEVEL = {"primary": "小学", "junior":"初中", "senior": "高中"}

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

instance_db = mongo.instance


def school_new_delta(mongo, channel_ids=[], begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                           end_time=datetime.datetime.now().strftime("%Y-%m-%d")):
    """
    学校新增数据
    :param request:
    :param channel_ids:
    :param begin_time:
    :param end_time:
    :return:
    """
    items = []
    if not begin_time:
        begin_time = begin_time
    if not end_time:
        end_time = end_time

    coll = mongo.school_per_day

    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids},
                    # "day": {"$gte": begin_time, "$lte": end_time}
                }
            },
            {
                "$project": {
                    "channel": 1,
                    "school_id": 1,
                    "day": 1,
                    "student_number": 1,
                    "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                    "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                    "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                    "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                    "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                    "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},
                    "guardian_unique_count_range": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$guardian_unique_count", 0]},
                    "guardian_unique_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_unique_count", 0]},
                    "pay_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_number", 0]},
                    "pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]}
                }
            },

            {"$group": {"_id": {"channel": "$channel", "school_id": "$school_id"},
                        "total_school": {"$sum": "$school_number"},
                        "total_exercise": {"$sum": "$valid_exercise_count"},
                        "total_reading": {"$sum": "$valid_reading_count"},
                        "total_word": {"$sum": "$valid_word_count"},
                        "total_images": {"$sum": {"$add": ["$e_image_c", "$w_image_c"]}},
                        "total_unique_guardian_range": {"$sum": "$guardian_unique_count_range"},
                        "total_unique_guardian": {"$sum": "$total_unique_guardian_range"},
                        "total_pay_number": {"$sum": "$pay_number"},
                        "total_pay_amount": {"$sum": "$pay_amount"},
                        "total_student": {'$sum': "$student_number"},
                        "e_image_c": {"$sum": "$e_image_c"},
                        "w_image_c": {"$sum": "$w_image_c"}
                        }
             },
            {
                "$project": {
                    "total_school": 1,
                    "total_exercise": 1,
                    "total_reading": 1,
                    "total_word": 1,
                    "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},
                    "total_unique_guardian": 1,
                    "total_pay_number": 1,
                    "total_pay_amount": 1,
                    "bind_ratio": {"$cond": [{"$eq": ["$total_student", 0]}, 0,
                                             {"$divide": ["$total_unique_guardian",
                                                          "$total_student"]}]},
                }

            }

        ])

    for item in item_count:
        items.append(item)

    return items

def channel_new_delta( mongo, channel_ids=[], begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                            end_time=datetime.datetime.now().strftime("%Y-%m-%d")):
    """

    :param request:
    :param channel_ids:
    :return:
    """
    items = []
    if not begin_time:
        begin_time = begin_time
    if not end_time:
        end_time = end_time

    coll = mongo.channel_per_day
    print(begin_time, end_time, '00000000000000000000')
    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids},
                    # "day": {"$gte": begin_time, "$lte": end_time}
                }
            },
            {
                "$project": {
                    "channel": 1,
                    "day": 1,
                    "student_number": 1,
                    "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                    "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                    "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                    "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                    "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                    "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},
                    "guardian_unique_count_range": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$guardian_unique_count", 0]},
                    "guardian_unique_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_unique_count", 0]},
                    "pay_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_number", 0]},
                    "pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]}
                }
            },

            {"$group": {"_id": "$channel",
                        "total_school": {"$sum": "$school_number"},
                        "total_exercise": {"$sum": "$valid_exercise_count"},
                        "total_reading": {"$sum": "$valid_reading_count"},
                        "total_word": {"$sum": "$valid_word_count"},
                        # "total_images": {"$sum": {"$add": ["$e_image_c", "$w_image_c"]}}, #这种用法是错误的，如果某一条记录其中一个字段不存在所有的都会返回0
                        "total_unique_guardian_range": {"$sum": "$guardian_unique_count_range"},
                        "total_unique_guardian": {"$sum": "$total_unique_guardian_range"},
                        "total_pay_number": {"$sum": "$pay_number"},
                        "total_pay_amount": {"$sum": "$pay_amount"},
                        "total_student": {'$sum': "$student_number"},
                        "e_image_c": {"$sum": "$e_image_c"},
                        "w_image_c": {"$sum": "$w_image_c"}
                        }
             },
            {
                "$project": {
                    "total_school": 1,
                    "total_exercise": 1,
                    "total_reading": 1,
                    "total_word": 1,
                    "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},
                    "total_unique_guardian": 1,
                    "total_pay_number": 1,
                    "total_pay_amount": 1,
                    "bind_ratio": {"$cond": [{"$eq": ["$total_student", 0]}, 0,
                                             {"$divide": ["$total_unique_guardian",
                                                          "$total_student"]}]},
                }

            }

        ])

    for item in item_count:
        items.append(item)

    return items


def sheet( items=[], area_map=None, channel_map=None, school_items=[], channel_mysql_map={}, school_mysql_map={}, area_name_extra='',begin='',end=''):
    """
    生成excel
    :param items:
    :param area_map:
    :param channel_map:
    :return:
    """

    wb = Workbook('./channel_sheet/'+area_name_extra+ '|'+ begin + "--" + end+".xlsx")
    default = {
        "total_school": 0,
        "total_exercise": 0,
        "total_reading": 0,
        "total_word": 0,
        "total_images": 0,
        "total_unique_guardian": 0,
        "total_pay_number": 0,
        "total_pay_amount": 0,
        "bind_ratio": 0
    }
    sheet_name = "sheet"
    if items:
        sheet_name = "渠道新增数据"
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]


        item_map = {}
        for item in items:
            item_map[item['_id']] = item


        for channel_id, data in channel_map.items():
            if not item_map.get(channel_id):
                default['_id'] = channel_id
                items.append(default)
        ws.append(['大区名称', '渠道名称', '新增和学校数量', '新增测试使用次数', '新增阅读使用次数',
                   '新增单词使用次数', '新增图片数量', '新增绑定人数数量', '新增付费人数', '新增付费金额', '绑定率'])
        for row in range(0, len(items)):
            ws.append([channel_map.get(items[row]['_id']).get("area_name"), channel_map.get(items[row]['_id']).get("name"),
                       items[row]['total_school'], items[row]['total_exercise'],
                       items[row]['total_reading'], items[row]['total_word'],
                       items[row]['total_images'], items[row]['total_unique_guardian'],
                       items[row]['total_pay_number'], items[row]['total_pay_amount'], items[row]['bind_ratio']])
    if school_items:
        sheet_name = "学校新增数据"
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]


        item_map = {}
        for item in school_items:
            item_map[item['_id']['school_id']] = item

        for school_id, data in school_mysql_map.items():
            if not item_map.get(school_id):
                default["_id"] = {}
                default['_id']['channel'] = data['owner_id']
                default['_id']['school_id'] = data['id']
                school_items.append(default)
        ws.append(['渠道', '学校名称', '新增测试使用次数', '新增阅读使用次数',
                   '新增单词使用次数', '新增图片数量', '新增绑定人数数量', '新增付费人数', '新增付费金额', '绑定率'])

        print("school_mysql_map", school_mysql_map)
        for row in range(0, len(school_items)):

            print(school_items[row]['_id']['school_id'],
                  school_mysql_map.get(school_items[row]['_id']['school_id']), "school_mysql_map.get(school_items[row]['_id']['school_id'])")
            if not school_mysql_map.get(school_items[row]['_id']['school_id']):
                continue
            ws.append([channel_mysql_map.get(school_items[row]['_id']['channel']).get("name"),
                       school_mysql_map.get(school_items[row]['_id']['school_id']).get("full_name"),
                       school_items[row]['total_exercise'],
                       school_items[row]['total_reading'], school_items[row]['total_word'],
                       school_items[row]['total_images'], school_items[row]['total_unique_guardian'],
                       school_items[row]['total_pay_number'], school_items[row]['total_pay_amount'],
                       school_items[row]['bind_ratio']])

    wb.save('./channel_sheet/'+area_name_extra+ '|'+ begin + "--" + end+".xlsx")






channels = mongo.instance.find({"role": 3, "status": 1})
channels = list(channels)
channels_id = [item['old_id'] for item in channels]
school_sql = "select id, owner_id, full_name from sigma_account_ob_school where available = 1 and owner_id in (%s) " % (",".join([str(id) for id in channels_id]))
channels_info = "select id,username,name from sigma_account_us_user where available = 1 and id in (%s)" % (",".join([str(id) for id in channels_id]))
cursor = connection.cursor()
cursor.execute(school_sql)
real_schools = cursor.fetchall()
cursor = connection.cursor()
cursor.execute(channels_info)
channels_info = cursor.fetchall()
schools_id = [item['id'] for item in real_schools]
grade_sql = "select grade,school_id from sigma_account_ob_group where available = 1 and school_id in (%s) group by school_id, grade" % (",".join([str(id) for id in schools_id]))
cursor = connection.cursor()
cursor.execute(grade_sql)
grades = cursor.fetchall()

all_group_id_sql = "select id from sigma_account_ob_group where available =1 and school_id in (%s) "% (",".join([str(id) for id in schools_id]))
cursor = connection.cursor()
cursor.execute(all_group_id_sql)
all_group_id = cursor.fetchall()

all_student_sql = "select g.school_id,g.grade, gu.user_id,u.student_vip_expire from sigma_account_ob_group as g join sigma_account_re_groupuser as gu  on g.available = 1 and gu.available =1 and g.id = gu.group_id and gu.role_id = 2 join sigma_account_us_user as u on u.available =1 and u.id = gu.user_id where g.school_id in (%s) " % (",".join([str(id) for id in schools_id]))
cursor = connection.cursor()
cursor.execute(all_student_sql)
all_student = cursor.fetchall()

all_student_map = defaultdict(list)
for student in all_student:
    if student['student_vip_expire'] > 0 and time.time() < student['student_vip_expire']:
        all_student_map[str(student['school_id']) + '@' + str(student['grade'])].append(1)
school_channel_map = {}
for school in real_schools:
    school_channel_map[school['id']] = school

channel_name_map = {}
for channel in channels_info:
    channel_name_map[channel['id']] = channel

items = mongo.grade_per_day.aggregate([

    {
        "$match": {"channel": {"$in": channels_id}}
    },
    {
        "$project": {
            "day": 1,
            "channel": 1,
            "grade": 1,
            "school_id": 1,
            "student_number": 1,
            "guardian_unique_count": 1,
            "valid_exercise_count_semester": {"$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$valid_exercise_count", 0]},
            "valid_word_count_semester": {"$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$valid_word_count", 0]},
            "valid_exercise_count_month": {"$cond": [{"$and": [{"$lte": ["$day", "2019-01-14"]}, {
                            "$gte": ["$day", "2018-12-14"]}]}, "$valid_exercise_count", 0]},
        }
    },
    {
        "$group": {
            "_id": {"school_id": "$school_id", "grade": "$grade"},
            "total_guardian_unique_count": {"$sum": "$guardian_unique_count"},
            "total_valid_exercise_count_semester": {"$sum": "$valid_exercise_count_semester"},
            "total_valid_word_count_semester": {"$sum": "$valid_word_count_semester"},
            "total_valid_exercise_count_month": {"$sum": "$valid_exercise_count_month"},
            "total_student_number": {"$sum": "$student_number"}
        }
    },
    {
        "$project" : {
            "total_guardian_unique_count": 1,
            "total_valid_exercise_count_semester": 1,
            "total_valid_word_count_semester": 1,
            "total_valid_exercise_count_month": 1,
            "total_student_number": 1,
            "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_guardian_unique_count", "$total_student_number"]}]},
        }
    }

])
items_stat= []
for item in items:
    items_stat.append(item)


items_map = defaultdict(list)
for item in items_stat:
    items_map[str(item['_id']["school_id"]) + "@" + str(item['_id']['grade'])] = item
    # items_map[item['_id']["school_id"]].append(item)

default = {
"total_guardian_unique_count": 0,
            "total_valid_exercise_count_semester": 0,
            "total_valid_word_count_semester": 0,
            "total_valid_exercise_count_month": 0,
            "total_student_number": 0,
            "bind_ratio": 0
}
grade_map =  defaultdict(list)
for grade in grades :
    default['_id']={}
    default['_id']['school_id']=grade['school_id']
    default['_id']['grade'] = grade['grade']
    grade.update(items_map.get(str(grade['school_id']) + "@" + grade['grade'], default))
    grade_map[grade['school_id']].append(grade)

wb = Workbook('./all_school_certain_data.xlsx')
sheet_name = "年级数据"
wb.create_sheet(title=sheet_name)
ws = wb[sheet_name]

ws.append(['渠道', '学校名称', '学校id', '年级', '绑定学生数',
                   '绑定率', '本学期考试', '本学期单词', '最近一个月考试', 'vip开通率'])

for school_id, grade_list in grade_map.items():
    for row in range(0, len(grade_list)):
        print(grade_list[row])
        ws.append([channel_name_map.get(school_channel_map.get(grade_list[row]['_id']['school_id']).get("owner_id")).get("name"),
                   school_channel_map.get(grade_list[row]['_id']['school_id']).get("full_name"),
                   school_id, grade_list[row]['_id']['grade'],
                   grade_list[row].get("total_guardian_unique_count"),
                   grade_list[row].get("bind_ratio") if grade_list[row].get("bind_ratio") <= 1 else 1,
                   grade_list[row].get("total_valid_exercise_count_semester"),
                   grade_list[row].get("total_valid_word_count_semester"),
                   grade_list[row].get("total_valid_exercise_count_month"),

                   sum(all_student_map.get(str(grade_list[row]['_id']['school_id']) + "@" + grade_list[row]["_id"]['grade'], [])) / grade_list[row].get("total_student_number") if grade_list[row].get("total_student_number") else 0
                   ])
wb.save('./all_school_certain_data.xlsx')
print(json.dumps(items_stat,indent=4))

server_mongo.stop()
server.stop()
connection.close()






