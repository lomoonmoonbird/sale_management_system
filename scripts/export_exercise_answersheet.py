import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook
import requests
from operator import itemgetter

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

def custom_sort(strings):
    d = defaultdict(list)
    for i in strings:

        ii = i.split('-')
        if len(ii) < 2:
            d[int(ii[0])] = []
        elif len(ii) == 2:
            d[int(ii[0])].append(int(ii[1]))

        else:
            pass


    final = []

    for key, value in d.items():
        # print(value)
        d[key] = sorted(value)
    keys = list(d.keys())
    keys.sort()
    # print(keys)
    # print(d)
    for dd in keys:
        # print(dd)
        value = d.get(dd)
        # print(type(d))
        # print(d, '-', dd, '---', value)

        # print('ddddd', dd)
        key = dd
        if not value:
            # print('keykeykey', key)
            final.append(str(key))
        else:
            final += [str(key) + "-"+ str(inner) for inner in value]
    # for key, value in d.items():

    # print(final)
    return final

class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime.datetime)):
            return str(obj)
        return json.JSONEncoder.default(self, obj)

cursor = connection.cursor()

QUESTION_TYPE = {
    1: "选择",
    2: "判断",
    3: "填空",
    4: "作答",
    5: "短文改错",
    6: "材料",
    7: "选做",
    8: "书面表达",
    9: "作文"
}


QUESTION_TYPE_2 = {
    "choice": 1,
    # ""
}

EXERCISE = {
    "七年级": ["1f7ba06274", "3a2774a2cd", "45a8a458c8", "8b14430d2b", "9f0b66bcf0", "c0249afc94", "c88fdf9861", "e3d01432fd", "f1a3432609"],
    "八年级": ["2c6c8cc04f", "72d9625982", "817258da01", "82202cbd9f", "9b6e8cd6be", "b0343a4d0e", "b5f4a01733", "e0a7c4e60a", "ec96a9dd03"],
    # "八年级": ["2c6c8cc04f", "72d9625982", "817258da01", "82202cbd9f", "9b6e8cd6be", "b0343a4d0e", "b5f4a01733", "e0a7c4e60a"],
    "九年级": ["4f3787fb92", "65238b26de", "71dd0eea33", "61a9ecdd29", "e8b107edf2", "51ef59ec39", "ba1dcec96b", "fe43307dba", "6ae9395126"]
}

# EXERCISE = {
#     "七年级": ["3a2774a2cd"],
#     # "八年级": ["2c6c8cc04f", "72d9625982", "817258da01", "82202cbd9f", "9b6e8cd6be", "b0343a4d0e", "b5f4a01733", "e0a7c4e60a", "ec96a9dd03"],
#     # "九年级": ["4f3787fb92", "65238b26de", "71dd0eea33", "61a9ecdd29", "e8b107edf2", "51ef59ec39", "ba1dcec96b", "fe43307dba", "6ae9395126"]
# }

# EXERCISE = {
#     # "七年级": ["1f7ba06274", "3a2774a2cd", "45a8a458c8", "8b14430d2b", "9f0b66bcf0", "c0249afc94", "c88fdf9861", "e3d01432fd", "f1a3432609"],
#     # "八年级": ["2c6c8cc04f", "72d9625982", "817258da01", "82202cbd9f", "9b6e8cd6be", "b0343a4d0e", "b5f4a01733", "e0a7c4e60a", "ec96a9dd03"],
#     "八年级": ["2c6c8cc04f"],
#     # "九年级": ["4f3787fb92", "65238b26de", "71dd0eea33", "61a9ecdd29", "e8b107edf2", "51ef59ec39", "ba1dcec96b", "fe43307dba", "6ae9395126"]
# }

def get_choice_table(exercise_uid):
    x = exercise_uid[0]
    return "sigma_pool_as_choice_"+str(x)

def get_answersheet_table(exercise_uid):
    x = exercise_uid[0]
    # print('x', x, type(x))
    return "sigma_exercise_ob_answersheetitem_" + str(x)


subject_sql = "select * from sigma_pool_st_subject "
cursor.execute(subject_sql)
subjects = cursor.fetchall() #!!!!!
subject_map = {}
for s in subjects:
    subject_map[s['id']] = s['cn_name']
# print(json.dumps(list(set([item['question_type'] for item in requests.get("http://sigma-stable-gz.oss-cn-shanghai.aliyuncs.com/sigma/exercise/004c4e8db4/origin.json").json()])), indent=4))
for grade, exerciseid_list in EXERCISE.items(): #每个年级的考试
    for exerciseid in exerciseid_list: #每个考试
        exercise_info_sql = "select * from sigma_exercise_ob_exercise where uid = '%s'" % exerciseid

        cursor.execute(exercise_info_sql)
        exercise_info = cursor.fetchall() #!!!!!
        exercise_info = exercise_info[0] if exercise_info else {}

        answersheet_sql = "select * from %s where exercise_id = %s and available =1 " % (get_answersheet_table(exerciseid), exercise_info.get('id'))
        # answersheet_doi_sql = "select (distinct doi) from %s where exercise_id = %s" % (
        # get_answersheet_table(exerciseid), exercise_info.get('id'))
        # print(answersheet_sql)
        cursor.execute(answersheet_sql)
        answersheet = cursor.fetchall() #!!!!
        # cursor.execute(answersheet_doi_sql)

        answersheet_map = defaultdict(list)
        answersheet_user_map = defaultdict(list)
        for a_s in answersheet:
            answersheet_map[str(a_s['student_id'])+"@"+str(a_s['question_type'])].append(a_s)
            answersheet_user_map[a_s['student_id']].append(a_s)

        userids = list(set([item['student_id'] for item in answersheet]))

        choice_sql = "select * from %s where exercise_id = %s and available = 1" % (
        get_choice_table(exerciseid), exercise_info.get('id'))
        # print(choice_sql)
        cursor.execute(choice_sql)
        choices = cursor.fetchall()  # !!!!

        choices_map = {}
        for c in choices:
            choices_map[c['item_id']] = c
        print(userids)
        users_sql = "select * from sigma_account_us_user where id in (%s)" % ','.join([str(id) for id in userids])
        cursor.execute(users_sql)
        users = cursor.fetchall() # !!!!!
        user_map = {}
        for u in users:
            # if u['number'] == '17370006':
            #     print(u)
            user_map[u['id']] = u

        school_ids = [item['school_id'] for item in users]

        usergroup_sql = "select group_id, user_id from sigma_account_re_groupuser where user_id in (%s)" % ','.join([str(id) for id in userids])
        cursor.execute(usergroup_sql)
        usergroup = cursor.fetchall()  # !!!!!

        usergroup_map = {}
        for ug in usergroup:
            usergroup_map[ug['user_id']] = ug

        group_ids = [item['group_id'] for item in usergroup]

        group_sql = "select id, name, school_id from sigma_account_ob_group where id in (%s)" % ','.join(
            [str(id) for id in group_ids])
        # print('group_sql', group_sql)
        cursor.execute(group_sql)
        groups = cursor.fetchall()  # !!!!!

        group_map = {}
        for g in groups:
            group_map[g['id']] = g


        school_sql = "select * from sigma_account_ob_school where id in (%s)" % ','.join([str(id) for id in school_ids])
        # print(school_sql)
        cursor.execute(school_sql)
        schools = cursor.fetchall()  # !!!!!

        school_map = {}
        for school in schools:
            school_map[school['id']] = school


        wb = Workbook("./exercises/"+grade + "@" + exerciseid + ".xlsx")
        wb.create_sheet(title="aaa")
        ws = wb['aaa']
        flag = False
        users.insert(0, users[0]) #站位
        for index, user in enumerate(users):

            # if user['number'] == '17370006':
            #     print(user)
            #选择
            user_choice = answersheet_map.get(str(user['id']) + "@" + str(1))

            user_choice = sorted(user_choice, key=lambda k : k['doi']) if user_choice else []
            # print(user_choice)
            user_choice_map={}
            if user_choice:
                for u_c in user_choice:
                    user_choice_map[u_c['doi']] = u_c
            user_choice_len = len(user_choice)
            user_choice_select_column = [item['doi'] for item in user_choice]
            # user_choice_select_column = custom_sort(user_choice_select_column)

            # user_choice_select_column = [str(id) for id in user_choice_select_column]
            user_choice_score_column = [item['doi'] for item in user_choice]
            # user_choice_score_column = custom_sort(user_choice_score_column)
            # user_choice_score_column = [str(id) for id in user_choice_score_column]
            # print(user_choice_score_column)
            #判断
            user_judge = answersheet_map.get(str(user['id']) + "@" + str(2))
            user_judge_map = {}
            if user_judge:
                for u_c in user_judge:
                    user_judge_map[u_c['doi']] = u_c
            user_judge = sorted(user_judge, key=lambda k: k['doi']) if user_judge else []
            user_judge_len = len(user_judge) if user_judge else 0
            user_judge_column = [item['doi'] for item in user_judge] if user_judge else []

            #填空
            user_blank = answersheet_map.get(str(user['id']) + "@" + str(3))
            user_blank_map = {}
            if user_blank:
                for u_c in user_blank:
                    user_blank_map[u_c['doi']] = u_c
            user_blank = sorted(user_blank, key=lambda k: k['doi']) if user_blank else []
            user_blank_len = len(user_blank) if user_blank else 0
            user_blank_column = [item['doi'] for item in user_blank] if user_blank else []
            # 作答
            user_an = answersheet_map.get(str(user['id']) + "@" + str(4))
            user_an_map = {}
            if user_an:
                for u_c in user_an:
                    user_an_map[u_c['doi']] = u_c
            user_an = sorted(user_an, key=lambda k: k['doi']) if user_an else []
            user_an_len = len(user_an) if user_an else 0
            user_an_column = [ item['doi'] for item in user_an] if user_an else []
            # 短文改错
            user_coo = answersheet_map.get(str(user['id']) + "@" + str(5))
            user_coo_map = {}
            if user_coo:
                for u_c in user_coo:
                    user_coo_map[u_c['doi']] = u_c
            user_coo = sorted(user_coo, key=lambda k: k['doi']) if user_coo else []
            user_coo_len = len(user_coo) if user_coo else 0
            user_coo_column = [ item['doi'] for item in user_coo] if user_coo else []
            # 材料
            user_ma = answersheet_map.get(str(user['id']) + "@" + str(6))
            user_ma_map = {}
            if user_ma:
                for u_c in user_ma:
                    user_ma_map[u_c['doi']] = u_c
            user_ma = sorted(user_ma, key=lambda k: k['doi']) if user_ma else []
            user_ma_len = len(user_ma) if user_ma else 0
            user_ma_column = [item['doi'] for item in user_ma] if user_ma else []
            # 选做
            user_op = answersheet_map.get(str(user['id']) + "@" + str(7))
            user_op_map = {}
            if user_op:
                for u_c in user_op:
                    user_op_map[u_c['doi']] = u_c
            user_op = sorted(user_op, key=lambda k: k['doi']) if user_op else []
            user_op_len = len(user_op) if user_op else 0
            user_op_column = [ item['doi'] for item in user_op] if user_op else []
            # 书面表达
            user_exp = answersheet_map.get(str(user['id']) + "@" + str(8))
            user_exp_map = {}
            if user_exp:
                for u_c in user_exp:
                    user_exp_map[u_c['doi']] = u_c
            user_exp = sorted(user_exp, key=lambda k: k['doi']) if user_exp else []
            user_exp_len = len(user_exp)  if user_exp else 0
            user_exp_column = [ item['doi'] for item in user_exp]  if user_exp else []
            # 作文
            user_ar = answersheet_map.get(str(user['id']) + "@" + str(9))
            user_ar_map = {}
            if user_ar:
                for u_c in user_ar:
                    user_ar_map[u_c['doi']] = u_c
            user_ar = sorted(user_ar, key=lambda k: k['doi']) if user_ar else []
            user_ar_len = len(user_ar) if user_ar else 0
            user_ar_column = [item['doi'] for item in user_ar] if user_ar else []


            if flag == False:
                # print(user_choice_select_column)

                one = ["", "", "", "", "", "", "", ""]
                if user_choice:
                    one += ["选择"]*len(user_choice_select_column)
                if user_choice:
                    one +=["选择"] * len(user_choice_score_column)
                if user_judge:
                    one +=["判断"] * len(user_judge_column)
                if user_blank:
                    one +=["填空"] * len(user_blank_column)
                if user_an:
                    one +=["作答"] * len(user_an_column)
                if user_coo:
                    one +=["短文改错"] * len(user_coo_column)
                if user_ma:
                    one +=["材料"] * len(user_ma_column)
                if user_op:
                    one +=["选做"] * len(user_op_column)
                if user_exp:
                    one +=["书面表达"] * len(user_exp_column)
                if user_ar:
                    one +=["作文"] * len(user_ar_column)
                ws.append(one)
                ws.append(["学校", "班级", "科目", "考号", "姓名", "唯一号", "学号", "总分"] +
                          custom_sort(user_choice_select_column) +
                          custom_sort(user_choice_score_column) +
                          custom_sort(user_judge_column)+
                          custom_sort(user_blank_column)+
                          custom_sort(user_an_column)+
                          custom_sort(user_coo_column)+
                          custom_sort(user_ma_column)+
                          custom_sort(user_op_column)+
                          custom_sort(user_exp_column)+
                          custom_sort(user_ar_column)

                          )
                flag = True
                # print(["学校", "班级", "科目", "考号", "姓名", "唯一号", "学号", "总分"] +
                #           user_choice_select_column +
                #           user_choice_score_column +
                #           user_judge_column+
                #           user_blank_column+
                #           user_an_column+
                #           user_coo_column+
                #           user_ma_column+
                #           user_op_column+
                #           user_exp_column+
                #           user_ar_column)
            else:

                aaaa = [
                    school_map.get(user['school_id'], {}).get("full_name", ""),
                    group_map.get(usergroup_map.get(user['id'], {}).get("group_id", ""), {}).get("name"),
                    subject_map.get(exercise_info['subject_id']),
                    user['number'],
                    user['name'],
                    user['number'],
                    user['number'],
                    sum(item['final_score'] for item in answersheet_user_map.get(user['id'],[]))

                ]

                # print("aaaa", aaaa)

                if user_choice:

                    # aaaa = aaaa+[choices_map.get(item['id'], {}).get('correct_choice', "") for item in user_choice_select_column]
                    aaaa = aaaa + [choices_map.get(user_choice_map.get(item, {}).get("id", ""), {}).get("correct_choice") for item in custom_sort(user_choice_select_column)]
                if user_choice:
                    aaaa = aaaa + [user_choice_map.get(item, {}).get('final_score') for item in custom_sort(user_choice_select_column)]
                if user_judge:
                    aaaa = aaaa +[user_judge_map.get(item, {}).get('final_score') for item in custom_sort(user_judge_column)]
                if user_blank:
                    aaaa = aaaa +[user_blank_map.get(item, {}).get('final_score') for item in custom_sort(user_blank_column)]
                if user_an:
                    aaaa = aaaa +[user_an_map.get(item,{}).get('final_score') for item in custom_sort(user_an_column)]
                if user_coo:
                    aaaa = aaaa +[user_coo_map.get(item,{}).get('final_score') for item in custom_sort(user_coo_column)]
                if user_ma:
                    aaaa = aaaa +[user_ma_map.get(item, {}).get('final_score') for item in custom_sort(user_ma_column)]
                if user_op:
                    aaaa = aaaa +[user_op_map.get(item,{}).get('final_score') for item in custom_sort(user_op_column)]
                if user_exp:
                    aaaa = aaaa +[user_exp_map.get(item,{}).get('final_score') for item in custom_sort(user_exp_column)]
                if user_ar:
                    aaaa = aaaa +[user_ar_map.get(item,{}).get('final_score') for item in custom_sort(user_ar_column)]
                ws.append(aaaa)


        wb.save("./exercises/"+grade + "@" + exerciseid + ".xlsx")







       # for key, content in channel_school_defaultdict.items():
       #     wb = Workbook("./sheet/" + key.split('@')[2])
       #     # print(content)
       #     for subkey, subcontent in content.items():
       #         wb.create_sheet(title=subkey.split("@")[1])
       #         ws = wb[subkey.split("@")[1]]
       #         # print(ws)
       #         print(subcontent)
       #         ws.append(['渠道id', '学校id', '学段', '班级', '年级'])
       #         for row in range(0, len(subcontent)):
       #             ws.append([subcontent[row]['channel_id'], subcontent[row]['school_id'],
       #                        LEVEL.get(subcontent[row]['stage'], "没找到学段"), subcontent[row]['class'],
       #                        subcontent[row]['grade']])
       #             # ws.cell(row=row, column=1).value = subcontent[row]['stage']
       #             # ws.cell(row=row, column=2).value = subcontent[row]['class']
       #             # ws.cell(row=row, column=3).value = subcontent[row]['grade']
       #     wb.save("./sheet/" + key.split('@')[2] + '.xlsx')


        # print(json.dumps(answersheet_map, indent=4, cls=CustomEncoder))


server_mongo.stop()
server.stop()
cursor.close()
connection.close()