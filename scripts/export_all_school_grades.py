import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()


stat = mongo.grade_per_day.aggregate([
    {
        "$project": {
            "day": 1,
            "channel": 1,
            "school_id": 1,
            "grade": 1,

            "e_image_c": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2019-01-15"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$e_image_c", 0]},
            "w_image_c": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2019-01-15"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$w_image_c", 0]},


            "valid_exercise_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-12-01"]}]}, "$valid_exercise_count", 0]},

            "valid_word_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-12-01"]}]}, "$valid_word_count", 0]},
            "valid_reading_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-12-01"]}]}, "$valid_reading_count", 0]},

            "pay_amount": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2019-01-15"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$pay_amount", 0]},

        }
    },
    {
        "$group": {
            "_id": {"school_id": "$school_id", "grade": "$grade"},
            "e_image_c": {"$sum": "$e_image_c"},
            "w_image_c": {"$sum": "$w_image_c"},
            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
            "valid_word_count": {"$sum": "$valid_word_count"},
            "valid_reading_count": {"$sum": "$valid_reading_count"},
            "pay_amount": {"$sum": "$pay_amount"},

        }
    },
    {
        "$project": {
            "e_image_c": 1,
            "w_image_c": 1,
            "valid_exercise_count": 1,
            "valid_word_count": 1,
            "valid_reading_count": 1,
            "pay_amount": 1,
        }
    }
])

stat = list (stat)

# print(json.dumps(stat,indent=4))
stat_map = {}
for s in stat:
    stat_map[str(s["_id"]['school_id']) + "@" + s['_id']['grade']] = s

curr_page = 1
per_page = 100
total = 0
school_q = "select ss.id, ss.president, ss.full_name, ss.owner_id, ss.time_create " \
           "from (select id from sigma_account_ob_school where available = 1 order by id limit %s, %s) " \
           "s join sigma_account_ob_school ss on s.id = ss.id order by ss.id"


schoo_include_group_map = defaultdict(list)

default = {
            "e_image_c": 0,
            "w_image_c": 0,
            "valid_exercise_count": 0,
            "valid_word_count": 0,
            "valid_reading_count": 0,
            "pay_amount": 0,
}
schools_map = {}
while True:

    cursor.execute(school_q % ((curr_page-1)*per_page, per_page) )
    schools = cursor.fetchall()
    if not schools:
        break
    curr_page+=1
    total+=len(schools)


    for school in schools:
        schools_map[school['id']] = school

    school_ids = [item['id'] for item in schools]
    group_sql = "select grade, school_id, time_create from sigma_account_ob_group where available = 1 and school_id in (%s) group by grade, school_id " % (",".join([str(id) for id in school_ids]))

    cursor.execute(group_sql)
    groups = cursor.fetchall()

    group_map = {}
    for g in groups:
        group_map[g['school_id']] = g

    for g in groups:
        schoo_include_group_map[g['school_id']].append(g)
    # for s in schools:
    #     tmp = group_map.get(s['id'])
    #     if tmp:
    #         schoo_include_group_map[s['id']].append(tmp)

channel_sql = "select id, name from sigma_account_us_user where available = 1 and role_id = 6"
cursor.execute(channel_sql)
channels = cursor.fetchall()
channel_map = {}
for channel in channels:
    channel_map[channel['id']] = channel
print(len(schoo_include_group_map.keys()))
# print(len(test))
no_grade_schools = list(set(list(schools_map.keys())).difference(set(list(schoo_include_group_map.keys()))))
# print(len(schools_map.keys()))
wb = Workbook('./all_school_grade.xlsx')
sheet_name = "all_school_grade"
wb.create_sheet(title=sheet_name)
ws = wb[sheet_name]

ws.append(['渠道名', '学校', '负责人', '年级', '开通时间', '测试图像|2018-09-01--2019-01-15', '单词图像|2018-09-01--2019-01-15',
           '测试使用次数|2018-12-01--2018-12-31', '阅读使用次数|2018-12-01--2018-12-31',
           '词汇使用次数|2018-12-01--2018-12-31', '总和|2018-12-01--2018-12-31', '上学期付费金额|2018-09-01--2019-01-15'])


for school_id, s_g in schoo_include_group_map.items():
    for sg in s_g:
        ws.append([
            channel_map.get(schools_map[school_id].get("owner_id", ""), {}).get("name"),
            schools_map[school_id].get("full_name", ""),
            schools_map[school_id].get("president", ""),
            sg['grade'],
            sg['time_create'],
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("e_image_c"),
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("w_image_c"),
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_exercise_count"),
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_reading_count"),
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_word_count"),
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_exercise_count")+
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_reading_count")+
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_word_count"),
            stat_map.get(str(school_id) + "@" + sg['grade'], default).get("pay_amount"),
        ])




stat = mongo.school_per_day.aggregate([
    {
        "$match": {
            "school_id": {"$in": no_grade_schools}
        }
    },
    {
        "$project": {
            "day": 1,
            "channel": 1,
            "school_id": 1,

            "e_image_c": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2019-01-15"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$e_image_c", 0]},
            "w_image_c": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2019-01-15"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$w_image_c", 0]},


            "valid_exercise_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-12-01"]}]}, "$valid_exercise_count", 0]},

            "valid_word_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-12-01"]}]}, "$valid_word_count", 0]},
            "valid_reading_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2018-12-31"]}, {
                            "$gte": ["$day", "2018-12-01"]}]}, "$valid_reading_count", 0]},

            "pay_amount": {
                        "$cond": [{"$and": [{"$lte": ["$day", "2019-01-15"]}, {
                            "$gte": ["$day", "2018-09-01"]}]}, "$pay_amount", 0]},

        }
    },
    {
        "$group": {
            "_id": "$school_id",
            "e_image_c": {"$sum": "$e_image_c"},
            "w_image_c": {"$sum": "$w_image_c"},
            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
            "valid_word_count": {"$sum": "$valid_word_count"},
            "valid_reading_count": {"$sum": "$valid_reading_count"},
            "pay_amount": {"$sum": "$pay_amount"},

        }
    },
    {
        "$project": {
            "e_image_c": 1,
            "w_image_c": 1,
            "valid_exercise_count": 1,
            "valid_word_count": 1,
            "valid_reading_count": 1,
            "pay_amount": 1,
        }
    }
])
stat = list(stat)
stat_map = {}
for s in stat:
    stat_map[str(s["_id"])] = s


print(json.dumps(stat, indent=4, cls=CustomEncoder))

no_grade_school_sql = "select id ,full_name, owner_id, time_create from sigma_account_ob_school where available = 1 and id in (%s)" % ",".join([str(id) for id in no_grade_schools])
cursor.execute(no_grade_school_sql)
nogradeschools = cursor.fetchall()
nogradeschools_map = {}
for no in nogradeschools:
    nogradeschools_map[no['id']] =no

sheet_name = "all_school_no_grade"
wb.create_sheet(title=sheet_name)
ws = wb[sheet_name]

ws.append(['渠道名', '学校', '负责人', '开通时间', '测试图像|2018-09-01--2019-01-15', '单词图像|2018-09-01--2019-01-15',
           '测试使用次数|2018-12-01--2018-12-31', '阅读使用次数|2018-12-01--2018-12-31',
           '词汇使用次数|2018-12-01--2018-12-31', '总和|2018-12-01--2018-12-31', '上学期付费金额|2018-09-01--2019-01-15'])

for school_id, s_g in nogradeschools_map.items():

    ws.append([
        channel_map.get(nogradeschools_map[school_id].get("owner_id", ""), {}).get("name"),
        nogradeschools_map[school_id].get("full_name", ""),
        nogradeschools_map[school_id].get("president", ""),
        nogradeschools_map[school_id].get("time_create", ""),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("e_image_c"),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("w_image_c"),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_exercise_count"),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_reading_count"),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_word_count"),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_exercise_count")+
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_reading_count")+
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("valid_word_count"),
        stat_map.get(str(school_id) + "@" + sg['grade'], default).get("pay_amount"),
    ])






wb.save('./all_school_grade.xlsx')




# print(json.dumps(schools, indent=4, cls=CustomEncoder))





server_mongo.stop()
server.stop()
cursor.close()
connection.close()