import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
    ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init, worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook
from bson import ObjectId


class CustomEncoder(json.JSONEncoder):
    def encode(self, o):
        if isinstance(o, ObjectId):
            o = str(ObjectId)
        return o

term_start = "2019-02-15"
term_end = "2019-04-21"

day30_start = "2019-03-20"
day30_end = "2019-04-21"




def long_task(start, end):
    try:
        server = SSHTunnelForwarder(
            ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

            ssh_password="PengKim@89527",
            ssh_username="jinpeng",
            remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
        server.start()

        server_mongo = SSHTunnelForwarder(
            ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

            ssh_password="PengKim@89527",
            ssh_username="jinpeng",
            remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
        server_mongo.start()

        mongo = pymongo.MongoClient('127.0.0.1',
                                    username='root',
                                    password='sigmaLOVE2017',
                                    authMechanism='SCRAM-SHA-1',
                                    port=server_mongo.local_bind_port).sales
        connection = pymysql.connect(host="127.0.0.1",
                                     port=server.local_bind_port,
                                     user="sigma",
                                     password="sigmaLOVE2017",
                                     db=MYSQL_NAME,
                                     charset='utf8mb4',
                                     cursorclass=pymysql.cursors.DictCursor)

        cursor = connection.cursor()

        exclude_channel_sql = "select id from sigma_account_us_user" \
                              " WHERE id in (select channel_id from sigma_account_re_channel_group where group_id = 9)"
        cursor.execute(exclude_channel_sql)
        exclude_channels = cursor.fetchall()
        exclude_channel_ids = [item['id'] for item in exclude_channels] + [8] + [81]

        channels_sql = "select * from sigma_account_us_user where role_id = 6 " \
                       "and available = 1 " \
                       "and id not in (%s)" % ','.join([str(id) for id in exclude_channel_ids])

        # channels_sql = "select * from sigma_account_us_user where role_id = 6 " \
        #                "and available = 1 " \
        #                "and id =727722"

        cursor.execute(channels_sql)
        channels = cursor.fetchall()
        channels_map = {}
        for channel in channels:
            channels_map[channel['id']] = channel

        # print(channels)

        total_count = 0
        all_school = []
        for channel in channels[start:end]:
            school_grade_stage_defaultdict = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
            wb = Workbook("./all_channels_process/%s.xlsx" % channel['name'])
            sheet_name = "data"
            wb.create_sheet(title=sheet_name)
            ws = wb[sheet_name]

            ws.append(["渠道", "学校", "年级", "学段", "年级创建日期", "试用天数",
                       "本学期使用次数", "30天使用次数", "30天智能批改", "30天分级阅读", "30天词汇", "30天听说",
                       "在册学生数", "未绑定人数", "绑定率", "付费金额", "付费人数", "退款"])
            total_count += 1
            # print(channel['name'], total_count)
            school_sql = "select id,full_name, owner_id from sigma_account_ob_school where available =1  and owner_id = %s" % \
                         channel['id']
            cursor.execute(school_sql)
            schools = cursor.fetchall()
            all_school += schools
            for school in schools:  # 每个学校
                # print('school', school['id'])
                # print("school", school)
                school_id = school['id']

                groups_sql = "select * from sigma_account_ob_group where school_id = %s" % school_id
                cursor.execute(groups_sql)
                groups = cursor.fetchall()
                unique_exercise = set()
                unique_listen = set()
                unique_read = set()
                for group in groups:  # 每个班级

                    # print("group", group['id'])
                    group_id = group['id']
                    grade = group['grade']
                    stage = group['stage']
                    school_id = group['school_id']
                    school_grade_stage_key = str(school_id) + "@" + str(grade) + "@" + stage
                    school_groupid_grade_stage_key = str(school_id)  + "@" + str(grade) + "@" + stage + "@"+ str(group_id)
                    school_grade_stage_defaultdict[school_grade_stage_key]['date']["date"].append(group['time_create'])
                    user_sql = "select * from sigma_account_re_groupuser where group_id=%s and role_id=2" % group_id
                    teacher_sql = "select * from sigma_account_re_groupuser where available =1 and group_id=%s and role_id = 1" % group_id
                    cursor.execute(user_sql)
                    users = cursor.fetchall()
                    cursor.execute(teacher_sql)
                    teachers = cursor.fetchall()

                    hermes_term = []
                    hermes_day = []
                    word_term_id = []
                    word_day_id = []
                    if users:
                        # hermes
                        hermes_term_sql = "select student_id, exercise_id, time_create, count(distinct student_id) as total_user " \
                                          "from sigma_pool_as_hermes " \
                                          "where time_create >='%s' " \
                                          "and time_create <= '%s' " \
                                          " and available = 1 " \
                                          "and student_id in (%s) group by exercise_id having total_user >=10" \
                                          % (term_start, term_end, ",".join([str(item['user_id']) for item in users]))


                        hermes_day_sql = "select student_id, exercise_id, time_create, count(distinct student_id) as total_user " \
                                         "from sigma_pool_as_hermes " \
                                         "where time_create >='%s' " \
                                         "and time_create <= '%s' " \
                                         " and available = 1 " \
                                         "and student_id in (%s) group by exercise_id having total_user >=10" \
                                         % (day30_start, day30_end, ",".join([str(item['user_id']) for item in users]))
                        cursor.execute(hermes_term_sql)
                        hermes_term = cursor.fetchall()
                        cursor.execute(hermes_day_sql)
                        hermes_day = cursor.fetchall()


                    if hermes_term:

                        word_term_sql = 'select * from sigma_exercise_re_exercisemeta ' \
                                        'where available =1 ' \
                                        'and value = "word" ' \
                                        'and exercise_id in (%s)' % (
                                            ','.join([str(id['exercise_id']) for id in hermes_term]))
                        # print(word_term_sql)
                        cursor.execute(word_term_sql)
                        word_term = cursor.fetchall()
                        word_term_id = [item['exercise_id'] for item in word_term]
                    if hermes_day:
                        word_day_sql = 'select * from sigma_exercise_re_exercisemeta ' \
                                       'where available =1 ' \
                                       'and value = "word" ' \
                                       'and exercise_id in (%s)' % (
                                           ','.join([str(id['exercise_id']) for id in hermes_day]))
                        cursor.execute(word_day_sql)
                        word_day = cursor.fetchall()
                        word_day_id = [item['exercise_id'] for item in word_day]

                    for hermesterm in hermes_term:

                        if  school_grade_stage_key +"@"+str(hermesterm['exercise_id']) not in unique_exercise:
                            if hermesterm['exercise_id'] in word_term_id:
                                school_grade_stage_defaultdict[school_grade_stage_key]['word']["term"].append(1)
                            else:
                                school_grade_stage_defaultdict[school_grade_stage_key]['exercise']["term"].append(1)

                    for hermesday in hermes_day:
                        if school_grade_stage_key + "@" + str(hermesday['exercise_id']) not in unique_exercise:
                            if hermesday['exercise_id'] in word_day_id:
                                school_grade_stage_defaultdict[school_grade_stage_key]['word']["day"].append(1)
                            else:
                                school_grade_stage_defaultdict[school_grade_stage_key]['exercise']["day"].append(1)

                    [unique_exercise.add(school_grade_stage_key + "@" + str(e['exercise_id'])) for e in hermes_term]
                    [unique_exercise.add(school_grade_stage_key + "@" + str(e['exercise_id'])) for e in hermes_day]

                    # 阅读
                    readings_term = []
                    readings_day = []
                    if teachers:
                        # if school_id == 3720:
                        #     print(teachers)
                        reading_all_sql = """
                        select t.id,t.teacher_id,st.student_id,t.time_create ,count(st.id) as st_count 
                        from sigma_graded_reading_task as t
                        join  sigma_graded_reading_student_task as st 
                        on t.id = st.task_id
                        and t.available = 1
                        and t.time_create >='%s' and t.time_create<='%s'
                        and t.teacher_id in (%s) 
                        group by t.id having st_count >= 10;
                        """ % (term_start, term_end, ','.join([str(id['user_id']) for id in teachers]))
                        # print(reading_all_sql)
                        cursor.execute(reading_all_sql)
                        readings_term = cursor.fetchall()
                        # for item in reading_term:
                        #     print( str(item['time_create']))
                        readings_day = [item for item in readings_term if str(item['time_create'])[:-9] >= day30_start
                                        and str(item['time_create'])[:-9] <= day30_end]
                    for r_term in readings_term:
                        if school_grade_stage_key + "@" + str(r_term['id']) not in unique_read:
                            school_grade_stage_defaultdict[school_grade_stage_key]['read']["term"].append(1)

                    for r_day in readings_day:
                        if school_grade_stage_key + "@" + str(r_day['id']) not in unique_read:
                            school_grade_stage_defaultdict[school_grade_stage_key]['read']["day"].append(1)
                    [unique_read.add(school_grade_stage_key + "@" + str(e['id'])) for e in readings_term]
                    [unique_read.add(school_grade_stage_key + "@" + str(e['id'])) for e in readings_day]
                    # 听说
                    listen_term = []
                    listen_day = []
                    if teachers:
                        reading_all_sql = """
    
                                   select t.id,t.teacher_id,st.student_id,t.time_create ,count(st.id) as st_count 
                                   from sigma_listening_speaking_task as t
                                   join  sigma_listening_speaking_student_task as st 
                                   on t.id = st.task_id
                                   and t.available = 1
                                   and t.time_create >='%s' and t.time_create<='%s'
                                   and t.teacher_id in (%s) 
                                   group by t.id having st_count >= 10;
                                   """ % (term_start, term_end, ','.join([str(id['user_id']) for id in teachers]))

                        cursor.execute(reading_all_sql)
                        listen_term = cursor.fetchall()
                        # if school_id == 3720:
                        #     print(listen_term)
                        # print(school_id, group['id'], group['grade'], group['stage'], teachers)
                        listen_day = [item for item in listen_term if str(item['time_create'])[:-9] >= day30_start
                                      and str(item['time_create'])[:-9] <= day30_end]
                        # if school_id == 3720:
                        #     print(group_id, [item for item in listen_term if str(item['time_create'])[:-9] >= day30_start
                        #               and str(item['time_create'])[:-9] <= day30_end])

                    for l_term in listen_term:
                        if school_grade_stage_key + "@" + str(l_term['id']) not in unique_listen:
                            school_grade_stage_defaultdict[school_grade_stage_key]['listen']["term"].append(
                                1)

                    for l_day in listen_day:
                        if school_grade_stage_key + "@" + str(l_day['id']) not in unique_listen:
                            school_grade_stage_defaultdict[school_grade_stage_key]['listen']["day"].append(
                                1)
                    [unique_listen.add(school_grade_stage_key + "@" + str(e['id'])) for e in listen_term]
                    [unique_listen.add(school_grade_stage_key + "@" + str(e['id'])) for e in listen_day]



                    # 人数
                    bindings = {}
                    bindings['total_bind'] = 0
                    if users:
                        bind_sql = "select count(distinct user_id) as total_bind from sigma_account_re_userwechat where available = 1 and user_id in (%s)" % (
                            ",".join([str(item['user_id']) for item in users]))
                        # print(bind_sql)
                        cursor.execute(bind_sql)
                        bindings = cursor.fetchone()
                    school_grade_stage_defaultdict[school_grade_stage_key]['user']["total"].append(len(users))
                    totaluser = len(users)
                    bind_number = bindings["total_bind"] if bindings["total_bind"] else 0
                    school_grade_stage_defaultdict[school_grade_stage_key]['user']["unbind"].append(totaluser - bind_number)

                    # 付费
                    pays = []
                    if users:
                        pay_sql = "select id,user_id, coupon_amount from sigma_pay_ob_order where available = 1 and  coupon_amount >= 1 and status=3 and user_id in (%s)" % (
                            ",".join([str(item['user_id']) for item in users]))
                        cursor.execute(pay_sql)
                        pays = cursor.fetchall()
                        order_id = [item['id'] for item in pays]
                    school_grade_stage_defaultdict[school_grade_stage_key]['pay']["number"].append(
                        len(list({item['user_id'] for item in pays})))
                    school_grade_stage_defaultdict[school_grade_stage_key]['pay']["amount"].append(
                        sum([item['coupon_amount'] for item in pays]))
                    # 退款
                    refund = {}
                    refund['total_refund'] = 0
                    if order_id:
                        refund_sql = "select sum(amount) as total_refund from sigma_pay_ob_refund where status=6 and order_id in (%s)" % (
                            ",".join([str(id) for id in order_id]))
                        # print(refund_sql)
                        cursor.execute(refund_sql)
                        refund = cursor.fetchone()
                        if refund:
                            # print(refund, 'refund')
                            school_grade_stage_defaultdict[school_grade_stage_key]['refund']["amount"].append(
                                refund['total_refund'] if refund['total_refund'] else 0)
                        else:
                            school_grade_stage_defaultdict[school_grade_stage_key]['refund']["amount"].append(0)

            all_school_map = {}
            for school in all_school:
                all_school_map[int(school['id'])] = school

            # print(school_grade_stage_defaultdict)
            for school_grade_stage_key, data in school_grade_stage_defaultdict.items():
                s = school_grade_stage_key.split('@')
                school_id = int(s[0])
                grade = s[1]
                stage = s[2]

                bind_rate = (sum(data.get("user", {}).get("total", [0])) - sum(
                    data.get("user", {}).get("unbind", [0]))) / sum(
                    data.get("user", {}).get("total", [0])) if sum(data.get("user", {}).get("total", [0])) else 0

                # if school_id == 3720:
                #     print(school_grade_stage_key, data.get("listen", {}).get("day", [0]), unique_listen)
                ws.append([
                    channels_map.get(all_school_map.get(school_id, {}).get("owner_id", ""), {}).get("name"),
                    all_school_map.get(school_id, {}).get("full_name", ""),
                    grade,
                    stage,
                    sorted(data['date']['date'])[0],
                    (datetime.datetime.now() - sorted(data['date']['date'])[0]).days,

                    len(data.get("word", {}).get("term", [])) + len(data.get("exercise", {}).get("term", [])) + sum(
                        data.get("read", {}).get("term", [0])) + sum(data.get("listen", {}).get("term", [0])),

                    len(data.get("word", {}).get("day", [])) + len(data.get("exercise", {}).get("day", [])) + sum(
                        data.get("read", {}).get("day", [0])) + sum(data.get("listen", {}).get("day", [0])),

                    len(data.get("exercise", {}).get("day", [])),
                    sum(data.get("read", {}).get("day", [0])),
                    len(data.get("word", {}).get("day", [])),
                    sum(data.get("listen", {}).get("day", [0])),
                    sum(data.get("user", {}).get("total", [0])),
                    sum(data.get("user", {}).get("unbind", [0])),
                    bind_rate if bind_rate <= 1 else 1,
                    sum(data.get("pay", {}).get("amount", [0])),
                    sum(data.get("pay", {}).get("number", [0])),
                    sum(data.get("refund", {}).get("amount", [0])),

                ])
            # print(11111)
            wb.save("./all_channels_process/%s.xlsx" % channel['name'])

        """
    
        "渠道","学校", "年级", "学段", "年级创建日期", "试用天数",
                   "本学期使用次数", "30天使用次数","30天智能批改", "30天分级阅读","30天词汇","30天听说",
                   "在册学生数","未绑定人数","绑定率",,"付费金额","付费人数","退款"
        """
        cursor.close()
        connection.close()
        server_mongo.stop()
        server.stop()
    except:
        import traceback
        traceback.print_exc()


from multiprocessing import Pool
import os, time, random



if __name__ == '__main__':
    print('Parent process %s.' % os.getpid())
    total = 130
    n = 20
    howmany = total/n
    p = Pool(n)  # 创建4个进程
    for i in range(total//n + 1):
        print(i)
        p.apply_async(long_task, args=(n*i,n*(i+1)))
    print('Waiting for all subprocesses done...')

    p.close()
    p.join()

    # print(130//20)

