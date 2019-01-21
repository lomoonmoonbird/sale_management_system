import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook




server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)


school_sql = "select id, owner_id, full_name from sigma_account_ob_school where available = 1 and owner_id=596004"
cursor = connection.cursor()
cursor.execute(school_sql)
schools = cursor.fetchall()

school_map = {}
for school in schools:
    school_map[school['id']] = school

school_ids = [item['id'] for item in schools]
group_sql = "select id, school_id, name from sigma_account_ob_group where available = 1 and school_id in (%s)" % (",".join([str(id) for id in school_ids]))
cursor.execute(group_sql)
groups = cursor.fetchall()

school_group_map = {}
for group in groups:
    print(group)
    school_group_map[str(group['school_id']) + '@' + str(group['id'])] = group


items = mongo.class_per_day.aggregate([

    {
        "$match": {"channel": 596004}
    },
    {
        "$project": {
            "channel": 1,
            "school_id": 1,
            "group_id": 1,
            "day": 1,
            "student_number": 1,
            "valid_exercise_count": 1,
            "valid_word_count": 1,
            "valid_reading_count": 1,
            "guardian_unique_count": 1,
            "pay_number": 1,

        }
    },
    {
        "$group":{
            "_id": {"school_id": "$school_id", "group_id": "$group_id"},
            "total_student_number": {"$sum": "$student_number"},
            "total_valid_exercise_count": {"$sum": "$valid_exercise_count"},
            "total_valid_word_count": {"$sum": "$valid_word_count"},
            "total_valid_reading_count": {"$sum": "$valid_reading_count"},
            "total_guardian_unique_count": {"$sum": "$guardian_unique_count"},
            "total_pay_number": {"$sum": "$pay_number"},
        }
    },
    {
        "$project": {
            "_id": 1,
            "total_student_number": 1,
            "total_valid_exercise_count": 1,
            "total_valid_word_count": 1,
            "total_valid_reading_count": 1,
            "total_guardian_unique_count": 1,
            'total_pay_number': 1,
            "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_guardian_unique_count", "$total_student_number"]}]},
        }
    }

])

items_map = {}
for item in items:
    items_map[str(item['_id']['school_id']) + "@" + str(item['_id']['group_id'])] = item

# data = list(items)

# for school_group_id, data in school_group_map.items():


wb = Workbook("./certain_channel_school_group.xlsx")
# print(content)
# for school_group_id, data in school_group_map.items():
wb.create_sheet(title="sheet1")

ws = wb['sheet1']
# print(ws)
ws.append(['学校', '班级', '班级人数','班级考试数','班级单词数','班级阅读数','班级家长数','班级绑定率','班级付费人数'])
for school_group_id, data in school_group_map.items():

    ws.append([school_map.get(int(school_group_id.split('@')[0]), {}).get("full_name"),
              school_group_map.get(school_group_id, {}).get("name"),
               items_map.get(school_group_id, {}).get("total_student_number", 0),
               items_map.get(school_group_id, {}).get("total_valid_exercise_count", 0),
               items_map.get(school_group_id, {}).get("total_valid_word_count", 0),
               items_map.get(school_group_id, {}).get("total_valid_reading_count", 0),
               items_map.get(school_group_id, {}).get("total_guardian_unique_count", 0),
               items_map.get(school_group_id, {}).get("bind_ratio", 0),
               items_map.get(school_group_id, {}).get("total_pay_number", 0)])
    # ws.cell(row=row, column=1).value = subcontent[row]['stage']
    # ws.cell(row=row, column=2).value = subcontent[row]['class']
    # ws.cell(row=row, column=3).value = subcontent[row]['grade']
wb.save("./certain_channel_school_group.xlsx")

server_mongo.stop()
server.stop()
cursor.close()
connection.close()