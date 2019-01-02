import pymongo
import re
from collections import defaultdict
import json

from celery import Task
from enumconstant import Roles
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

LEVEL = {"primary": "小学", "junior":"初中", "senior": "高中"}

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

instance_db = mongo.instance


def school_new_delta(mongo, channel_ids=[], begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                           end_time=datetime.datetime.now().strftime("%Y-%m-%d")):
    """
    学校新增数据
    :param request:
    :param channel_ids:
    :param begin_time:
    :param end_time:
    :return:
    """
    items = []
    if not begin_time:
        begin_time = begin_time
    if not end_time:
        end_time = end_time

    coll = mongo.school_per_day_coll

    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids},
                    # "day": {"$gte": begin_time, "$lte": end_time}
                }
            },
            {
                "$project": {
                    "channel": 1,
                    "school_id": 1,
                    "day": 1,
                    "student_number": 1,
                    "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                    "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                    "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                    "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                    "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                    "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},
                    "guardian_unique_count_range": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$guardian_unique_count", 0]},
                    "guardian_unique_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_unique_count", 0]},
                    "pay_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_number", 0]},
                    "pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]}
                }
            },

            {"$group": {"_id": {"channel": "$channel", "school_id": "$school_id"},
                        "total_school": {"$sum": "$school_number"},
                        "total_exercise": {"$sum": "$valid_exercise_count"},
                        "total_reading": {"$sum": "$valid_reading_count"},
                        "total_word": {"$sum": "$valid_word_count"},
                        "total_images": {"$sum": {"$add": ["$e_image_c", "$w_image_c"]}},
                        "total_unique_guardian_range": {"$sum": "$guardian_unique_count_range"},
                        "total_unique_guardian": {"$sum": "$guardian_unique_count"},
                        "total_pay_number": {"$sum": "$pay_number"},
                        "total_pay_amount": {"$sum": "$pay_amount"},
                        "total_student": {'$sum': "$student_number"},
                        "e_image_c": {"$sum": "$e_image_c"},
                        "w_image_c": {"$sum": "$w_image_c"}
                        }
             },
            {
                "$project": {
                    "total_school": 1,
                    "total_exercise": 1,
                    "total_reading": 1,
                    "total_word": 1,
                    "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},
                    "total_unique_guardian": 1,
                    "total_pay_number": 1,
                    "total_pay_amount": 1,
                    "bind_ratio": {"$cond": [{"$eq": ["$total_student", 0]}, 0,
                                             {"$divide": ["$total_unique_guardian",
                                                          "$total_student"]}]},
                }

            }

        ])

    for item in item_count:
        items.append(item)

    return items

def channel_new_delta( mongo, channel_ids=[], begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                            end_time=datetime.datetime.now().strftime("%Y-%m-%d")):
    """

    :param request:
    :param channel_ids:
    :return:
    """
    items = []
    if not begin_time:
        begin_time = begin_time
    if not end_time:
        end_time = end_time

    coll = mongo.channel_per_day

    item_count = coll.aggregate(
        [
            {
                "$match": {
                    "channel": {"$in": channel_ids},
                    # "day": {"$gte": begin_time, "$lte": end_time}
                }
            },
            {
                "$project": {
                    "channel": 1,
                    "day": 1,
                    "student_number": 1,
                    "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                    "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                    "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                    "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                    "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                    "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},
                    "guardian_unique_count_range": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$guardian_unique_count", 0]},
                    "guardian_unique_count": {
                        "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_unique_count", 0]},
                    "pay_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_number", 0]},
                    "pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                        "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]}
                }
            },

            {"$group": {"_id": "$channel",
                        "total_school": {"$sum": "$school_number"},
                        "total_exercise": {"$sum": "$valid_exercise_count"},
                        "total_reading": {"$sum": "$valid_reading_count"},
                        "total_word": {"$sum": "$valid_word_count"},
                        # "total_images": {"$sum": {"$add": ["$e_image_c", "$w_image_c"]}}, #这种用法是错误的，如果某一条记录其中一个字段不存在所有的都会返回0
                        "total_unique_guardian_range": {"$sum": "$guardian_unique_count_range"},
                        "total_unique_guardian": {"$sum": "$guardian_unique_count"},
                        "total_pay_number": {"$sum": "$pay_number"},
                        "total_pay_amount": {"$sum": "$pay_amount"},
                        "total_student": {'$sum': "$student_number"},
                        "e_image_c": {"$sum": "$e_image_c"},
                        "w_image_c": {"$sum": "$w_image_c"}
                        }
             },
            {
                "$project": {
                    "total_school": 1,
                    "total_exercise": 1,
                    "total_reading": 1,
                    "total_word": 1,
                    "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},
                    "total_unique_guardian": 1,
                    "total_pay_number": 1,
                    "total_pay_amount": 1,
                    "bind_ratio": {"$cond": [{"$eq": ["$total_student", 0]}, 0,
                                             {"$divide": ["$total_unique_guardian",
                                                          "$total_student"]}]},
                }

            }

        ])

    for item in item_count:
        items.append(item)

    return items


def sheet( items=[], area_map=None, channel_map=None, school_items=[], channel_mysql_map={}, school_mysql_map={}, area_name_extra='',begin='',end=''):
    """
    生成excel
    :param items:
    :param area_map:
    :param channel_map:
    :return:
    """

    wb = Workbook('./channel_sheet/'+area_name_extra+ '|'+ begin + "--" + end+".xlsx")
    default = {
        "total_school": 0,
        "total_exercise": 0,
        "total_reading": 0,
        "total_word": 0,
        "total_images": 0,
        "total_unique_guardian": 0,
        "total_pay_number": 0,
        "total_pay_amount": 0,
        "bind_ratio": 0
    }
    sheet_name = "sheet"
    if items:
        sheet_name = "渠道新增数据"
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]


        item_map = {}
        for item in items:
            item_map[item['_id']] = item


        for channel_id, data in channel_map.items():
            if not item_map.get(channel_id):
                default['_id'] = channel_id
                items.append(default)
        ws.append(['大区名称', '渠道名称', '新增和学校数量', '新增测试使用次数', '新增阅读使用次数',
                   '新增单词使用次数', '新增图片数量', '新增绑定人数数量', '新增付费人数', '新增付费金额', '绑定率'])
        for row in range(0, len(items)):
            ws.append([channel_map.get(items[row]['_id']).get("area_name"), channel_map.get(items[row]['_id']).get("name"),
                       items[row]['total_school'], items[row]['total_exercise'],
                       items[row]['total_reading'], items[row]['total_word'],
                       items[row]['total_images'], items[row]['total_unique_guardian'],
                       items[row]['total_pay_number'], items[row]['total_pay_amount'], items[row]['bind_ratio']])
    if school_items:
        sheet_name = "学校新增数据"
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]


        item_map = {}
        for item in school_items:
            item_map[item['_id']['school_id']] = item

        for school_id, data in school_mysql_map.items():
            if not item_map.get(school_id):
                default["_id"] = {}
                default['_id']['channel'] = data['owner_id']
                default['_id']['school_id'] = data['id']
                school_items.append(default)
        ws.append(['渠道', '学校名称', '新增测试使用次数', '新增阅读使用次数',
                   '新增单词使用次数', '新增图片数量', '新增绑定人数数量', '新增付费人数', '新增付费金额', '绑定率'])

        for row in range(0, len(school_items)):
            print(channel_mysql_map)
            print(school_items[row]['_id']['channel'], channel_mysql_map.get(school_items[row]['_id']['channel']), "channel_mysql_map.get(school_items[row]['_id']['channel'])")
            ws.append([channel_mysql_map.get(school_items[row]['_id']['channel']).get("name"),
                       school_mysql_map.get(school_items[row]['_id']['school_id']).get("full_name"),
                       school_items[row]['total_exercise'],
                       school_items[row]['total_reading'], school_items[row]['total_word'],
                       school_items[row]['total_images'], school_items[row]['total_unique_guardian'],
                       school_items[row]['total_pay_number'], school_items[row]['total_pay_amount'],
                       school_items[row]['bind_ratio']])

    wb.save('./channel_sheet/'+area_name_extra+ '|'+ begin + "--" + end+".xlsx")




# daterange = [("2018-01-01", "2018-01-31"), ("2018-02-01", "2018-02-28"), ("2018-03-01", "2018-03-31"), ("2018-04-01", "2018-04-30"),
#              ("2018-05-01", "2018-05-31"), ("2018-06-01", "2018-06-30"), ("2018-07-01", "2018-07-31"),("2018-08-01","2018-08-31"),
#              ("2018-09-01", "2018-09-30"),("2018-10-01", "2018-10-31"),("2018-11-01", "2018-11-30"),("2018-12-01","2018-12-31")]

daterange = [("2019-01-01", "2019-01-01")]
area_names = ["华东大区","华南大区", "华北大区","西部大区","华中大区","西南大区","东南大区","东北大区"]
channel_names = ''

for area_name in area_names:
    for d in daterange:
        area_name = area_name
        begin_time = d[0]
        end_time = d[1]
        channel_name = ''
        #渠道 begin
        query = { "role": Roles.AREA.value, "status": 1}
        if area_name:
            query.update({"name": {"$in": area_name.split(',')}})

        areas = mongo.instance.find(query)
        areas= list(areas)
        area_ids = [str(item['_id']) for item in areas]
        # print(areas)
        channels = mongo.instance.find({"parent_id": {"$in": area_ids},
                                                                          "role": Roles.CHANNEL.value,
                                                                      "status": 1})
        channels = list(channels)
        # print(channels)
        old_ids = [item['old_id'] for item in channels]


        area_map = {}
        for area in areas:
            area_map[str(area['_id'])] = area
        print('area_map', area_map)
        sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
              ','.join([str(id) for id in old_ids])

        cursor = connection.cursor()
        cursor.execute(sql)
        real_channels = cursor.fetchall()


        channels_map = {}
        for channel in channels:
            channels_map[channel['old_id']] = channel

        print("channels_map", channels_map)
        channel_mongo_map = {}
        for channel in real_channels:
            channel.update({"name": channel.get("name", ""),
                            "area_id": channels_map.get(channel['id'],{}).get("parent_id", ""),
                            "area_name": area_map.get(channels_map.get(channel['id'],{}).get("parent_id", ""), {}).get("name","")})
            channel_mongo_map[channel['id']] = channel

        items = channel_new_delta(mongo, old_ids, begin_time, end_time)
        # print(json.dumps(items, indent=4))
        #渠道 end
        #学校 begin
        sql = "select id, name from sigma_account_us_user where available = 1 and name in (%s) " % \
              ','.join(["'"+str(username)+"'" for username in channel_name.split(",")])

        cursor = connection.cursor()
        cursor.execute(sql)
        real_channels = cursor.fetchall()


        channel_ids = [item['id'] for item in real_channels]

        school_items= []
        channel_mysql_map = {}
        school_mysql_map = {}

        print("channel_ids", channel_ids)
        if channel_ids:
            school_sql = "select id, full_name, owner_id " \
                         "from sigma_account_ob_school " \
                         "where available = 1 " \
                         "and owner_id in (%s)" % (','.join([str(id) for id in channel_ids]))
            cursor = connection.cursor()
            cursor.execute(school_sql)
            real_schools = cursor.fetchall()




            for channel in real_channels:
                channel_mysql_map[channel['id']] = channel

            for school in real_schools:
                school_mysql_map[school['id']] = school
            school_items = school_new_delta(mongo, old_ids, begin_time, end_time)


        sheet(items,area_map,channel_mongo_map, school_items,channel_mysql_map,school_mysql_map, area_name,begin_time, end_time)


server_mongo.stop()
server.stop()
connection.close()






