import pymongo
import re
from collections import defaultdict
import json
from celery import Task
from models.mysql.centauri import ob_school, us_user, re_userwechat, ob_group, \
ob_groupuser, ob_exercise, as_hermes, ob_order, re_userwechat, Roles, StageEnum, \
    StudentRelationEnum, ExerciseTypeEnum, ob_exercisemeta, st_location, ob_reading
import pymysql
import pymongo
from pymongo import InsertOne, DeleteMany, ReplaceOne, UpdateOne
from sqlalchemy import select, func, asc, distinct, text, desc
from sqlalchemy.sql import and_, or_, not_
from sqlalchemy.dialects import mysql
from configs import MONGODB_CONN_URL
from loggings import logger
import pickle
import time
import datetime
from datetime import timedelta, date
from collections import defaultdict
from sshtunnel import SSHTunnelForwarder
from tasks.celery_base import BaseTask, CustomEncoder
from pymongo.errors import BulkWriteError
from celery.signals import worker_process_init,worker_process_shutdown, beat_init
from configs import DEBUG, MONGODB_CONN_URL, MYSQL_NAME, MYSQL_USER, MYSQL_PASSWORD, MYSQL_HOST, MYSQL_PORT
from openpyxl import load_workbook, Workbook

server = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('rr-uf6247jo85269bp6e.mysql.rds.aliyuncs.com', 3306))
server.start()

server_mongo = SSHTunnelForwarder(
    ssh_address_or_host=('139.196.77.128', 5318),  # 跳板机

    ssh_password="PengKim@89527",
    ssh_username="jinpeng",
    remote_bind_address=('dds-uf6fcc4e461ee5a41.mongodb.rds.aliyuncs.com', 3717))
server_mongo.start()

mongo = pymongo.MongoClient('127.0.0.1',
                            username='root',
                            password = 'sigmaLOVE2017',
                            authMechanism='SCRAM-SHA-1',
                            port=server_mongo.local_bind_port).sales
connection = pymysql.connect(host="127.0.0.1",
                                  port=server.local_bind_port,
                                  user="sigma",
                                  password="sigmaLOVE2017",
                                  db=MYSQL_NAME,
                                  charset='utf8mb4',
                                  cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()
group_names = [1801,1802,1803,1805,1806,1807,1808,1809,1810,1811,1812,1813,1814,1815,1701,1702,1703,1704,1705,1706,1707,1708,1709,1710,1711,1712,1713,1714,1601,1602,1603,1604,1605,1606,1607,1608,1609,1610,1611,1612,1613,1614]

group_sql =  "select * from sigma_account_ob_group where available = 1 and school_id = 3720 "
cursor.execute(group_sql)
groups = cursor.fetchall()


group_ids = [item['id'] for item in groups]
usergroup_sql = "select * from sigma_account_re_groupuser where available = 1 and role_id = 2 and group_id in (%s)" % (','.join([str(id) for id in group_ids]))
cursor.execute(usergroup_sql)
usergroup = cursor.fetchall()


user_ids = [item['user_id'] for item in usergroup]
wechat_sql = "select distinct(user_id) from sigma_account_re_userwechat where available =1  and user_id in (%s)" %(",".join([str(id) for id in user_ids]))
cursor.execute(wechat_sql)
wechat = cursor.fetchall()
wechat_userid = [item['user_id'] for item in wechat]

wechat_range_sql = "select distinct(user_id) from sigma_account_re_userwechat where available =1 and time_create>='2019-01-23'  and user_id in (%s)" %(",".join([str(id) for id in user_ids]))
cursor.execute(wechat_range_sql)
wechat_range = cursor.fetchall()
wechat_userid_range = [item['user_id'] for item in wechat]

teachers_sql = "select * from sigma_account_re_groupuser where available = 1 and role_id = 1 and group_id in (select id from sigma_account_ob_group where available =1 and school_id = 3720)"
cursor.execute(teachers_sql)
teachers = cursor.fetchall()

teachers_id = [item['user_id'] for item in teachers]

reading_sql = "select * from sigma_graded_reading_task where available =1 and time_create >='2019-01-21' and time_create <='2019-01-27' and teacher_id in (%s)" % (",".join([str(id) for id in teachers_id]))
cursor.execute(reading_sql)
readings = cursor.fetchall()

reading_map = defaultdict(list)
for r in readings:
    reading_map[r['teacher_id']].append(r)


teacher_map = defaultdict(list)
for t in teachers:
    teacher_map[t['group_id']].append(len(reading_map[t['user_id']]))


user_group_map = defaultdict(list)
user_group = {}
user_bind = defaultdict(list)
user_range_bind = defaultdict(list)
for ug in usergroup:
    user_group_map[ug['group_id']].append(ug)
    user_group[ug['user_id']] = ug['group_id']
    if ug['user_id'] in wechat_userid:
        user_bind[ug['group_id']].append(1)
    if ug['user_id'] in wechat_userid_range:
        user_range_bind[ug['group_id']].append(1)



wb = Workbook('./reading.xlsx')
sheet_name = "班级数据"
wb.create_sheet(title=sheet_name)
ws = wb[sheet_name]

ws.append(['班级', '学生总数', '微信绑定率', '新增绑定人数', '阅读任务数'])
for g in groups:
    ws.append([
        g['name'],
        len(user_group_map[g['id']]),
        len(user_bind[g['id']]) / len(user_group_map[g['id']]) if len(user_group_map[g['id']]) else 0,
        len(user_range_bind[g['id']]),
        sum(teacher_map.get(g['id'], [0]))
    ])
wb.save('./reading.xlsx')
print(groups)
server_mongo.stop()
server.stop()
cursor.close()
connection.close()