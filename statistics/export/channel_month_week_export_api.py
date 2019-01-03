#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 导出全局 大区 渠道月报表格
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin

class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime)):
            return str(obj)
        if isinstance(obj, (ObjectId)):
            return str(obj)

        return json.JSONEncoder.default(self, obj)

class ChannelExportReport(BaseHandler, ExportBase, DataExcludeMixin):
    def __init__(self):
        self.db = 'sales'
        self.user_coll = 'sale_user'
        self.instance_coll = 'instance'
        self.class_per_day_coll = 'class_per_day'
        self.grade_per_day_coll = 'grade_per_day'
        self.channel_per_day_coll = 'channel_per_day'
        self.school_per_day_coll = 'school_per_day'
        self.thread_pool = ThreadPoolExecutor(20)


    @validate_permission(data_validation=True)
    async def month(self, request: Request):
        """
        渠道导出月报
        {
            "channel_id": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        exclude_schools_u = request['data_permission']['exclude_school']
        channel_id = request_param.get("channel_id", "")
        channel_old_id = -1
        channel_info = {}
        if request['user_info']['instance_role_id'] == Roles.CHANNEL.value:
            channel_id = request['user_info']['channel_id']
            channel_info = await request.app['mongodb'][self.db][self.instance_coll].find_one({"_id": ObjectId(channel_id),
                                                                                "role": Roles.CHANNEL.value,
                                                                                "status": 1}, {"old_id": 1, "_id": 0})

        else:
            channel_info = await request.app['mongodb'][self.db][self.instance_coll].find_one({"old_id": int(channel_id),
                                                                                                     "role": Roles.CHANNEL.value,
                                                                                                     "status": 1})

        if not channel_info:
            raise RequestError("channel doesn't exist")
        channel_old_id = channel_info.get("old_id", "")
        # schools = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": channel_id,
        #                                                                     "role": Roles.SCHOOL.value,"status": 1})
        # schools = await schools.to_list(10000)
        # schools_ids = [item['school_id'] for item in schools]

        exclude_schools = await self.exclude_schools(request.app['mysql'])
        if not exclude_schools:
            sql = "select id,full_name from sigma_account_ob_school " \
                  "where available = 1 " \
                  "and owner_id = %s "  % (channel_old_id)
        else:
            sql = "select id,full_name from sigma_account_ob_school " \
                  "where available = 1 " \
                  "and owner_id = %s " \
                  "and id not in (%s)" % (channel_old_id, ','.join(str(id) for id in exclude_schools))

        channel_one_sql = "select id,name from sigma_account_us_user where id = %s" % channel_old_id

        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(sql)
                schools = await cur.fetchall()
                await cur.execute(channel_one_sql)
                channel = await cur.fetchone()


        schools_ids = [item['id'] for item in schools]
        items = await self._list_month(request, schools_ids)
        channel_items = await self._list_channel_month(request, [channel_old_id])
        template_path = os.path.dirname(__file__) + "/templates/channel_month_template.xlsx"

        # print(json.dumps(items,indent=4))
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.sheet,
                                                       template_path,
                                                       items,
                                                       channel_items,
                                                       schools,
                                                       channel,
                                                       "month")

        return await self.replay_stream(sheet, "渠道月报-" + datetime.now().strftime("%Y-%m-%d"), request)

    @validate_permission(data_validation=True)
    async def week(self, request: Request):
        """
        渠道导出周报
        :param request:
        :return:
        """

        request_param = await get_params(request)
        exclude_schools_u = request['data_permission']['exclude_school']
        channel_id = request_param.get("channel_id", "")
        channel_old_id = -1
        channel_info = {}
        if request['user_info']['instance_role_id'] == Roles.CHANNEL.value:
            channel_id = request['user_info']['channel_id']
            channel_info = await request.app['mongodb'][self.db][self.instance_coll].find_one({"_id": ObjectId(channel_id),
                                                                                               "role": Roles.CHANNEL.value,
                                                                                               "status": 1},
                                                                                              {"old_id": 1, "_id": 0})

        else:
            channel_info = await request.app['mongodb'][self.db][self.instance_coll].find_one(
                {"old_id": int(channel_id),
                 "role": Roles.CHANNEL.value,
                 "status": 1})

        if not channel_info:
            raise RequestError("channel doesn't exist")
        channel_old_id = channel_info.get("old_id", "")


        # schools = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": channel_id,
        #                                                                     "role": Roles.SCHOOL.value,"status": 1})
        # schools = await schools.to_list(10000)
        # schools_ids = [item['school_id'] for item in schools]

        exclude_schools = await self.exclude_schools(request.app['mysql'])
        if not exclude_schools:
            sql = "select id,full_name from sigma_account_ob_school " \
                  "where available = 1 " \
                  "and owner_id = %s " % (channel_old_id)
        else:
            sql = "select id,full_name from sigma_account_ob_school " \
                  "where available = 1 " \
                  "and owner_id = %s " \
                  "and id not in (%s)" % (channel_old_id, ','.join(str(id) for id in exclude_schools))

        channel_one_sql = "select id,name from sigma_account_us_user where id = %s" % channel_old_id

        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(sql)
                schools = await cur.fetchall()
                await cur.execute(channel_one_sql)
                channel = await cur.fetchone()

        schools_ids = [item['id'] for item in schools]
        items = await self._list_week(request, schools_ids)
        channel_items = await self._list_channel_week(request, [channel_old_id])
        template_path = os.path.dirname(__file__) + "/templates/channel_week_template.xlsx"

        # print(json.dumps(items,indent=4))
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.sheet,
                                                       template_path,
                                                       items,
                                                       channel_items,
                                                       schools,
                                                       channel,
                                                       "month")

        return await self.replay_stream(sheet, "渠道周报-" + datetime.now().strftime("%Y-%m-%d"), request)


    @validate_permission(data_validation=True)
    async def month_school(self, request: Request):
        """
        渠道学校月报
        :param request:
        :return:
        """


    async def _list_month(self, request: Request, school_ids: list):
        """
        月报
        :param request:
        :param channel_ids:
        :return:
        """

        coll = request.app['mongodb'][self.db][self.school_per_day_coll]
        items = []

        last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day,\
        curr_month_first_day, curr_month_last_day = self._curr_and_last_and_last_last_month()

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "school_id": {"$in": school_ids}
                    }
                },
                {
                    "$project": {
                        "school_id": 1,
                        "channel": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "total_pay_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_number", 0]},
                        "total_pay_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_amount", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_count", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_count", 0]},

                        "guardian_unique_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_unique_count", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$school_id",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$total_pay_number"},
                            "total_pay_amount": {"$sum": "$total_pay_amount"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                            "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_unique_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "guardian_unique_number_curr_month": 1,
                        "guardian_unique_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,
                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                 {"$divide": ["$total_guardian_number", "$total_student_number"]}]},
                        # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _list_week(self, request: Request, school_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.school_per_day_coll]
        items = []
        last_week = self.last_week_from_7_to_6()
        last_last_week = self.last_last_week_from_7_to_6()
        last_week_first_day, last_week_last_day, \
        last_last_week_first_day, last_last_week_last_day =  last_week[0], last_week[6],\
                                                             last_last_week[0], last_last_week[6]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "school_id": {"$in": school_ids}
                    }
                },
                {
                    "$project": {
                        "school_id": 1,
                        "channel": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "total_pay_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_number", 0]},
                        "total_pay_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_amount", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$guardian_count", 0]},
                        "guardian_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_count", 0]},

                        "guardian_unique_number_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_unique_count", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$school_id",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$total_pay_number"},
                            "total_pay_amount": {"$sum": "$total_pay_amount"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                            "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_unique_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "guardian_unique_number_curr_month": 1,
                        "guardian_unique_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,
                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                 {"$divide": ["$total_guardian_number", "$total_student_number"]}]},
                        # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _list_channel_month(self, request: Request, channel_ids=[]):
        """
        渠道月统计
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []

        last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day, \
        curr_month_first_day, curr_month_last_day = self._curr_and_last_and_last_last_month()

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "school_id": 1,
                        "channel": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "total_pay_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_number", 0]},
                        "total_pay_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_amount", 0]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_count", 0]},
                        "guardian_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_count", 0]},

                        "guardian_unique_number_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_unique_count", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$school_id",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$total_pay_number"},
                            "total_pay_amount": {"$sum": "$total_pay_amount"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                            "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"},
                            "total_images": {"$sum": "$total_images"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_school_number": 1,
                        "total_unique_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "guardian_unique_number_curr_month": 1,
                        "guardian_unique_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,
                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                 {"$divide": ["$total_guardian_number", "$total_student_number"]}]},
                        # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _list_channel_week(self, request: Request, channel_ids=[]):
        """
        渠道月统计
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []

        last_week = self.last_week_from_7_to_6()
        last_last_week = self.last_last_week_from_7_to_6()
        last_week_first_day, last_week_last_day, \
        last_last_week_first_day, last_last_week_last_day = last_week[0], last_week[6], \
                                                            last_last_week[0], last_last_week[6]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "school_id": 1,
                        "channel": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "total_pay_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_number", 0]},
                        "total_pay_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_amount", 0]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$guardian_count", 0]},
                        "guardian_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_count", 0]},

                        "guardian_unique_number_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_unique_count", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$school_id",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$total_pay_number"},
                            "total_pay_amount": {"$sum": "$total_pay_amount"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                            "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"},
                            "total_images": {"$sum": "$total_images"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_school_number": 1,
                        "total_unique_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "guardian_unique_number_curr_month": 1,
                        "guardian_unique_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,
                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                 {"$divide": ["$total_guardian_number", "$total_student_number"]}]},
                        # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items


    def sheet(self, template, items, channel_items,  schools, channel, report_type):
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]

        area_dimesion_items = {}

        start_point = 4
        if channel_items:
            row1 = sheet[1]
            if report_type == 'week':
                last_week = self.last_week()
                row1[0].value = "渠道_" + last_week[0] + "-" + last_week[6] + "周报数据"
            elif report_type == 'month':
                _, _, last_month, _, _, _ = self._curr_and_last_and_last_last_month()
                month = datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
                row1[0].value = "渠道" + str(month) + "月报数据"
            for one in list(sheet[2:3]):
                for cell in one:
                    cell.font = self._white_font()
                    cell.fill = self._background_header_color()
                    cell.border = self._border()
            summary_map = defaultdict(list)
            for index, channel_item in enumerate(channel_items):
                row = sheet[index + 4]
                index += 1
                for cell in row:
                    cell.font = self._black_font()
                    cell.alignment = self._alignment()
                    cell.border = self._border()
                    #地市 渠道名
                    row[0].value = channel.get("name", "")
                    #学校
                    row[1].value = channel_item.get("total_school_number", 0)
                    row[2].value = str(channel_item.get("school_number_last_month", 0)) + "/" + str(channel_item.get("school_number_curr_month", 0))
                    summary_map[1].append(row[1].value)
                    summary_map[2].append(row[2].value)
                    #教师
                    row[3].value = str(channel_item.get("teacher_number_last_month", 0)) + "/" + str(channel_item.get("teacher_number_curr_month", 0))
                    summary_map[3].append(row[3].value)
                    #学生
                    row[4].value = str(channel_item.get("student_number_last_month", 0)) + "/" + str(
                        channel_item.get("student_number_curr_month", 0))
                    summary_map[3].append(row[3].value)
                    #考试
                    row[5].value = channel_item.get("valid_exercise_count", 0)
                    row[6].value = str(channel_item.get("valid_exercise_count_last_month", 0)) + "/" + str(
                        channel_item.get("valid_exercise_count_curr_month", 0))
                    summary_map[5].append(row[5].value)
                    summary_map[6].append(row[6].value)
                    #单词
                    row[7].value = channel_item.get("valid_word_count", 0)
                    row[8].value = str(channel_item.get("valid_word_count_last_month", 0)) + "/" + str(
                        channel_item.get("valid_word_count_curr_month", 0))
                    summary_map[7].append(row[7].value)
                    summary_map[8].append(row[8].value)
                    #阅读
                    row[9].value = channel_item.get("valid_reading_count", 0)
                    row[10].value = str(channel_item.get("valid_reading_count_last_month", 0)) + "/" + str(
                        channel_item.get("valid_reading_count_curr_month", 0))
                    summary_map[9].append(row[9].value)
                    summary_map[10].append(row[10].value)
                    #图像
                    row[11].value = channel_item.get("total_images", 0)
                    row[12].value = str(channel_item.get("e_image_c_last_month", 0)+channel_item.get("w_image_c_last_month", 0)) + "/" + str(
                        channel_item.get("e_image_c_curr_month", 0)+channel_item.get("w_image_c_curr_month", 0))
                    summary_map[11].append(row[11].value)
                    summary_map[12].append(row[12].value)
                    #绑定
                    row[13].value = channel_item.get("total_unique_guardian_number", 0)
                    row[14].value = self.percentage(channel_item.get("total_unique_guardian_number", 0)
                                                    / channel_item.get("total_student_number", 0)
                                                    if channel_item.get("total_student_number", 0) else 0)
                    row[15].value = str(channel_item.get("guardian_unique_number_last_month", 0)) + "/" + str(
                        channel_item.get("guardian_unique_number_curr_month", 0))
                    summary_map[13].append(row[13].value)
                    summary_map[14].append(row[14].value)
                    summary_map[15].append(row[15].value)
                    #付费
                    row[16].value = channel_item.get("total_pay_amount", 0)
                    row[17].value = str(channel_item.get("pay_amount_last_month", 0)) + "/" + str(
                        channel_item.get("pay_amount_curr_month", 0))
                    summary_map[16].append(row[16].value)
                    summary_map[17].append(row[17].value)
                offset = start_point + len(channel_items)
                for index, cell in enumerate(sheet[offset]):
                    cell.font = self._black_font()
                    cell.font = self._black_font()
                    cell.alignment = self._alignment()
                    cell.border = self._border()
                    if index == 0:
                        cell.value = "环比新增率"
                        continue
                    cell.alignment = self._alignment()
                    if index in (1, 5, 7, 9, 11, 13, 14, 16):  # 不带/last_summary
                        cell.value = "/"
                    elif index in (2, 3, 4, 6, 8, 10, 12, 15, 17):  # 带/
                        last_summary = sum(
                            [float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
                        curr_summary = sum(
                            [float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
                        cell.value = self.percentage(
                            (curr_summary - last_summary) / last_summary if last_summary else 0)
                    else:
                        pass
                    if isinstance(cell.value, (int, float)):
                        if cell.value == 0:
                            cell.font = self._red_font()
                    if isinstance(cell.value, (str)):
                        if ("/" in cell.value and cell.value.split('/')[0] == "0") or \
                                ("/" in cell.value and cell.value.split('/')[1] == "0") or \
                                ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                            cell.font = self._red_font()
                start_point += offset + 1 + 1

        if schools:
            default = {
                "_id": '',
                "valid_reading_count": 0,
                "valid_exercise_count": 0,
                "valid_word_count": 0,
                "e_image_c": 0,
                "w_image_c": 0,
                "total_teacher_number": 0,
                "total_student_number": 0,
                "total_guardian_number": 0,
                "total_unique_guardian_number": 0,
                "total_pay_number": 0,
                "total_pay_amount": 0,
                "teacher_number_curr_month": 0,
                "teacher_number_last_month": 0,
                "student_number_curr_month": 0,
                "student_number_last_month": 0,
                "guardian_number_curr_month": 0,
                "guardian_number_last_month": 0,
                "guardian_unique_number_curr_month": 0,
                "guardian_unique_number_last_month": 0,
                "pay_number_curr_month": 0,
                "pay_number_last_month": 0,
                "pay_amount_curr_month": 0,
                "pay_amount_last_month": 0,
                "valid_reading_count_curr_month": 0,
                "valid_reading_count_last_month": 0,
                "valid_exercise_count_curr_month": 0,
                "valid_exercise_count_last_month": 0,
                "valid_word_count_curr_month": 0,
                "valid_word_count_last_month": 0,
                "e_image_c_curr_month": 0,
                "e_image_c_last_month": 0,
                "w_image_c_curr_month": 0,
                "w_image_c_last_month": 0,
                "pay_ratio": 0.0,
                "bind_ratio": 0.0
            }

            school_map = {}
            items_map = {}
            for item in items:
                items_map[item['_id']] = item
            for school in schools:
                default['_id'] = school['id']
                school.update(items_map.get(school['id'], default))
                school_map[school['id']] = school
            offset = 0
            sheet.merge_cells(start_row=start_point, start_column=1, end_row=start_point + 1,
                              end_column=21)
            offset += 1
            title_row1 = sheet[start_point + 2]
            title_row2 = sheet[start_point + 3]
            offset += 3
            for cell in title_row1:
                cell.fill = self._background_header_color()
                cell.font = self._white_font()
                cell.border = self._border()
                cell.alignment = self._alignment()
            for cell in title_row2:
                cell.fill = self._background_header_color()
                cell.font = self._white_font()
                cell.border = self._border()
                cell.alignment = self._alignment()

            title_row1[0].value = "学校名称"
            title_row1[1].value = "教师总数"
            title_row1[2].value = "新增教师"
            title_row1[3].value = "学生总数"
            title_row1[4].value = "新增学生"
            title_row1[5].value = "考试总数"
            title_row1[6].value = "新增考试"
            title_row1[7].value = "单词总数"
            title_row1[8].value = "新增单词"
            title_row1[9].value = "阅读总数"
            title_row1[10].value = "新增阅读"
            title_row1[11].value = "绑定总数"
            title_row1[12].value = "绑定率"
            title_row1[13].value = "新增绑定"
            title_row1[14].value = "付费总数"
            title_row1[15].value = "新增付费"
            if report_type == 'week':
                title_row2[2].value = "上周/本周"
                title_row2[4].value = "上周/本周"
                title_row2[6].value = "上周/本周"
                title_row2[8].value = "上周/本周"
                title_row2[10].value = "上周/本周"
                title_row2[13].value = "上周/本周"
                title_row2[15].value = "上周/本周"
            else:
                title_row2[2].value = "上月/本月"
                title_row2[4].value = "上月/本月"
                title_row2[6].value = "上月/本月"
                title_row2[8].value = "上月/本月"
                title_row2[10].value = "上月/本月"
                title_row2[13].value = "上月/本月"
                title_row2[15].value = "上月/本月"
            index = 0
            print("school_len", len(schools))
            summary_channel_map = defaultdict(list)
            for index, shool_data in enumerate(schools):

                row = sheet[start_point + offset +index]

                for cell in row:
                    cell.font = self._black_font()
                    cell.alignment = self._alignment()
                    cell.border = self._border()

                # 学校名
                row[0].value = shool_data.get("full_name", "")
                row[0].width = len(shool_data.get("full_name", ""))
                # 教师
                row[1].value = shool_data['total_teacher_number']
                row[2].value = str(shool_data['teacher_number_last_month']) + '/' + str(shool_data['teacher_number_curr_month'])
                summary_channel_map[1].append(row[1].value)
                summary_channel_map[2].append(row[2].value)
                # 学生
                row[3].value = shool_data['total_student_number']
                row[4].value = str(shool_data['student_number_last_month']) + '/' + str(shool_data['student_number_curr_month'])
                summary_channel_map[3].append(row[3].value)
                summary_channel_map[4].append(row[4].value)
                # 考试
                row[5].value = shool_data['valid_exercise_count']
                row[6].value = str(shool_data['valid_exercise_count_last_month']) + '/' + str(
                    shool_data['valid_exercise_count_curr_month'])
                summary_channel_map[5].append(row[5].value)
                summary_channel_map[6].append(row[6].value)
                # 单词
                row[7].value = shool_data['valid_word_count']
                row[8].value = str(shool_data['valid_word_count_last_month']) + '/' + str(
                    shool_data['valid_word_count_curr_month'])
                summary_channel_map[7].append(row[7].value)
                summary_channel_map[8].append(row[8].value)
                # 阅读
                row[9].value = shool_data['valid_reading_count']
                row[10].value = str(shool_data['valid_reading_count_last_month']) + '/' + str(
                    shool_data['valid_reading_count_curr_month'])
                summary_channel_map[9].append(row[9].value)
                summary_channel_map[10].append(row[10].value)
                # 绑定
                row[11].value = shool_data['total_unique_guardian_number']
                row[12].value = self.percentage(
                    shool_data['guardian_unique_number_last_month'] / shool_data['total_student_number'] if shool_data[
                                                                                                    'total_student_number'] > 0 else 0)
                row[13].value = str(shool_data['guardian_unique_number_last_month']) + '/' + str(
                    shool_data['guardian_unique_number_curr_month'])
                summary_channel_map[11].append(row[11].value)
                summary_channel_map[12].append(
                    str(shool_data['guardian_unique_number_last_month']) + '/' + str(shool_data['total_student_number']))
                summary_channel_map[13].append(row[13].value)
                # 付费
                row[14].value = shool_data['total_pay_amount']
                row[15].value = str(shool_data['pay_amount_last_month']) + '/' + str(
                    shool_data['pay_amount_curr_month'])
                summary_channel_map[14].append(row[14].value)
                summary_channel_map[15].append(row[15].value)
                index += 1
            print(summary_channel_map)
            start_point +=  len(schools) + offset
            for index, cell in enumerate(sheet[start_point ]):

                cell.font = self._black_font()
                cell.alignment = self._alignment()
                cell.border = self._border()

                cell.font = self._black_font()
                if index == 0:
                    cell.value = "总计"
                    continue
                cell.alignment = self._alignment()
                if index in (1, 3, 5, 7, 9, 11, 14):  # 不带/ 总和
                    cell.value = self.rounding(sum((summary_channel_map.get(index, [0]))))
                elif index in (2, 4, 6, 8, 10, 13, 15):  # 带/
                    cell.value = str(
                        sum([float(item.split('/')[0]) for item in
                             summary_channel_map.get(index, "0/0")])) + "/" + str(
                        sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, "0/0")]))
                elif index in (12,):
                    total_guardian = sum(
                        [float(item.split('/')[0]) for item in summary_channel_map.get(index, ["0/0"])])
                    total_student = sum(
                        [float(item.split('/')[1]) for item in summary_channel_map.get(index, ["0/0"])])
                    cell.value = self.percentage(total_guardian / total_student if total_student else 0)
                else:
                    pass
                if isinstance(cell.value, (int, float)):
                    if cell.value == 0:
                        cell.font = self._red_font()
                if isinstance(cell.value, (str)):
                    if ("/" in cell.value and cell.value.split('/')[0] == "0.0") or \
                            ("/" in cell.value and cell.value.split('/')[1] == "0.0") or \
                            ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                        cell.font = self._red_font()

            start_point += 1
            for index, cell in enumerate(sheet[start_point ]):
                cell.font = self._black_font()
                cell.font = self._black_font()
                cell.alignment = self._alignment()
                cell.border = self._border()
                if index == 0:
                    cell.value = "环比新增率"
                    continue
                cell.alignment = self._alignment()
                if index in (1, 3, 5, 7, 9, 11, 12, 14):  # 不带/last_summary
                    cell.value = "/"
                elif index in (2, 4, 6, 8, 10, 13, 15):  # 带/
                    last_summary = sum(
                        [float(item.split('/')[0]) for item in summary_channel_map.get(index, ["0/0"])])
                    curr_summary = sum(
                        [float(item.split('/')[1]) for item in summary_channel_map.get(index, ["0/0"])])
                    cell.value = self.percentage(
                        (curr_summary - last_summary) / last_summary if last_summary else 0)
                else:
                    pass
                if isinstance(cell.value, (int, float)):
                    if cell.value == 0:
                        cell.font = self._red_font()
                if isinstance(cell.value, (str)):
                    if ("/" in cell.value and cell.value.split('/')[0] == "0") or \
                            ("/" in cell.value and cell.value.split('/')[1] == "0") or \
                            ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                        cell.font = self._red_font()

        # row1 = sheet[1]
        # if report_type == 'week':
        #     last_week = self.last_week()
        #     row1[0].value = "渠道_" + last_week[0] + "-" + last_week[6] + "周报数据"
        # elif report_type == 'month':
        #     _, _, last_month, _, _, _ = self._curr_and_last_and_last_last_month()
        #     month = datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
        #     row1[0].value = "渠道" + str(month) + "月报数据"
        # for one in list(sheet[2:3]):
        #     for cell in one:
        #         cell.font = self._white_font()
        #         cell.fill = self._background_header_color()
        #         cell.border = self._border()
        #
        # summary_map = defaultdict(list)
        # for index, school_data in enumerate(items):
        #     row = sheet[index + 4]
        #     index += 1
        #     for cell in row:
        #         cell.font = self._black_font()
        #         cell.alignment = self._alignment()
        #         cell.border = self._border()
        #     # 大区名字
        #     row[0].value = index
        #     row[1].value = school_map.get(school_data.get("_id", ""), {}).get("full_name", "")
        #
        #     # 新增教师数量
        #     total_teacher_number = school_data.get("total_teacher_number", 0)
        #     teacher_number_curr_month = school_data.get("teacher_number_curr_month", 0)
        #     teacher_number_last_month = school_data.get("teacher_number_last_month", 0)
        #     mom = (teacher_number_curr_month - teacher_number_last_month) / teacher_number_last_month \
        #         if teacher_number_last_month else 0
        #     row[2].value = total_teacher_number
        #     row[3].value = teacher_number_last_month
        #     row[4].value = teacher_number_curr_month
        #     row[5].value = self.percentage(mom)
        #     summary_map[2].append(row[2].value)
        #     summary_map[3].append(row[3].value)
        #     summary_map[4].append(row[4].value)
        #     summary_map[5].append(str(teacher_number_last_month) + '/' + str(teacher_number_curr_month))
        #
        #     # 新增学生数量
        #     total_student_number = school_data.get("total_student_number", 0)
        #     student_number_curr_month = school_data.get("student_number_curr_month", 0)
        #     student_number_last_month = school_data.get("student_number_last_month", 0)
        #     mom = (student_number_curr_month - student_number_last_month) / student_number_last_month \
        #         if student_number_last_month else 0
        #     row[6].value = total_student_number
        #     row[7].value = student_number_last_month
        #     row[8].value = student_number_curr_month
        #     row[9].value = self.percentage(mom)
        #     summary_map[6].append(row[6].value)
        #     summary_map[7].append(row[7].value)
        #     summary_map[8].append(row[8].value)
        #     summary_map[9].append(str(student_number_last_month) + '/' + str(student_number_curr_month))
        #     # 新增考试数量
        #     valid_exercise_count = school_data.get("valid_exercise_count", 0)
        #     valid_exercise_count_curr_month = school_data.get("valid_exercise_count_curr_month", 0)
        #     valid_exercise_count_last_month = school_data.get("valid_exercise_count_last_month", 0)
        #     mom = (valid_exercise_count_curr_month - valid_exercise_count_last_month) / valid_exercise_count_last_month \
        #         if valid_exercise_count_last_month else 0
        #     row[10].value = valid_exercise_count
        #     row[11].value = valid_exercise_count_last_month
        #     row[12].value = valid_exercise_count_curr_month
        #     row[13].value = self.percentage(mom)
        #     summary_map[10].append(row[10].value)
        #     summary_map[11].append(row[11].value)
        #     summary_map[12].append(row[12].value)
        #     summary_map[13].append(str(valid_exercise_count_last_month) + '/' + str(valid_exercise_count_curr_month))
        #     # 新增考试图片数量
        #     e_image_c_curr_month = school_data.get("e_image_c_curr_month", 0)
        #     e_image_c_last_month = school_data.get("e_image_c_last_month", 0)
        #
        #     mom = (e_image_c_curr_month - e_image_c_last_month) / e_image_c_last_month \
        #         if e_image_c_last_month else 0
        #     row[14].value = e_image_c_last_month
        #     row[15].value = e_image_c_curr_month
        #     row[16].value = self.percentage(mom)
        #     summary_map[14].append(row[14].value)
        #     summary_map[15].append(row[15].value)
        #     summary_map[16].append(str(e_image_c_last_month) + '/' + str(e_image_c_curr_month))
        #     # 新增单词数量
        #     valid_word_count = school_data.get("valid_word_count", 0)
        #     valid_word_count_curr_month = school_data.get("valid_word_count_curr_month", 0)
        #     valid_word_count_last_month = school_data.get("valid_word_count_last_month", 0)
        #     mom = (valid_word_count_curr_month - valid_word_count_last_month) / valid_word_count_last_month \
        #         if valid_word_count_last_month else 0
        #     row[17].value = valid_word_count
        #     row[18].value = valid_word_count_last_month
        #     row[19].value = valid_word_count_curr_month
        #     row[20].value = self.percentage(mom)
        #     summary_map[17].append(row[17].value)
        #     summary_map[18].append(row[18].value)
        #     summary_map[19].append(row[19].value)
        #     summary_map[20].append(str(valid_word_count_last_month) + '/' + str(valid_word_count_curr_month))
        #     # 新增单词图像数量
        #     w_image_c = school_data.get("w_image_c", 0)
        #     w_image_c_curr_month = school_data.get("w_image_c_curr_month", 0)
        #     w_image_c_last_month = school_data.get("w_image_c_last_month", 0)
        #     mom = (w_image_c_curr_month - w_image_c_last_month) / w_image_c_last_month \
        #         if w_image_c_last_month else 0
        #     row[21].value = w_image_c_last_month
        #     row[22].value = w_image_c_curr_month
        #     row[23].value = self.percentage(mom)
        #     summary_map[21].append(row[21].value)
        #     summary_map[22].append(row[22].value)
        #     summary_map[23].append(str(w_image_c_last_month) + '/' + str(w_image_c_curr_month))
        #     # 新增阅读数量
        #     valid_reading_count = school_data.get("valid_reading_count", 0)
        #     valid_reading_count_curr_month = school_data.get("valid_reading_count_curr_month", 0)
        #     valid_reading_count_last_month = school_data.get("valid_reading_count_last_month", 0)
        #     mom = (valid_reading_count_curr_month - valid_reading_count_last_month) / valid_reading_count_last_month \
        #         if valid_reading_count_last_month else 0
        #     row[24].value = valid_reading_count
        #     row[25].value = valid_reading_count_last_month
        #     row[26].value = valid_reading_count_curr_month
        #     row[27].value = self.percentage(mom)
        #     summary_map[24].append(row[24].value)
        #     summary_map[25].append(row[25].value)
        #     summary_map[26].append(row[26].value)
        #     summary_map[27].append(str(valid_reading_count_last_month) + '/' + str(valid_reading_count_curr_month))
        #     # 新增家长数量
        #     total_guardian_number = school_data.get("total_guardian_number", 0)
        #     total_unique_guardian_number = school_data.get("total_unique_guardian_number", 0)
        #     guardian_number_curr_month = school_data.get("guardian_number_curr_month", 0)
        #     guardian_unique_number_curr_month = school_data.get("guardian_unique_number_curr_month", 0)
        #     guardian_number_last_month = school_data.get("guardian_number_last_month", 0)
        #     guardian_unique_number_last_month = school_data.get("guardian_unique_number_last_month", 0)
        #     mom = (guardian_unique_number_curr_month - guardian_unique_number_last_month) / guardian_unique_number_last_month if guardian_unique_number_last_month else 0
        #     avg = total_unique_guardian_number / total_student_number if total_student_number > 0 else 0
        #     row[28].value = self.percentage(avg)
        #     row[29].value = guardian_number_last_month
        #     row[30].value = guardian_number_curr_month
        #     row[31].value = self.percentage(mom)
        #     summary_map[28].append(str(total_unique_guardian_number) + "/" + str(total_student_number))
        #     summary_map[29].append(row[29].value)
        #     summary_map[30].append(row[30].value)
        #     summary_map[31].append(str(guardian_number_last_month) + '/' + str(guardian_number_curr_month))
        #     # 新增付费
        #     total_pay_amount = school_data.get("total_pay_amount", 0)
        #     pay_amount_curr_month = school_data.get("pay_amount_curr_month", 0)
        #     pay_amount_last_month = school_data.get("pay_amount_last_month", 0)
        #     mom = (pay_amount_curr_month - pay_amount_last_month) / pay_amount_last_month \
        #         if pay_amount_last_month else 0
        #     row[32].value = total_pay_amount
        #     row[33].value = pay_amount_last_month
        #     row[34].value = pay_amount_curr_month
        #     row[35].value = self.percentage(mom)
        #     summary_map[32].append(row[32].value)
        #     summary_map[33].append(row[33].value)
        #     summary_map[34].append(row[34].value)
        #     summary_map[35].append(str(pay_amount_last_month) + '/' + str(pay_amount_curr_month))
        #     for one in row:
        #         if isinstance(one.value, (int, float)):
        #             if one.value == 0:
        #                 one.font = self._red_font()
        #
        #         if isinstance(one.value, (str)):
        #             if ("/" in one.value and one.value.split('/')[0] == "0") or\
        #                     ("/" in one.value and one.value.split('/')[1] == "0") or \
        #                     ("%" in one.value and one.value.split("%")[0] == '0.00'):
        #                 one.font = self._red_font()
        #
        # total_offset = len(items) + 4
        # divider = len(items)
        # for index, cell in enumerate(sheet[total_offset]):
        #     cell.border = self._border()
        #     cell.alignment = self._alignment()
        #     # print("summary_map.get(index, [0])", index, summary_map.get(index, [0]))
        #     if index == 0:
        #         cell.value = "总计"
        #         continue
        #     if index in (5, 9, 13, 16, 20, 23, 27, 31, 35): #平均值
        #         # print(summary_map.get(index, [0]))
        #         # cell.value = self.percentage(sum((summary_map.get(index, [0]))) / divider if divider > 0 else 0)
        #         last_summary = sum([float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
        #         curr_summary = sum([float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
        #         cell.value = self.percentage((curr_summary - last_summary) / last_summary if last_summary else 0)
        #     elif index in (28,):
        #         total_guardian = sum([float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
        #         total_student = sum([float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
        #         cell.value = self.percentage(total_guardian / total_student if total_student else 0)
        #     else:
        #         cell.value = self.rounding(sum(summary_map.get(index,[0])))
        #     if isinstance(cell.value, (int, float)):
        #         if cell.value == 0:
        #             cell.font = self._red_font()
        #     if isinstance(cell.value, (str)):
        #         if ("/" in cell.value and cell.value.split('/')[0] == "0.0") or \
        #                 ("/" in cell.value and cell.value.split('/')[1] == "0.0") or \
        #                 ("%" in cell.value and cell.value.split("%")[0] == '0.00') or cell.value == '0.00':
        #             cell.font = self._red_font()



        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream