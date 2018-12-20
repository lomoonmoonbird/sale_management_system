#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 按日期导出增量数据
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
from operator import itemgetter
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin
from statistics.area_list_api import AreaList
from utils import CustomEncoder
from statistics.export.export_base import ExportBase

class DateRangeExport(BaseHandler, ExportBase, DataExcludeMixin):
    """
    按时间范围导出报告
    """
    def __init__(self):
        self.db = 'sales'
        self.user_coll = 'sale_user'
        self.instance_coll = 'instance'
        self.class_per_day_coll = 'class_per_day'
        self.grade_per_day_coll = 'grade_per_day'
        self.channel_per_day_coll = 'channel_per_day'
        self.sale_user_coll = 'sale_user'
        self.school_per_day_coll = 'school_per_day'
        self.thread_pool = ThreadPoolExecutor(20)


    @validate_permission()
    async def area_list_export(self, request: Request):
        """
        总部
        {
            "begin_time": "",
            "end_time": ""
        }
        :param request:
        :return:
        """

        request_param = await get_params(request)
        begin_time = request_param.get('begin_time', 0)
        end_time = request_param.get('end_time', 0)
        if not begin_time or not end_time:
            raise RequestError("begin_time or end_time must not be empty")

        areas = request.app['mongodb'][self.db][self.instance_coll].find(
            {"parent_id": request['user_info']['global_id'],
             "role": Roles.AREA.value,
             "status": 1})
        areas = await areas.to_list(100000)

        areas_ids = [str(item['_id']) for item in areas]
        channels = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": {"$in": areas_ids},
                                                                             "role": Roles.CHANNEL.value, "status": 1})
        channels = await channels.to_list(100000)

        old_ids = list(set([item['old_id'] for item in channels]))
        areas_map = {}
        for area in areas:
            areas_map[str(area['_id'])] = area
        channels_map = {}
        for channel in channels:
            channels_map[channel['old_id']] = channel['parent_id']

        exclude_channels = await self.exclude_channel(request.app['mysql'])
        old_ids = list(set(old_ids).difference(set(exclude_channels)))

        items = await self._channel_date_range_list(request, old_ids, begin_time, end_time)

        area_list_obj = AreaList()
        items = await area_list_obj.compute_area_list(request, areas, areas_map, channels_map, items)
        title = "总部 -"+ begin_time + "--" +  end_time + " 大区数据"
        template_path = os.path.dirname(__file__) + "/templates/global_area_date_range_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.area_sheet,
                                                       template_path,
                                                       items,
                                                       title)

        return await self.replay_stream(sheet,
                                        title, request)

    @validate_permission()
    async def channel_list_export(self, request: Request):
        """
        渠道时间范围列表导出
        {
            "area_id": ""
            "begin_time": "",
            "end_time": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        area_id = request_param.get("area_id")
        begin_time = request_param.get('begin_time', 0)
        end_time = request_param.get('end_time', 0)
        if not begin_time or not end_time:
            raise RequestError("begin_time or end_time must not be empty")
        if request['user_info']['instance_role_id'] == Roles.AREA.value:
            area_id = request['user_info']['area_id']
        elif request['user_info']['instance_role_id'] == Roles.GLOBAL.value:
            area_id = area_id

        if not area_id:
            raise RequestError("area_id not found")

        channels = request.app['mongodb'][self.db][self.instance_coll].find(
            {"parent_id": area_id,
             "role": Roles.CHANNEL.value,
             "status": 1
             })

        channels = await channels.to_list(10000)
        items = []
        old_ids = [item['old_id'] for item in channels]
        if old_ids:
            sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
                  ','.join([str(id) for id in old_ids])
            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(sql)
                    real_channels = await cur.fetchall()

            channels_map = {}
            for channel in real_channels:
                channels_map[channel["id"]] = channel

            if request['user_info']['instance_role_id'] == Roles.GLOBAL.value:
                exclude_channels = await self.exclude_channel(request.app['mysql'])
                old_ids = list(set(old_ids).difference(set(exclude_channels)))
            items = await self._channel_date_range_list(request, old_ids, begin_time, end_time)
            for item in items:
                item['contest_coverage_ratio'] = 0
                item['contest_average_per_person'] = 0
                item["channel_info"] = channels_map.get(item["_id"], 0)
            items = items

        title = "大区 -" + begin_time + "--" + end_time + " 渠道数据"
        template_path = os.path.dirname(__file__) + "/templates/area_channel_date_range_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.channel_sheet,
                                                       template_path,
                                                       items,
                                                       title)

        return await self.replay_stream(sheet,
                                        title, request)

    @validate_permission()
    async def market_list_export(self, request: Request):
        """
        时间切片市场列表导出
        {
            "channel_id": ""
            "begin_time": "",
            "end_time": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        begin_time = request_param.get('begin_time', 0)
        end_time = request_param.get('end_time', 0)
        channel_id = request_param.get('channel_id', '')
        if not begin_time or not end_time:
            raise RequestError("begin_time or end_time must not be empty")

        if request['user_info']['instance_role_id'] == Roles.CHANNEL.value:
            channel_id = request['user_info']['channel_id']
        else:
            channel_id = channel_id

        if not channel_id:
            raise RequestError("channel_id not found")

        markets = request.app['mongodb'][self.db][self.sale_user_coll].find({"channel_id": channel_id,
                                                                             "instance_role_id": Roles.MARKET.value,
                                                                             "status": 1})

        markets = await markets.to_list(10000)
        market_user_ids = [int(item['user_id']) for item in markets]
        schools = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": channel_id,
                                                                            "role": Roles.SCHOOL.value,
                                                                            "user_id": {"$in": market_user_ids},
                                                                            "status": 1})
        schools = await schools.to_list(10000)

        channel_info = await request.app['mongodb'][self.db][self.instance_coll].find_one(
            {"_id": ObjectId(channel_id), "status": 1})
        data = []
        if channel_info:
            old_ids = [channel_info.get("old_id", -1)]
            if request['user_info']['instance_role_id'] == Roles.GLOBAL.value:
                exclude_channels = await self.exclude_channel(request.app['mysql'])
                old_ids = list(set(old_ids).difference(set(exclude_channels)))
            items = await self._school_list(request, old_ids, begin_time, end_time)
            item_map = {}
            for item in items:
                item_map[item['_id']] = item
            from collections import defaultdict
            market_school_map = defaultdict(lambda: defaultdict(dict))
            for school in schools:
                default = {
                    "total_school_number": 0,
                    "total_teacher_number": 0,
                    "total_student_number": 0,
                    "total_guardian_number": 0,
                    "total_pay_number": 0,
                    "total_pay_amount": 0,
                    "total_valid_exercise_number": 0,
                    "total_valid_word_number": 0,
                    "total_exercise_image_number": 0,
                    "total_word_image_number": 0,
                    "total_valid_reading_number": 0,
                    "pay_ratio": 0,
                    "bind_ratio": 0,
                    "contest_coverage_ratio": 0,
                    "contest_average_per_person": 0,
                }
                school.update(item_map.get(school['school_id'], default))
                if market_school_map[school['user_id']]['n']:
                    market_school_map[school['user_id']]['n'].append(1)
                else:
                    market_school_map[school['user_id']]['n'] = [1]
                if market_school_map[school['user_id']]['schools']:
                    market_school_map[school['user_id']]['schools'].append(school)
                else:
                    market_school_map[school['user_id']]['schools'] = [school]


            for market in markets:
                tmp = {
                    "market_name": market['nickname'],
                    "total_school_number": 0,
                    "total_teacher_number": 0,
                    "total_student_number": 0,
                    "total_guardian_number": 0,
                    "total_pay_number": 0,
                    "total_pay_amount": 0,
                    "total_valid_exercise_number": 0,
                    "total_valid_word_number": 0,
                    "total_exercise_image_number": 0,
                    "total_word_image_number": 0,
                    "total_valid_reading_number": 0,
                    "pay_ratio": 0,
                    "bind_ratio": 0,
                    "contest_coverage_ratio": 0,
                    "contest_average_per_person": 0,
                }
                one_market = market_school_map.get(int(market['user_id']), tmp)
                if one_market:

                    tmp["total_school_number"] = sum(one_market.get('n')) if one_market.get('n') else 0
                    if one_market.get('schools'):
                        for school in one_market['schools']:
                            tmp["total_teacher_number"] += school['total_teacher_number']
                            tmp['total_student_number'] += school['total_student_number']
                            tmp['total_guardian_number'] += school['total_guardian_number']
                            tmp['total_pay_number'] += school['total_pay_number']
                            tmp['total_pay_amount'] += school['total_pay_amount']
                            tmp['total_valid_exercise_number'] += school['total_valid_exercise_number']
                            tmp['total_valid_word_number'] += school['total_valid_word_number']
                            tmp['total_exercise_image_number'] += school['total_exercise_image_number']
                            tmp['total_word_image_number'] += school['total_word_image_number']
                            tmp['total_valid_reading_number'] += school['total_valid_reading_number']
                            tmp['contest_coverage_ratio'] += school.get('contest_coverage_ratio', 0)
                            tmp['contest_average_per_person'] += school.get("contest_average_per_person", 0)

                    tmp['pay_ratio'] = tmp['total_pay_number'] / tmp['total_student_number'] if tmp[
                                                                                                    'total_student_number'] > 0 else 0
                    tmp['bind_ratio'] = tmp['total_guardian_number'] / tmp['total_student_number'] if tmp[
                                                                                                          'total_student_number'] > 0 else 0

                    data.append(tmp)


        title = "渠道 " + begin_time + "--" + end_time + " 市场数据"
        template_path = os.path.dirname(__file__) + "/templates/channel_market_date_range_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.market_sheet,
                                                       template_path,
                                                       data,
                                                       title)

        return await self.replay_stream(sheet,
                                        title, request)

    def area_sheet(self, template, items, title):
        """
        製作表格
        :param template:
        :param items:
        :param title:
        :return:
        """
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]
        row_title = sheet[1]
        row_title[0].value = title
        row_title[0].alignment = self._alignment()
        row_title[0].font = self._black_font()
        row_title[0].border = self._border()
        row_title[0].fill = self._white_background_color()
        row_header = sheet[2]
        for cell in row_header:
            cell.alignment = self._alignment()
            cell.font = self._white_font()
            cell.border = self._border()
            cell.fill = self._background_header_color()
        for index, item in enumerate(items):
            row = sheet[index+4]
            row[0].value = item['area_info']['name']
            row[1].value = item['total_school_number']
            row[2].value = item['range_school_number']
            row[3].value = item['total_student_number']
            row[4].value = item['total_teacher_number']
            row[5].value = "暂无数据"
            row[6].value = "暂无数据"

            row[7].value = item['total_valid_exercise_number']
            row[8].value = item['range_valid_exercise_number']

            row[9].value = item['total_exercise_image_number']
            row[10].value = item['range_exercise_image_number']

            row[11].value = item['total_valid_reading_number']
            row[12].value = item['range_valid_reading_number']

            row[13].value = item['total_valid_word_number']
            row[14].value = item['range_valid_word_number']
            row[15].value = item['total_word_image_number']
            row[16].value = item['range_word_image_number']

            row[17].value = item['total_guardian_unique_count']
            row[18].value = item['range_guardian_unique_count']

            row[19].value = item['total_pay_amount']
            row[20].value = item['range_pay_amount']

            for cell in row:
                cell.alignment = self._alignment()
                cell.font = self._black_font()
                cell.border = self._border()
                cell.fill = self._white_background_color()


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def channel_sheet(self, template, items, title):
        """
        製作表格
        :param template:
        :param items:
        :param title:
        :return:
        """
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]
        row_title = sheet[1]
        row_title[0].value = title
        row_title[0].alignment = self._alignment()
        row_title[0].font = self._black_font()
        row_title[0].border = self._border()
        row_title[0].fill = self._white_background_color()
        row_header = sheet[2]
        for cell in row_header:
            cell.alignment = self._alignment()
            cell.font = self._white_font()
            cell.border = self._border()
            cell.fill = self._background_header_color()
        for index, item in enumerate(items):
            row = sheet[index+4]
            row[0].value = item['channel_info']['name']
            row[0].value = item['area_info']['name']
            row[1].value = item['total_school_number']
            row[2].value = item['range_school_number']
            row[3].value = item['total_student_number']
            row[4].value = item['total_teacher_number']
            row[5].value = "暂无数据"
            row[6].value = "暂无数据"

            row[7].value = item['total_valid_exercise_number']
            row[8].value = item['range_valid_exercise_number']

            row[9].value = item['total_exercise_image_number']
            row[10].value = item['range_exercise_image_number']

            row[11].value = item['total_valid_reading_number']
            row[12].value = item['range_valid_reading_number']

            row[13].value = item['total_valid_word_number']
            row[14].value = item['range_valid_word_number']
            row[15].value = item['total_word_image_number']
            row[16].value = item['range_word_image_number']

            row[17].value = item['total_guardian_unique_count']
            row[18].value = item['range_guardian_unique_count']

            row[19].value = item['total_pay_amount']
            row[20].value = item['range_pay_amount']

            for cell in row:
                cell.alignment = self._alignment()
                cell.font = self._black_font()
                cell.border = self._border()
                cell.fill = self._white_background_color()


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def market_sheet(self, template, items, title):
        """
        製作表格
        :param template:
        :param items:
        :param title:
        :return:
        """
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]
        row_title = sheet[1]
        row_title[0].value = title
        row_title[0].alignment = self._alignment()
        row_title[0].font = self._black_font()
        row_title[0].border = self._border()
        row_title[0].fill = self._white_background_color()
        row_header = sheet[2]
        for cell in row_header:
            cell.alignment = self._alignment()
            cell.font = self._white_font()
            cell.border = self._border()
            cell.fill = self._background_header_color()
        for index, item in enumerate(items):
            row = sheet[index+4]
            row[0].value = item['market_name']
            row[1].value = item['total_school_number']
            row[2].value = item['total_teacher_number']
            row[3].value = item['total_student_number']
            row[4].value = item['total_valid_exercise_number']
            row[5].value = item['total_exercise_image_number']
            row[6].value = item['total_valid_word_number']
            row[7].value = item['total_word_image_number']
            row[8].value = item['total_valid_reading_number']
            row[9].value = item['total_guardian_number']
            row[10].value = item['bind_ratio']
            row[11].value = item['total_pay_number']
            row[12].value = item['pay_ratio']
            row[13].value = item['total_pay_amount']

            for cell in row:
                cell.alignment = self._alignment()
                cell.font = self._black_font()
                cell.border = self._border()
                cell.fill = self._white_background_color()


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    async def _channel_date_range_list(self, request: Request, channel_ids: list, begin_time: str, end_time: str):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []
        # yesterday = datetime.strptime(end_time, "%Y-%m-%d") - timedelta(1)
        # yesterday_before_30day = yesterday - timedelta(30)
        # yesterday_str = yesterday.strftime("%Y-%m-%d")
        # yesterday_before_30day_str = yesterday_before_30day.strftime("%Y-%m-%d")
        #

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids},
                        # "day": {"$gte": begin_time, "$lte": end_time}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "range_school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                        "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$school_number", 0]},

                        "student_number": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$school_number", 0]},

                        "teacher_number": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$school_number", 0]},

                        "range_valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$valid_exercise_count", 0]},

                        "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, ]}, "$e_image_c", 0]},
                        "range_e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                        "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, ]}, "$w_image_c", 0]},
                        "range_w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},

                        "range_valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$valid_reading_count", 0]},

                        "range_valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                        "valid_word_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$valid_word_count", 0]},

                        "range_guardian_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$guardian_count", 0]},
                        "guardian_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_count", 0]},

                        "range_guardian_unique_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$guardian_count", 0]},
                        "guardian_unique_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_count", 0]},

                        "range_pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]},
                        "pay_amount": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$pay_amount", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "total_school_number": {"$sum": "$school_number"},
                            "range_school_number": {"$sum": "$range_school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "range_guardian_count": {"$sum": "$range_guardian_count"},
                            "total_guardian_unique_count": {"$sum": "$guardian_unique_count"},
                            "range_guardian_unique_count": {"$sum": "$range_guardian_unique_count"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "range_pay_amount": {"$sum": "$range_pay_amount"},
                            "total_valid_reading_number": {"$sum": "$valid_reading_count"},
                            "range_valid_reading_count": {"$sum": "$range_valid_reading_count"},
                            "total_valid_exercise_number": {"$sum": "$valid_exercise_count"},
                            "range_valid_exercise_count": {"$sum": "range_valid_exercise_count"},
                            "total_valid_word_number": {"$sum": "$valid_word_count"},
                            "range_valid_word_count": {"$sum": "$range_valid_word_count"},
                            "total_exercise_image_number": {"$sum": "$e_image_c"},
                            "range_exercise_image_number": {"$sum": "$range_e_image_c"},
                            "total_word_image_number": {"$sum": "$w_image_c"},
                            "range_word_image_number": {"$sum": "$range_w_image_c"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_school_number": 1,
                        "range_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "range_guardian_count": 1,
                        "total_guardian_unique_count": 1,
                        "range_guardian_unique_count": 1,
                        "total_pay_amount": 1,
                        "range_pay_amount": 1,
                        "total_valid_reading_number": 1,
                        "range_valid_reading_count": 1,
                        "total_valid_exercise_number": 1,
                        "range_valid_exercise_count": 1,
                        "total_valid_word_number": 1,
                        "range_valid_word_count": 1,
                        "total_exercise_image_number": 1,
                        "range_exercise_image_number": 1,
                        "total_word_image_number": 1,
                        "range_word_image_number": 1
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _school_list(self, request: Request, channel_ids: list, begin_time:str, end_time: str):
        """
        学校统计
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.school_per_day_coll]
        items = []
        yesterday = datetime.strptime(end_time, "%Y-%m-%d") - timedelta(1)
        yesterday_before_30day = yesterday - timedelta(30)
        yesterday_str = yesterday.strftime("%Y-%m-%d")
        yesterday_before_30day_str = yesterday_before_30day.strftime("%Y-%m-%d")

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids},
                        "day": {"$gte": begin_time, "$lte": end_time}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "school_id": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count":1 ,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$valid_reading_count", 0]},
                        "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$valid_exercise_count", 0]},
                        "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$e_image_c", 0]},
                        "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$valid_word_count", 0]},
                        "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$school_id",
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_guardian_unique_count": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "total_valid_reading_number": {"$sum": "$valid_reading_count"},
                            "total_valid_exercise_number": {"$sum": "$valid_exercise_count"},
                            "total_valid_word_number": {"$sum": "$valid_word_count"},
                            "total_exercise_image_number": {"$sum": "$e_image_c"},
                            "total_word_image_number": {"$sum": "$w_image_c"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_guardian_unique_count": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "total_valid_reading_number": 1,
                        "total_valid_exercise_number": 1,
                        "total_valid_word_number": 1,
                        "total_exercise_image_number": 1,
                        "total_word_image_number": 1,
                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0, {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_guardian_unique_count", "$total_student_number"]}]},
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items