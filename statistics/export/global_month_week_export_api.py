#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 导出全局 大区 渠道月报表格
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
from operator import itemgetter
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin
from tasks.celery_base import BaseTask

class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime)):
            return str(obj)
        if isinstance(obj, (ObjectId)):
            return str(obj)

        return json.JSONEncoder.default(self, obj)

class GlobalExportReport(BaseHandler, ExportBase, DataExcludeMixin):
    def __init__(self):
        self.db = 'sales'
        self.user_coll = 'sale_user'
        self.instance_coll = 'instance'
        self.class_per_day_coll = 'class_per_day'
        self.grade_per_day_coll = 'grade_per_day'
        self.channel_per_day_coll = 'channel_per_day'
        self.start_time = BaseTask().start_time
        self.thread_pool = ThreadPoolExecutor(20)


    @validate_permission(data_validation=True)
    async def month(self, request: Request):
        """
        导出表格
        :param request:
        :return:
        """
        exclude_areas = request['data_permission']['exclude_area']
        exclude_channels_u = request['data_permission']['exclude_channel']
        exclude_area_objectid = [ObjectId(id) for id in exclude_areas]
        instance_coll = request.app['mongodb'][self.db][self.instance_coll]
        user_coll = request.app['mongodb'][self.db][self.user_coll]
        areas = instance_coll.find({"_id": {"$nin": exclude_area_objectid},
                                    "role": Roles.AREA.value,
                                    "status": 1})
        areas = await areas.to_list(100000)
        area_ids = [str(item['_id']) for item in areas]
        stat_areas_channel = await self._get_areas_channels(request, areas)
        channels = instance_coll.find({"parent_id": {"$in": area_ids},
                                       "old_id": {"$nin": exclude_channels_u},
                                       "role": Roles.CHANNEL.value,
                                       "status": 1})

        channels = await channels.to_list(100000)

        area_users = user_coll.find({"area_id": {"$in": area_ids},
                                     "instance_role_id": Roles.AREA.value,
                                     "status": 1})
        area_users = await area_users.to_list(10000)
        area_map = {}
        area_name_id_map = {}
        for area in areas:
            area["_id"] = str(area['_id'])
            area_map[str(area['_id'])] = area
            area_name_id_map[str(area['name'])+"@"+str(area['_id'])] = area
        channel_map = {}
        for channel in channels:

            channel.update({"area_info": area_map.get(channel.get("parent_id", ""), {})})
            channel_map[channel['old_id']] = channel

        for user in area_users:
            user['area_info'] = area_map.get(user['area_id'], {})
        old_ids = [item['old_id'] for item in channels]

        exclude_channels = await self.exclude_channel(request.app['mysql'])
        old_ids = list(set(old_ids).difference(set(exclude_channels+exclude_channels_u)))
        sheet = b''
        if old_ids:
            sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
                  ','.join([str(id) for id in old_ids])
            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(sql)
                    real_channels = await cur.fetchall()
            items = await self._list_month(request, old_ids)
            template_path = os.path.dirname(__file__) + "/templates/global_month_template.xlsx"
            sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                           self.sheet,
                                                           template_path,
                                                           items,
                                                           channel_map,
                                                           area_name_id_map,
                                                           real_channels,
                                                           area_users,
                                                           stat_areas_channel,
                                                           "month")



        return await self.replay_stream(sheet, "总部月报-"+datetime.now().strftime("%Y-%m-%d"), request)

    @validate_permission(data_validation=True)
    async def week(self, request: Request):
        """
        导出表格
        :param request:
        :return:
        """
        exclude_areas = request['data_permission']['exclude_area']
        exclude_channels_u = request['data_permission']['exclude_channel']
        exclude_area_objectid = [ObjectId(id) for id in exclude_areas]
        instance_coll = request.app['mongodb'][self.db][self.instance_coll]
        user_coll = request.app['mongodb'][self.db][self.user_coll]
        areas = instance_coll.find({"_id": {"$nin": exclude_area_objectid},
                                    "role": Roles.AREA.value,
                                    "status": 1})
        areas = await areas.to_list(100000)
        area_ids = [str(item['_id']) for item in areas]
        stat_areas_channel = await self._get_areas_channels(request, areas)
        channels = instance_coll.find({"parent_id": {"$in": area_ids},
                                       "old_id": {"$nin": exclude_channels_u},
                                       "role": Roles.CHANNEL.value,
                                       "status": 1})

        channels = await channels.to_list(100000)

        area_users = user_coll.find({"area_id": {"$in": area_ids},
                                     "instance_role_id": Roles.AREA.value,
                                     "status": 1})
        area_users = await area_users.to_list(10000)
        area_map = {}
        area_name_id_map = {}
        for area in areas:
            area["_id"] = str(area['_id'])
            area_map[str(area['_id'])] = area
            area_name_id_map[str(area['name'])+"@"+str(area['_id'])] = area
        channel_map = {}
        for channel in channels:

            channel.update({"area_info": area_map.get(channel.get("parent_id", ""), {})})
            channel_map[channel['old_id']] = channel

        for user in area_users:
            user['area_info'] = area_map.get(user['area_id'], {})
        old_ids = [item['old_id'] for item in channels]

        exclude_channels = await self.exclude_channel(request.app['mysql'])
        old_ids = list(set(old_ids).difference(set(exclude_channels+exclude_channels_u)))
        sheet = b''
        if old_ids:
            sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
                  ','.join([str(id) for id in old_ids])
            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(sql)
                    real_channels = await cur.fetchall()

            items = await self._list_week(request, old_ids)

            template_path = os.path.dirname(__file__) + "/templates/global_week_template.xlsx"
            sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                           self.sheet,
                                                           template_path,
                                                           items,
                                                           channel_map,
                                                           area_name_id_map,
                                                           real_channels,
                                                           area_users,
                                                           stat_areas_channel,
                                                           "week")


        return await self.replay_stream(sheet, "总部周报-"+datetime.now().strftime("%Y-%m-%d"), request)


    def sheet(self, template, items, channel_map, area_name_id_map, real_channels, users, stat_areas_channel, report_type):
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]

        real_channel_id_name_map = {}
        for r_c_id_name in real_channels:
            real_channel_id_name_map[r_c_id_name['id']] = r_c_id_name['name']

        area_dimesion_items = {}
        channel_dimesion_items = {}
        for item in items:
            channel_dimesion_items[item['_id']] = item

            # 地市
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")), {}).setdefault(
                'total_city_number', []).append(item['total_city_number'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('city_number_last_month', []).append(item['city_number_last_month'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('city_number_curr_month', []).append(
                item['city_number_curr_month'])
            #学校
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")), {}).setdefault('total_school_number', []).append(item['total_school_number'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('school_number_last_month', []).append(item['school_number_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('school_number_curr_month', []).append(
                item['school_number_curr_month'])



            # 新增考试数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_exercise_count', []).append(
                item['valid_exercise_count'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_exercise_count_last_month', []).append(
                item['valid_exercise_count_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_exercise_count_curr_month', []).append(
                item['valid_exercise_count_curr_month'])


            # 新增单词数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_word_count', []).append(
                item['valid_word_count'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_word_count_last_month', []).append(
                item['valid_word_count_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_word_count_curr_month', []).append(
                item['valid_word_count_curr_month'])


            # 新增阅读数
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('valid_reading_count', []).append(
                item['valid_reading_count'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('valid_reading_count_last_month', []).append(
                item['valid_reading_count_last_month'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('valid_reading_count_curr_month', []).append(
                item['valid_reading_count_curr_month'])
            # 学生
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('total_student_number', []).append(
                item['total_student_number'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('student_number_last_month', []).append(
                item['student_number_last_month'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('student_number_curr_month', []).append(
                item['student_number_curr_month'])
            # 新增家长数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('total_guardian_number', []).append(
                item['total_guardian_number'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('guardian_number_last_month', []).append(
                item['guardian_number_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('guardian_number_curr_month', []).append(
                item['guardian_number_curr_month'])

            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('total_unique_guardian_number', []).append(
                item['total_unique_guardian_number'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('guardian_unique_number_last_month', []).append(
                item['guardian_unique_number_last_month'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('guardian_unique_number_curr_month', []).append(
                item['guardian_unique_number_curr_month'])

            # 新增付费数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('total_pay_amount', []).append(
                item['total_pay_amount'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('pay_amount_last_month', []).append(
                item['pay_amount_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('pay_amount_curr_month', []).append(
                item['pay_amount_curr_month'])

            #图像数量
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('total_images', []).append(
                item['total_images'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('total_images_last_month', []).append(
                item['total_images_last_month'])
            area_dimesion_items.setdefault(
                channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(
                    channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                {}).setdefault('total_images_curr_month', []).append(
                item['total_images_curr_month'])

        summary_map = defaultdict(list)
        row1 = sheet[1]
        if report_type == 'week':
            last_week = self.last_week()
            row1[0].value = "全局市场_" + last_week[0] + "-" + last_week[6] + "周报数据"
            row1[0].border = self._border()
            row1[0].font = self._black_font()
        elif report_type == 'month':
            _, _, last_month, _, _, _ = self._curr_and_last_and_last_last_month()
            month = datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
            row1[0].value = "全局市场" + str(month) + "月报数据"
            row1[0].border = self._border()
            row1[0].font = self._black_font()
        for one in list(sheet[2:3]):
            for cell in one:
                cell.font = self._white_font()
                cell.fill = self._background_header_color()
                cell.border = self._border()
        index = 0
        area_dimesion_items_sorted_map = {}
        sorted_area = []

        for name_id, data in area_name_id_map.items():
            data = area_dimesion_items.get(name_id, {})
            area_dimesion_items_sorted_map[name_id] = sum(data.get('valid_exercise_count_curr_month',[0]))
            sorted_area.append({"name_id": name_id,
                                "valid_exercise_count_curr_month": sum(data.get('valid_exercise_count_curr_month',[0]))})

        sorted_area = sorted(sorted_area, key = itemgetter("valid_exercise_count_curr_month"), reverse= True)
        print(json.dumps(sorted_area, indent=4, cls=CustomEncoder))

        for sort_data in sorted_area:
            area_name = sort_data['name_id']
            row = sheet[index + 4]
            # 大区名字
            for cell in row:
                cell.font = self._black_font()
                cell.alignment = self._alignment()
                cell.border = self._border()
            row[0].value = area_name.split('@')[0]

            #地市总数
            last_month_channel = 0
            last_month_channel = 0
            if report_type == 'month':
                last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day, \
                curr_month_first_day, curr_month_last_day = self._curr_and_last_and_last_last_month()
                tmp_last = [item for item in stat_areas_channel.get(area_name.split('@')[1], {}) if
                                          item.get('time_create',"") >= last_last_month_first_day and item.get(
                                              'time_create',"") <= last_last_month_last_day]
                last_month_channel = len(tmp_last) if tmp_last else 0
                tmp_curr = [item for item in stat_areas_channel.get(area_name.split('@')[1], {}) if
                                          item.get('time_create',"") >= last_month_first_day and item.get(
                                              'time_create',"") <= last_month_last_day]
                curr_month_channel = len(tmp_curr) if tmp_curr else 0
            else:
                last_week = self.last_week_from_7_to_6()
                last_last_week = self.last_last_week_from_7_to_6()
                last_week_first_day, last_week_last_day, last_last_week_first_day, last_last_week_last_day = \
                    last_week[0], last_week[6], last_last_week[0], last_last_week[6]
                tmp_last = [item for item in stat_areas_channel.get(area_name.split('@')[1], {}) if
                                          item.get('time_create',"") >= last_last_week_first_day and item.get(
                                              'time_create',"") <= last_last_week_last_day]
                last_month_channel = len(tmp_last) if tmp_last else 0
                tmp_curr = [item for item in stat_areas_channel.get(area_name.split('@')[1], {}) if
                                          item.get('time_create',"") >= last_week_first_day and item.get(
                                              'time_create',"") <= last_week_last_day]
                curr_month_channel = len(tmp_curr) if tmp_curr else 0

            if area_dimesion_items.get(area_name):
                area_data = area_dimesion_items.get(area_name)

                #新增地市 渠道数量


                mom = (sum(area_data['city_number_curr_month']) - sum(area_data['city_number_last_month'])) / sum(
                    area_data['city_number_last_month']) \
                    if sum(area_data['city_number_last_month']) else 0
                print(area_name.split('@')[1])
                if area_name.split('@')[1] == '5be540070195cfd8c80602b5':
                    print("stat_areas_channel.get(area_name.split('@')[1])",stat_areas_channel.get(area_name.split('@')[1]))
                row[1].value = len(stat_areas_channel.get(area_name.split('@')[1], []))
                row[2].value = str(last_month_channel) + "/" + str(curr_month_channel)
                summary_map[1].append(row[1].value)
                summary_map[2].append(row[2].value)
                # 新增学校
                mom = (sum(area_data['school_number_curr_month']) -
                       sum(area_data['school_number_last_month']))/sum(area_data['school_number_last_month']) \
                    if sum(area_data['school_number_last_month']) else 0
                row[3].value = sum(area_data['total_school_number'])
                row[4].value = str(sum(area_data['school_number_last_month'])) + "/" \
                               + str(sum(area_data['school_number_curr_month']))
                summary_map[3].append(row[3].value)
                summary_map[4].append(row[4].value)

                # 新增考试数量
                mom = (sum(area_data['valid_exercise_count_curr_month']) -
                       sum(area_data['valid_exercise_count_last_month'])) / sum(area_data[
                    'valid_exercise_count_last_month']) \
                    if sum(area_data['valid_exercise_count_last_month']) else 0
                row[5].value = sum(area_data['valid_exercise_count'])
                row[6].value = str(sum(area_data['valid_exercise_count_last_month'])) \
                               +"/"+ str(sum(area_data['valid_exercise_count_curr_month']))
                summary_map[5].append(row[5].value)
                summary_map[6].append(row[6].value)

                # 新增单词数量
                mom = (sum(area_data['valid_word_count_curr_month']) -
                       sum(area_data['valid_word_count_last_month'])) / sum(area_data[
                    'valid_word_count_last_month']) \
                    if sum(area_data['valid_word_count_last_month']) else 0
                row[7].value = sum(area_data['valid_word_count'])
                row[8].value = str(sum(area_data['valid_word_count_last_month'])) + \
                               "/" + str(sum(area_data['valid_word_count_curr_month']))
                summary_map[7].append(row[7].value)
                summary_map[8].append(row[8].value)

                # 新增阅读数量
                mom = (sum(area_data['valid_reading_count_curr_month']) - sum(
                    area_data['valid_reading_count_last_month'])) / sum(area_data[
                                                                         'valid_reading_count_last_month']) \
                    if sum(area_data['valid_reading_count_last_month']) else 0
                row[9].value = sum(area_data['valid_reading_count'])
                row[10].value = str(sum(area_data['valid_reading_count_last_month'])) + \
                                "/" + str(sum(area_data['valid_reading_count_curr_month']))
                summary_map[9].append(row[9].value)
                summary_map[10].append(row[10].value)

                #新增图像(单词和考试)
                row[11].value = sum(area_data['total_images'])
                row[12].value = str(sum(area_data['total_images_last_month'])) + "/" + str(
                    sum(area_data['total_images_curr_month']))
                summary_map[11].append(row[11].value)
                summary_map[12].append(row[12].value)

                #绑定
                mom = (sum(area_data['guardian_unique_number_curr_month']) -
                       sum(area_data['guardian_unique_number_last_month'])) / sum(
                    area_data[
                        'guardian_unique_number_last_month']) \
                    if sum(area_data['guardian_unique_number_last_month']) else 0
                row[13].value = sum(area_data['total_unique_guardian_number'])
                row[14].value = self.percentage(mom)
                row[15].value = str(sum(area_data['guardian_unique_number_last_month'])) + "/" + str(
                    sum(area_data['guardian_unique_number_curr_month']))
                summary_map[13].append(row[13].value)
                summary_map[14].append(row[15].value)
                summary_map[15].append(row[15].value)

                # 新增付费
                mom = (sum(area_data['pay_amount_curr_month']) -
                       sum(area_data['pay_amount_last_month'])) / sum(area_data['pay_amount_last_month']) \
                    if sum(area_data['pay_amount_last_month']) else 0
                row[16].value = sum(area_data['total_pay_amount'])
                row[17].value = str(sum(area_data['pay_amount_last_month'])) + "/" + str(sum(area_data['pay_amount_curr_month']))
                summary_map[16].append(row[16].value)
                summary_map[17].append(row[17].value)
                for one in row:
                    if isinstance(one.value, (int, float)):
                        if one.value == 0:
                            one.font = self._red_font()
            else:
                # 新增地市 渠道数量
                row[1].value = len(stat_areas_channel.get(area_name.split('@')[1], []))
                row[2].value = str(last_month_channel) + "/" + str(curr_month_channel)
                summary_map[1].append(row[1].value)
                summary_map[2].append(row[2].value)
                # 新增学校
                mom = 0
                row[3].value = 0
                row[4].value = "0/0"
                summary_map[3].append(row[3].value)
                summary_map[4].append(row[4].value)

                # 新增考试数量
                mom = 0
                row[5].value = 0
                row[6].value = "0/0"
                summary_map[5].append(row[5].value)
                summary_map[6].append(row[6].value)

                # 新增单词数量
                mom = 0
                row[7].value = 0
                row[8].value = "0/0"
                summary_map[7].append(row[7].value)
                summary_map[8].append(row[8].value)

                # 新增阅读数量
                mom = 0
                row[9].value = 0
                row[10].value = "0/0"
                summary_map[9].append(row[9].value)
                summary_map[10].append(row[10].value)
                # 新增图像(单词和考试)
                mom = 0
                avg = 0
                row[11].value = 0
                row[12].value = "0/0"
                summary_map[11].append(0)
                summary_map[12].append(row[12].value)
                # 绑定
                mom = 0
                row[13].value = 0
                row[14].value = self.percentage(0)
                row[15].value = "0/0"
                summary_map[13].append(row[13].value)
                summary_map[14].append(row[15].value)
                summary_map[15].append(row[15].value)

                # 新增付费
                mom = 0
                row[16].value = 0
                row[17].value = "0/0"
                summary_map[16].append(row[16].value)
                summary_map[17].append(row[17].value)

            for one in row:
                # if "大区1" in str(one.value):
                if isinstance(one.value, (int, float)):
                    if one.value == 0:
                        one.font = self._red_font()
                if isinstance(one.value, (str)):
                    if ("/" in one.value and one.value.split('/')[0] == "0") or\
                            ("/" in one.value and one.value.split('/')[1] == "0") or \
                            ("%" in one.value and one.value.split("%")[0] == '0.00'):
                        one.font = self._red_font()
            index += 1


        #总计计算
        total_offset = len(area_name_id_map) + 4
        divider = len(area_name_id_map)
        for index, cell in enumerate(sheet[total_offset]):
            cell.font = self._black_font()
            cell.border = self._border()
            if index == 0:
                cell.value = "总计"
                continue
            cell.alignment = self._alignment()
            if index in (14,): #平均值
                # cell.value = self.percentage(sum((summary_map.get(index, [0]))) / divider if divider > 0 else 0)
                # print('summary_map.get(index, ["0/0"])',index, summary_map.get(index, ["0/0"]))
                last_summary = sum([float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
                curr_summary = sum([float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
                cell.value = self.percentage((curr_summary - last_summary) / last_summary if last_summary else 0)
            elif index in (1,3,5,7,9,11,13,16): #求和
                cell.value = sum([int(item) for item in summary_map.get(index, 0)])
            elif index in (2, 4, 6, 8, 10, 12, 15, 17):  # / 求和
                cell.value = str(
                    sum([float(item.split('/')[0]) for item in summary_map.get(index, "0/0")])) + "/" + str(
                    sum([float(item.split('/')[1]) for item in summary_map.get(index, "0/0")]))
            else:

                cell.value = self.rounding(sum([int(item) for item in summary_map.get(index, [0])]))
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.font = self._red_font()
            if isinstance(cell.value, (str)):
                if ("/" in cell.value and cell.value.split('/')[0] == "0.0") or \
                        ("/" in cell.value and cell.value.split('/')[1] == "0.0") or \
                        ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                    cell.font = self._red_font()

        #环比计算
        total_offset = len(area_name_id_map) + 4 + 1
        for index, cell in enumerate(sheet[total_offset]):
            cell.font = self._black_font()
            cell.font = self._black_font()
            cell.alignment = self._alignment()
            cell.border = self._border()
            if index == 0:
                cell.value = "环比新增率"
                continue
            cell.alignment = self._alignment()
            if index in (1, 3, 5, 7, 9, 11, 13, 14, 16):  # 不带/last_summary
                cell.value = "/"
            elif index in (2, 4, 6, 8, 10 , 12, 15, 17):  # 带/
                last_summary = sum([float(item.split('/')[0]) for item in summary_map.get(index, ["0/0"])])
                curr_summary = sum([float(item.split('/')[1]) for item in summary_map.get(index, ["0/0"])])
                cell.value = self.percentage((curr_summary - last_summary) / last_summary if last_summary else 0)
            else:
                pass
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.font = self._red_font()
            if isinstance(cell.value, (str)):
                if ("/" in cell.value and cell.value.split('/')[0] == "0") or \
                        ("/" in cell.value and cell.value.split('/')[1] == "0") or \
                        ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                    cell.font = self._red_font()



        #大区渠道表格
        start_point = total_offset + 2

        channel_id_map = {}
        for r_c in real_channels:
            channel_id_map[r_c['id']] = r_c['name']

        area_channel_dimesion_map = defaultdict(list)
        for old_id, data in channel_map.items():
            data['channel_stat'] = channel_dimesion_items.get(old_id, {})
            area_channel_dimesion_map[data['area_info']['name'] + "@"+ data['area_info']['_id']].append(data)

        area_with_channel = defaultdict(dict)
        for area_id_name, data in area_name_id_map.items():
            area_with_channel[area_id_name] = area_channel_dimesion_map.get(area_id_name, [])

        delta = 0
        for area_name_id, data in area_with_channel.items():
            if data:
                summary_channel_map = defaultdict(list)
                row = sheet[start_point+delta + 2]
                sheet.merge_cells(start_row=start_point + delta + 2, start_column=1, end_row=start_point + delta + 3,
                                  end_column=21)
                if report_type == 'week':
                    last_week = self.last_week()
                    row[0].value = area_name_id.split("@")[0] +"大区" + last_week[0] + "-" + last_week[6] + "周报数据"
                    row[0].border = self._border()
                    row[0].font = self._black_font()
                    row[0].alignment = self._alignment()
                elif report_type == 'month':
                    _, _, last_month, _, _, _ = self._curr_and_last_and_last_last_month()
                    month = datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
                    row[0].value = area_name_id.split("@")[0] +"大区" + str(month) + "月报数据"
                    row[0].border = self._border()
                    row[0].font = self._black_font()
                    row[0].alignment = self._alignment()


                title_row1 = sheet[start_point+delta + 4]
                title_row2 = sheet[start_point + delta + 5]

                for cell in title_row1:
                    cell.fill = self._background_header_color()
                    cell.font = self._white_font()
                    cell.border = self._border()
                    cell.alignment = self._alignment()
                for cell in title_row2:
                    cell.fill = self._background_header_color()
                    cell.font = self._white_font()
                    cell.border = self._border()
                    cell.alignment = self._alignment()

                title_row1[0].value = "省份"
                title_row1[1].value = "地市"
                title_row1[2].value = "学校总数"
                title_row1[3].value = "新增学校"
                title_row1[4].value = "新增教师"
                title_row1[5].value = "新增学生"
                title_row1[6].value = "考试总数"
                title_row1[7].value = "新增考试"
                title_row1[8].value = "单词总数"
                title_row1[9].value = "新增单词"
                title_row1[10].value = "阅读总数"
                title_row1[11].value = "新增阅读"
                title_row1[12].value = "绑定总数"
                title_row1[13].value = "绑定率"
                title_row1[14].value = "新增绑定"
                title_row1[15].value = "付费总数"
                title_row1[16].value = "新增付费"
                if report_type == 'week':
                    title_row2[3].value = "上周/本周"
                    title_row2[4].value = "上周/本周"
                    title_row2[5].value = "上周/本周"
                    title_row2[7].value = "上周/本周"
                    title_row2[9].value = "上周/本周"
                    title_row2[11].value = "上周/本周"
                    title_row2[14].value = "上周/本周"
                    title_row2[16].value = "上周/本周"
                else:
                    title_row2[3].value = "上月/本月"
                    title_row2[4].value = "上月/本月"
                    title_row2[5].value = "上月/本月"
                    title_row2[7].value = "上月/本月"
                    title_row2[9].value = "上月/本月"
                    title_row2[11].value = "上月/本月"
                    title_row2[14].value = "上月/本月"
                    title_row2[16].value = "上月/本月"

                index = 0
                for channel_stat in data:
                    row = sheet[start_point + delta + index + 6 ]

                    for cell in row:
                        cell.font = self._black_font()
                        cell.alignment = self._alignment()
                        cell.border = self._border()
                    #省份
                    row[0].value = channel_id_map.get(channel_stat['old_id'], "").split("渠道")[0]
                    #渠道名
                    row[1].value = channel_id_map.get(channel_stat['old_id'], "")
                    #学校
                    row[2].value = channel_stat.get("channel_stat", {}).get("total_school_number", 0)
                    row[3].value = str(channel_stat.get("channel_stat", {}).get("school_number_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("school_number_curr_month", 0))
                    summary_channel_map[2].append(row[2].value)
                    summary_channel_map[3].append(row[3].value)

                    #教师
                    row[4].value = str(
                        channel_stat.get("channel_stat", {}).get("teacher_number_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("teacher_number_curr_month", 0))
                    summary_channel_map[4].append(row[4].value)

                    #学生
                    row[5].value = str(
                        channel_stat.get("channel_stat", {}).get("student_number_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("student_number_curr_month", 0))
                    summary_channel_map[5].append(row[5].value)

                    #考试
                    row[6].value = channel_stat.get("channel_stat", {}).get("valid_exercise_count", 0)
                    row[7].value = str(
                        channel_stat.get("channel_stat", {}).get("valid_exercise_count_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("valid_exercise_count_curr_month", 0))

                    summary_channel_map[6].append(row[6].value)
                    summary_channel_map[7].append(row[7].value)

                    #单词
                    row[8].value = channel_stat.get("channel_stat", {}).get("valid_word_count", 0)
                    row[9].value = str(
                        channel_stat.get("channel_stat", {}).get("valid_word_count_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("valid_word_count_curr_month", 0))

                    summary_channel_map[8].append(row[8].value)
                    summary_channel_map[9].append(row[9].value)

                    #阅读
                    row[10].value = channel_stat.get("channel_stat", {}).get("valid_reading_count", 0)
                    row[11].value = str(
                        channel_stat.get("channel_stat", {}).get("valid_reading_count_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("valid_reading_count_curr_month", 0))
                    summary_channel_map[10].append(row[10].value)
                    summary_channel_map[11].append(row[11].value)

                    #绑定
                    avg = channel_stat.get("channel_stat", {}).get("total_unique_guardian_number", 0) / \
                          channel_stat.get("channel_stat", {}).get("total_student_number", 0) \
                        if channel_stat.get("channel_stat", {}).get("total_student_number", 0) > 0 else 0
                    row[12].value = channel_stat.get("channel_stat", {}).get("total_unique_guardian_number", 0)
                    row[13].value = self.percentage(avg)
                    row[14].value = str(
                        channel_stat.get("channel_stat", {}).get("guardian_number_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("guardian_number_curr_month", 0))
                    summary_channel_map[12].append(row[12].value)
                    summary_channel_map[13].append(str(channel_stat.get("channel_stat", {}).get("total_unique_guardian_number", 0))
                                                   + "/" + str(channel_stat.get("channel_stat", {}).get("total_student_number", 0)))
                    summary_channel_map[14].append(row[14].value)


                    #付费
                    row[15].value = channel_stat.get("channel_stat", {}).get("total_pay_amount", 0)
                    row[16].value = str(
                        channel_stat.get("channel_stat", {}).get("pay_amount_last_month", 0)) + "/" + str(
                        channel_stat.get("channel_stat", {}).get("pay_amount_curr_month", 0))
                    summary_channel_map[15].append(row[15].value)
                    summary_channel_map[16].append(row[16].value)

                    for one in row:
                        # if "大区1" in str(one.value):
                        if isinstance(one.value, (int, float)):
                            if one.value == 0:
                                one.font = self._red_font()
                        if isinstance(one.value, (str)):
                            if ("/" in one.value and one.value.split('/')[0] == "0") or \
                                    ("/" in one.value and one.value.split('/')[1] == "0") or \
                                    ("%" in one.value and one.value.split("%")[0] == '0.00'):
                                one.font = self._red_font()

                    index += 1

                delta += len(data) + 6
                divider = len(data)
                for index, cell in enumerate(sheet[start_point+delta ]):

                    cell.font = self._black_font()
                    cell.alignment = self._alignment()
                    cell.border = self._border()

                    cell.font = self._black_font()
                    if index == 0:
                        cell.value = "总计"
                        continue
                    cell.alignment = self._alignment()
                    if index in (2, 6, 8, 10, 12, 15):  #不带/ 总和
                        cell.value = self.rounding(sum((summary_channel_map.get(index, [0]))))
                    elif index in (3, 4, 5, 7, 9, 11, 14, 16): #带/
                        cell.value = str(
                            sum([float(item.split('/')[0]) for item in summary_channel_map.get(index, "0/0")])) + "/" + str(
                            sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, "0/0")]))
                    elif index in (13, ):
                        total_guardian = sum([float(item.split('/')[0]) for item in summary_channel_map.get(index, ["0/0"])])
                        total_student = sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, ["0/0"])])
                        cell.value = self.percentage(total_guardian/total_student if total_student else 0)
                    else:
                        pass
                    if isinstance(cell.value, (int, float)):
                        if cell.value == 0:
                            cell.font = self._red_font()
                    if isinstance(cell.value, (str)):
                        if ("/" in cell.value and cell.value.split('/')[0] == "0.0") or \
                                ("/" in cell.value and cell.value.split('/')[1] == "0.0") or \
                                ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                            cell.font = self._red_font()
                delta += 1
                for index, cell in enumerate(sheet[start_point+delta ]):
                    cell.font = self._black_font()
                    cell.font = self._black_font()
                    cell.alignment = self._alignment()
                    cell.border = self._border()
                    if index == 0:
                        cell.value = "环比新增率"
                        continue
                    cell.alignment = self._alignment()
                    if index in (2, 6, 8, 10, 13, 15):  #不带/last_summary
                        cell.value = "/"
                    elif index in (3, 4, 5, 7, 9, 11, 14, 16): #带/
                        last_summary = sum([float(item.split('/')[0]) for item in summary_channel_map.get(index, ["0/0"])])
                        curr_summary = sum([float(item.split('/')[1]) for item in summary_channel_map.get(index, ["0/0"])])
                        cell.value = self.percentage((curr_summary-last_summary) / last_summary if last_summary else 0)
                    else:
                        pass
                    if isinstance(cell.value, (int, float)):
                        if cell.value == 0:
                            cell.font = self._red_font()
                    if isinstance(cell.value, (str)):
                        if ("/" in cell.value and cell.value.split('/')[0] == "0") or \
                                ("/" in cell.value and cell.value.split('/')[1] == "0") or \
                                ("%" in cell.value and cell.value.split("%")[0] == '0.00'):
                            cell.font = self._red_font()

                delta += 2

        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def _analyze(self, area_dimesion_items, area_name_id_map, users, report_type):
        """
        分析
        :return:
        """
        user_area_map = defaultdict(list)
        for user in users:
            user_area_map[user['area_id']].append(user)

        notification = []
        for area_name, area_data in area_name_id_map.items():
            if area_dimesion_items.get(area_name):
                area_data.update(area_dimesion_items.get(area_name))
            else:
                area_data.update({"total_city_number": [],"city_number_last_month":[], "city_number_curr_month":[],
                                  "total_school_number": [], "school_number_last_month":[],
                                  "school_number_curr_month": [], "total_teacher_number": [],
                                  "teacher_number_last_month": [], "teacher_number_curr_month": [],
                                  "total_student_number": [], "student_number_last_month": [], "student_number_curr_month": [],
                                  "valid_exercise_count": [], "valid_exercise_count_last_month": [], "valid_exercise_count_curr_month":[],
                                  "e_image_c_last_month": [], "e_image_c_curr_month": [], "valid_word_count": [],
                                  "valid_word_count_last_month": [], "valid_word_count_curr_month": [], "w_image_c": [],
                                  "w_image_c_last_month": [], "w_image_c_curr_month": [], "total_guardian_number": [],
                                  "guardian_number_last_month": [], "guardian_number_curr_month":[], "total_pay_amount": [],
                                  "pay_amount_last_month": [], "pay_amount_curr_month": []})
        # print(json.dumps(area_name_id_map, indent=4, cls=CustomEncoder))
        for area_name, area_data in area_name_id_map.items():
            #大区无任何新增
            condition_1 = sum(  area_data['school_number_last_month'] + area_data['teacher_number_last_month'] +
                      area_data['student_number_last_month'] + area_data['valid_exercise_count_last_month'] +
                      area_data['e_image_c_last_month'] + area_data['valid_word_count_last_month'] +
                      area_data['w_image_c_last_month'] + area_data['guardian_number_last_month'] + area_data['pay_amount_last_month']  ) == 0

            a = (sum(area_data['school_number_curr_month']) - sum(area_data['school_number_last_month'])) / sum(
                area_data['school_number_last_month']) \
                if sum(area_data['school_number_last_month']) else 0
            b = (sum(area_data['teacher_number_curr_month']) - sum(area_data['teacher_number_last_month'])) / sum(
                area_data[
                    'teacher_number_last_month']) \
                if sum(area_data['teacher_number_last_month']) else 0
            c = (sum(area_data['student_number_curr_month']) - sum(area_data['student_number_last_month'])) / sum(
                area_data[
                    'student_number_last_month']) \
                if sum(area_data['student_number_last_month']) else 0
            d = (sum(area_data['valid_exercise_count_curr_month']) - sum(
                area_data['valid_exercise_count_last_month'])) / sum(area_data[
                                                                         'valid_exercise_count_last_month']) \
                if sum(area_data['valid_exercise_count_last_month']) else 0
            e = (sum(area_data['e_image_c_curr_month']) - sum(area_data['e_image_c_last_month'])) / sum(area_data[
                                                                                                            'e_image_c_last_month']) \
                if sum(area_data['e_image_c_last_month']) else 0
            f = (sum(area_data['valid_word_count_curr_month']) - sum(area_data['valid_word_count_last_month'])) / sum(
                area_data[
                    'valid_word_count_last_month']) \
                if sum(area_data['valid_word_count_last_month']) else 0
            g = (sum(area_data['w_image_c_curr_month']) - sum(area_data['w_image_c_last_month'])) / sum(area_data[
                                                                                                            'w_image_c_last_month']) \
                if sum(area_data['w_image_c_last_month']) else 0
            h = (sum(area_data['guardian_number_curr_month']) - sum(area_data['guardian_number_last_month'])) / sum(
                area_data[
                    'guardian_number_last_month']) \
                if sum(area_data['guardian_number_last_month']) else 0
            i = (sum(area_data['pay_amount_curr_month']) - sum(area_data['pay_amount_last_month'])) / sum(area_data[
                                                                                                              'pay_amount_last_month']) \
                if sum(area_data['pay_amount_last_month']) else 0
            #大区有数据且有增长率大于100%
            conditon_2 = (a > 1 or b > 1 or c > 1 or d > 1 or e > 1 or f > 1 or g > 1 or h > 1 or i > 1) and not condition_1
            #大区本月有新增使用，无新增家长
            condtion_3 = sum( area_data['valid_exercise_count_last_month'] +
                      area_data['valid_word_count_last_month']  ) > 0 and area_data['guardian_number_last_month'] == 0
            #大区本月无新增使用，有其他新增情况
            condition_4 = sum( area_data['valid_exercise_count_last_month'] +
                               area_data['valid_word_count_last_month']  ) == 0 \
                            and sum(area_data['school_number_last_month'] +
                                     area_data['teacher_number_last_month'] +
                                     area_data['student_number_last_month'] +
                                     area_data['e_image_c_last_month'] +
                                     area_data['w_image_c_last_month'] +
                                     area_data['guardian_number_last_month']+
                                     area_data['pay_amount_last_month']  ) > 0
            if report_type == 'month': #月报
                if  condition_1:
                    area_users = user_area_map.get(area_name.split('@')[1], "")
                    for area_user in area_users:
                        notification.append({"user": area_user,
                                             "info": "本月无任何新增情况，数据异常，建议重点关注"})


                if  conditon_2 :
                    area_users = user_area_map.get(area_name.split('@')[1], "")
                    for area_user in area_users:
                        notification.append({"user": area_user,
                                             "info": area_user.get("area_info",{}).get("name", "") + ',' + area_user.get("nickname", "") +"情况表现良好，有数据增长率大于100%，请继续保持"})

                if  condtion_3:
                    area_users = user_area_map.get(area_name.split('@')[1], "")
                    for area_user in area_users:
                        notification.append({"user": area_user,
                                             "info": "大区经理共使用平台" + str(sum( area_data['valid_exercise_count_last_month'] + area_data['valid_word_count_last_month']  )) + "次，本月有新增使用，但无新增家长，建议推动绑定", })
                if condition_4:

                    area_users = user_area_map.get(area_name.split('@')[1], "")
                    for area_user in area_users:
                        notification.append({"user": area_user,
                                             "info": area_user.get("area_info",{}).get("name", "") + "本月无新增使用，数据异常，建议推动使用"})
            elif report_type == 'week': #周报
                pass
            else:
                pass
        return notification

    async def _list_month(self, request: Request, channel_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """

        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []

        last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day,\
        curr_month_first_day, curr_month_last_day = self._curr_and_last_and_last_last_month()

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "city_number": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "refund_number": 1,
                        "refund_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "total_pay_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_number", 0]},
                        "total_pay_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_amount", 0]},

                        "total_refund_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$refund_number", 0]},
                        "total_refund_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$refund_amount", 0]},

                        "city_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$city_number", 0]},
                        "city_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$city_number", 0]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_count", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_count", 0]},

                        "guardian_unique_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_unique_count", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_images": {"$sum": "$total_images"},
                            "total_city_number": {"$sum": "$city_number"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$total_pay_number"},
                            "total_pay_amount": {"$sum": "$total_pay_amount"},
                            "total_refund_amount": {"$sum": "$total_refund_amount"},
                            "city_number_curr_month": {"$sum": "$city_number_curr_month"},
                            "city_number_last_month": {"$sum": "$city_number_last_month"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                            "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"},
                            "total_images_curr_month": {"$sum": {"$add": ["$e_image_c_curr_month", "$w_image_c_curr_month"]}},
                            "total_images_last_month": {"$sum": {"$add": ["$e_image_c_last_month", "$w_image_c_last_month"]}},
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_city_number": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_unique_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": {"$subtract": ["$total_pay_amount", "$total_refund_amount"]},
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": 1,
                        "total_images_curr_month":1,
                        "total_images_last_month": 1,
                        "city_number_curr_month": 1,
                        "city_number_last_month": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "guardian_unique_number_curr_month": 1,
                        "guardian_unique_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,
                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                 {"$divide": ["$total_unique_guardian_number", "$total_student_number"]}]},
                        # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _list_week(self, request: Request, channel_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []
        last_week = self.last_week_from_7_to_6()
        last_last_week = self.last_last_week_from_7_to_6()
        last_week_first_day, last_week_last_day, last_last_week_first_day, last_last_week_last_day =  \
            last_week[0], last_week[6], last_last_week[0], last_last_week[6]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "city_number": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "guardian_unique_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "refund_number": 1,
                        "refund_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "total_pay_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_number", 0]},
                        "total_pay_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$pay_amount", 0]},

                        "total_refund_number": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$refund_number", 0]},
                        "total_refund_amount": {
                            "$cond": [{"$and": [{"$gte": ["$day", request['data_permission']['pay_stat_start_time']]}]},
                                      "$refund_amount", 0]},

                        "city_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$city_number", 0]},
                        "city_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$city_number", 0]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$guardian_count", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_count", 0]},

                        "guardian_unique_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_number_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_unique_count", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_reading_count", 0]},


                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_images": {"$sum": "$total_images"},
                            "total_city_number": {"$sum": "$city_number"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_unique_guardian_number": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$total_pay_number"},
                            "total_pay_amount": {"$sum": "$total_pay_amount"},
                            "total_refund_amount": {"$sum": "$total_refund_amount"},
                            "city_number_curr_month": {"$sum": "$city_number_curr_month"},
                            "city_number_last_month": {"$sum": "$city_number_last_month"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "guardian_unique_number_curr_month": {"$sum": "$guardian_unique_number_curr_month"},
                            "guardian_unique_number_last_month": {"$sum": "$guardian_unique_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"},
                            "total_images_curr_month": {
                                "$sum": {"$add": ["$e_image_c_curr_month", "$w_image_c_curr_month"]}},
                            "total_images_last_month": {
                                "$sum": {"$add": ["$e_image_c_last_month", "$w_image_c_last_month"]}},
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_city_number": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_unique_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": {"$subtract": ["$total_pay_amount", "$total_refund_amount"]},
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": 1,
                        "total_images_curr_month": 1,
                        "total_images_last_month": 1,
                        "city_number_curr_month": 1,
                        "city_number_last_month": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "guardian_unique_number_curr_month": 1,
                        "guardian_unique_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,

                        "pay_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0, {"$divide": ["$total_pay_number", "$total_student_number"]}]},
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student_number", 0]}, 0,
                                                {"$divide": ["$total_unique_guardian_number", "$total_student_number"]}]},
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _get_areas_channels(self, request: Request, areas):
        """
        获取大区下的渠道
        :param request:
        :param areas:
        :return:
        """
        area_ids = [str(item['_id']) for item in areas]
        channels = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": {"$in": area_ids},
                                                                             "role": Roles.CHANNEL.value,
                                                                             "status": 1}, {"old_id": 1,
                                                                                            "parent_id": 1,
                                                                                            "_id":0})
        channels_mongo = await channels.to_list(None)
        channels_ids = [item['old_id'] for item in channels_mongo]



        sql = "select id, time_create " \
              "from sigma_account_us_user " \
              "where available = 1 " \
              "and role_id = 6 " \
              "and id in (%s) " \
              "and time_create >= '%s' " \
              "and time_create <= '%s'" % (",".join([str(id) for id in channels_ids]),
                                           self.start_time.strftime("%Y-%m-%d"),
                                           datetime.now().strftime("%Y-%m-%d") )

        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(sql)
                channels_mysql = await cur.fetchall()

        channel_map = {}
        for channel in channels_mysql:
            channel_map[channel["id"]] = channel
        from collections import defaultdict
        area_defaultdict = defaultdict(list)
        for channel in channels_mongo:
            if channel_map.get(channel['old_id']):
                one = channel_map.get(channel['old_id'])
                one["time_create"] = one["time_create"].strftime("%Y-%m-%d")
                channel.update(one)
            area_defaultdict[channel['parent_id']].append(channel)
        # print(json.dumps(area_defaultdict, indent=4, cls=CustomEncoder))

        return area_defaultdict
