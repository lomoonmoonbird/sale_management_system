#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 学校专项导出
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
from operator import itemgetter
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin


class GradeExportReport(BaseHandler, ExportBase):
    """
    学校相关项导出
    """

    def __init__(self):
        self.thread_pool = ThreadPoolExecutor(10)
        self.db = "sales"
        self.class_per_day_coll = "class_per_day"

    @validate_permission()
    async def contest_related(self, request: Request):
        """
        考试相关导出
        {
            "school_id": "",
            "grade": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        school_id = request_param.get("school_id")
        grade = request_param.get("grade")
        if not school_id or not grade:
            raise RequestError("school_id or grade must not be empty")
        school_id = int(school_id)
        sql = "select id, name from sigma_account_ob_group where available = 1 and school_id = %s and grade = %s" % (school_id, grade)
        school_sql = "select id, full_name from sigma_account_ob_school where available = 1 and id = %s" % school_id
        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(sql)
                clazzs = await cur.fetchall()
                await cur.execute(school_sql)
                school = await cur.fetchone()


        items = await self._list_school_clazz(request, school_id)
        items_map = {}
        for item in items:
            items_map[item['_id']] = item

        for cla in clazzs:
            default= {
                "total_student_number": 0,
                "total_valid_reading_count": 0,
                "total_valid_reading_count_month": 0,
                "total_valid_exercise_count": 0,
                "total_valid_word_count": 0,
                "total_valid_word_count_month": 0,
                "total_total_valid_contest": 0,
                "total_total_valid_contest_month": 0
            }
            cla['stat_info'] = items_map.get(cla['id'], default)

        title = school['full_name'] + str(grade) + "级 班级考试详情 "
        template_path = os.path.dirname(__file__) + "/templates/school_class_contest_related_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.contest_sheet,
                                                       template_path,
                                                       clazzs,
                                                       title)
        return await self.replay_stream(sheet,
                                        title +
                                        datetime.now().strftime("%Y-%m-%d"), request)

    @validate_permission()
    async def guardian_related(self, request: Request):
        """
        家长 绑定相关导出
        {
            "school_id": ""
            "grade": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        school_id = request_param.get("school_id")
        grade = request_param.get('grade')
        if not school_id or not grade:
            raise RequestError("school_id or grade must not be empty")

        bind_sql = "select DISTINCT user_id " \
                   "from sigma_account_re_userwechat " \
                   "where available = 1 and " \
                   "user_id in ( select id from sigma_account_us_user" \
                   " where school_id = %s and available = 1 and role_id = 2 )" % school_id

        student_sql = "select u.id, u.name, g.grade  " \
                      "from sigma_account_us_user as u " \
                      "join sigma_account_re_groupuser as gu " \
                      "on u.available = 1 and gu.available =1 and u.role_id = 2 and u.id = gu.user_id " \
                      "join sigma_account_ob_group as g " \
                      "on g.id = gu.group_id and g.school_id = %s and g.grade = %s" % (school_id, grade)

        school_sql = "select full_name from sigma_account_ob_school " \
                     "where available = 1 and id = %s " % school_id
        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(bind_sql)
                students_has_wechat = await cur.fetchall()
                await cur.execute(student_sql)
                students = await cur.fetchall()
                await cur.execute(school_sql)
                school = await cur.fetchone()

        title = school['full_name'] + str(grade) + "年级 班级学生绑定情况 "
        template_path = os.path.dirname(__file__) + "/templates/guardian_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.guardian_sheet,
                                                       template_path,
                                                       students_has_wechat,
                                                       students,
                                                       title)
        return await self.replay_stream(sheet,
                                        title +
                                        datetime.now().strftime("%Y-%m-%d"), request)

    async def _list_school_clazz(self, request: Request, school_id: int):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.class_per_day_coll]
        yesterday = datetime.now() - timedelta(1)
        yesterday_before_30day = yesterday - timedelta(30)
        yesterday_str = yesterday.strftime("%Y-%m-%d")
        yesterday_before_30day_str = yesterday_before_30day.strftime("%Y-%m-%d")

        items = []

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "school_id": school_id,
                    }
                },
                {
                    "$project": {
                        "school_id":1,
                        "group_id": 1,
                        "student_number": 1,
                        "valid_reading_count": 1,
                        "valid_reading_count_month": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$valid_reading_count", 0]},
                        "valid_exercise_count": 1,
                        "valid_exercise_count_month": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$valid_exercise_count", 0]},
                        "valid_word_count": 1,
                        "valid_word_count_month": {"$cond": [{"$and": [{"$lte": ["$day", yesterday_str]}, {
                            "$gte": ["$day", yesterday_before_30day_str]}]}, "$valid_word_count", 0]},
                        "total_valid_contest": {"$sum": ["$valid_exercise_count", "$valid_word_count", "$valid_reading_count"]},
                        "total_valid_contest_month": {
                            "$sum": ["$valid_exercise_count_month", "$valid_word_count_month", "$valid_reading_count_month"]},
                        "day": 1
                    }
                },

                {"$group": {"_id": "$group_id",
                            "total_student_number": {"$sum": "$student_number"},
                            "total_valid_reading_count": {"$sum": "$valid_reading_count"},
                            "total_valid_reading_count_month": {"$sum": "$valid_reading_count_month"},
                            "total_valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "total_valid_word_count": {"$sum": "$valid_word_count"},
                            "total_valid_word_count_month": {"$sum": "$valid_word_count_month"},
                            "total_total_valid_contest": {"$sum": "$total_valid_contest"},
                            "total_total_valid_contest_month": {"$sum": "$total_valid_contest_month"}

                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "group_id": 1,
                        "total_student_number": 1,
                        "total_valid_reading_count": 1,
                        "total_valid_reading_count_month": 1,
                        "total_valid_exercise_count": 1,
                        "total_valid_word_count": 1,
                        "total_valid_word_count_month": 1,
                        "total_total_valid_contest": 1,
                        "total_total_valid_contest_month": 1
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items


    def contest_sheet(self, template, items, title):
        """
        製作表格
        :param template:
        :param items:
        :param title:
        :return:
        """
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]
        row_title = sheet[1]
        row_title[0].value = title
        row_title[0].alignment = self._alignment()
        row_title[0].font = self._black_font()
        row_title[0].border = self._border()
        row_title[0].fill = self._white_background_color()
        row_header = sheet[2]
        for cell in row_header:
            cell.alignment = self._alignment()
            cell.font = self._white_font()
            cell.border = self._border()
            cell.fill = self._background_header_color()
        for index, item in enumerate(items):
            row = sheet[index+3]
            row[0].value = item['name']
            row[1].value = item['stat_info']['total_student_number']
            row[2].value = item['stat_info']['total_valid_exercise_count']
            row[3].value = item['stat_info']['total_total_valid_contest']
            row[4].value = item['stat_info']['total_total_valid_contest_month']
            row[5].value = item['stat_info']['total_valid_reading_count']
            row[6].value = item['stat_info']['total_valid_reading_count_month']
            row[7].value = item['stat_info']['total_valid_word_count']
            row[8].value = item['stat_info']['total_valid_word_count_month']

            for cell in row:
                cell.alignment = self._alignment()
                cell.font = self._black_font()
                cell.border = self._border()
                cell.fill = self._white_background_color()


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def guardian_sheet(self, template, students_has_wechat, students, title):
        """
        绑定相关 製作表格
        :param template:
        :param items:
        :param title:
        :return:
        """
        students_has_wechat_ids = [item['user_id'] for item in students_has_wechat]
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]
        row_title = sheet[1]
        row_title[0].value = title
        row_title[0].alignment = self._alignment()
        row_title[0].font = self._black_font()
        row_title[0].border = self._border()
        row_title[0].fill = self._white_background_color()
        row_header = sheet[2]
        for cell in row_header:
            cell.alignment = self._alignment()
            cell.font = self._white_font()
            cell.border = self._border()
            cell.fill = self._background_header_color()
        for index, item in enumerate(students):
            row = sheet[index+3]
            row[0].value = item['name']
            row[1].value = "是" if item['id'] in students_has_wechat_ids else "否"


            for cell in row:
                cell.alignment = self._alignment()
                cell.font = self._black_font()
                cell.border = self._border()
                cell.fill = self._white_background_color()


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream