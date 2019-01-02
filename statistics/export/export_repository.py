#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 导出仓库方法
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
from operator import itemgetter
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin
from webargs import fields
from webargs.aiohttpparser import parser
from tasks.celery_base import BaseTask
from openpyxl import load_workbook, Workbook

class ExportRepository(BaseHandler, ExportBase, DataExcludeMixin):
    """
    导出仓库方法
    """

    def __init__(self):
        self.db = "sales"
        self.channel_per_day_coll = "channel_per_day"
        self.school_per_day_coll = "school_per_day"
        self.start_time = BaseTask().start_time
        self.begin_time = BaseTask().start_time.strftime("%Y-%m-%d")
        self.end_time = datetime.now().strftime("%Y-%m-%d")

    async def daily_new_pay(self, request, channel_ids=[],begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                                end_time=datetime.now().strftime("%Y-%m-%d")):
        """
        每日新增付费
        :param request:
        :param channel_ids:
        :return:
        """
        items = []
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        if not begin_time:
            begin_time = self.begin_time
        if not end_time:
            end_time = self.end_time
        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids},
                        "day": {"$gte": begin_time, "$lte": end_time}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "day": 1,
                        "pay_amount": 1
                    }
                },

                {"$group": {"_id": {"day": "$day", "channel": "$channel"},
                            "total_pay_amount_day": {"$sum": "$pay_amount"},
                            }
                },
                {
                    "$project": {
                        "total_pay_amount_day": 1,

                    }

                },
                {
                    "$sort": {"_id.day": 1}
                }

            ])

        async for item in item_count:
            items.append(item)

        return items


    async def channel_new_delta(self, request, channel_ids=[], begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                                end_time=datetime.now().strftime("%Y-%m-%d")):
        """

        :param request:
        :param channel_ids:
        :return:
        """
        items = []
        if not begin_time:
            begin_time = self.begin_time
        if not end_time:
            end_time = self.end_time

        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids},
                        # "day": {"$gte": begin_time, "$lte": end_time}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "day": 1,
                        "student_number": 1,
                        "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                        "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                        "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                        "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                        "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                        "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},
                        "guardian_unique_count_range": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_count": {
                            "$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_unique_count", 0]},
                        "pay_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$pay_number", 0]},
                        "pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]}
                    }
                },

                {"$group": {"_id": "$channel",
                            "total_school": {"$sum": "$school_number"},
                            "total_exercise": {"$sum": "$valid_exercise_count"},
                            "total_reading": {"$sum": "$valid_reading_count"},
                            "total_word": {"$sum": "$valid_word_count"},
                            # "total_images": {"$sum": {"$add": ["$e_image_c", "$w_image_c"]}}, #这种用法是错误的，如果某一条记录其中一个字段不存在所有的都会返回0
                            "total_unique_guardian_range": {"$sum": "$guardian_unique_count_range"},
                            "total_unique_guardian": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "total_student": {'$sum': "$student_number"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"}
                            }
                 },
                {
                    "$project": {
                        "total_school": 1,
                        "total_exercise": 1,
                        "total_reading": 1,
                        "total_word": 1,
                        "total_images": {"$sum": [ "$e_image_c", "$w_image_c" ] },
                        "total_unique_guardian": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student", 0]}, 0,
                                                 {"$divide": ["$total_unique_guardian",
                                                              "$total_student"]}]},
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items



    async def school_new_delta(self, request, channel_ids=[], begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                                end_time=datetime.now().strftime("%Y-%m-%d")):
        """
        学校新增数据
        :param request:
        :param channel_ids:
        :param begin_time:
        :param end_time:
        :return:
        """
        items = []
        if not begin_time:
            begin_time = self.begin_time
        if not end_time:
            end_time = self.end_time

        coll = request.app['mongodb'][self.db][self.school_per_day_coll]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids},
                        # "day": {"$gte": begin_time, "$lte": end_time}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "school_id": 1,
                        "day": 1,
                        "student_number": 1,
                        "school_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$school_number", 0]},
                        "valid_exercise_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_exercise_count", 0]},
                        "valid_reading_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_reading_count", 0]},
                        "valid_word_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$valid_word_count", 0]},
                        "e_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$e_image_c", 0]},
                        "w_image_c": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$w_image_c", 0]},
                        "guardian_unique_count_range": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$guardian_unique_count", 0]},
                        "guardian_unique_count": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}]}, "$guardian_unique_count", 0]},
                        "pay_number": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$pay_number", 0]},
                        "pay_amount": {"$cond": [{"$and": [{"$lte": ["$day", end_time]}, {
                            "$gte": ["$day", begin_time]}]}, "$pay_amount", 0]}
                    }
                },

                {"$group": {"_id": {"channel": "$channel", "school_id": "$school_id"},
                            "total_school": {"$sum": "$school_number"},
                            "total_exercise": {"$sum": "$valid_exercise_count"},
                            "total_reading": {"$sum": "$valid_reading_count"},
                            "total_word": {"$sum": "$valid_word_count"},
                            "total_images": {"$sum": [ "$e_image_c", "$w_image_c" ] },
                            "total_unique_guardian_range": {"$sum": "$guardian_unique_count_range"},
                            "total_unique_guardian": {"$sum": "$guardian_unique_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "total_student": {'$sum': "$student_number"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"}
                            }
                 },
                {
                    "$project": {
                        "total_school": 1,
                        "total_exercise": 1,
                        "total_reading": 1,
                        "total_word": 1,
                        "total_images": {"$sum": [ "$e_image_c", "$w_image_c" ] },
                        "total_unique_guardian": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "bind_ratio": {"$cond": [{"$eq": ["$total_student", 0]}, 0,
                                                 {"$divide": ["$total_unique_guardian",
                                                              "$total_student"]}]},
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items


    def sheet(self, items=[], area_map=None, channel_map=None, school_items=[], channel_mysql_map={}, school_mysql_map={}):
        """
        生成excel
        :param items:
        :param area_map:
        :param channel_map:
        :return:
        """
        with NamedTemporaryFile() as tmp:
            wb = Workbook(tmp.name)
            default = {
                "total_school": 0,
                "total_exercise": 0,
                "total_reading": 0,
                "total_word": 0,
                "total_images": 0,
                "total_unique_guardian": 0,
                "total_pay_number": 0,
                "total_pay_amount": 0,
                "bind_ratio": 0
            }
            sheet_name = "sheet"
            if items:
                sheet_name = "渠道新增数据"
                wb.create_sheet(title=sheet_name)
                ws = wb[sheet_name]


                item_map = {}
                for item in items:
                    item_map[item['_id']] = item


                for channel_id, data in channel_map.items():
                    if not item_map.get(channel_id):
                        default['_id'] = channel_id
                        items.append(default)
                ws.append(['大区名称', '渠道名称', '新增和学校数量', '新增测试使用次数', '新增阅读使用次数',
                           '新增单词使用次数', '新增图片数量', '新增绑定人数数量', '新增付费人数', '新增付费金额', '绑定率'])
                for row in range(0, len(items)):
                    ws.append([channel_map.get(items[row]['_id']).get("area_name"), channel_map.get(items[row]['_id']).get("name"),
                               items[row]['total_school'], items[row]['total_exercise'],
                               items[row]['total_reading'], items[row]['total_word'],
                               items[row]['total_images'], items[row]['total_unique_guardian'],
                               items[row]['total_pay_number'], items[row]['total_pay_amount'], items[row]['bind_ratio']])
            if school_items:
                sheet_name = "学校新增数据"
                wb.create_sheet(title=sheet_name)
                ws = wb[sheet_name]


                item_map = {}
                for item in school_items:
                    item_map[item['_id']['school_id']] = item

                for school_id, data in school_mysql_map.items():
                    if not item_map.get(school_id):
                        default["_id"] = {}
                        default['_id']['channel'] = data['owner_id']
                        default['_id']['school_id'] = data['id']
                        school_items.append(default)
                ws.append(['渠道', '学校名称', '新增测试使用次数', '新增阅读使用次数',
                           '新增单词使用次数', '新增图片数量', '新增绑定人数数量', '新增付费人数', '新增付费金额', '绑定率'])

                for row in range(0, len(school_items)):
                    ws.append([channel_mysql_map.get(school_items[row]['_id']['channel']).get("name"),
                               school_mysql_map.get(school_items[row]['_id']['school_id']).get("full_name"),
                               school_items[row]['total_exercise'],
                               school_items[row]['total_reading'], school_items[row]['total_word'],
                               school_items[row]['total_images'], school_items[row]['total_unique_guardian'],
                               school_items[row]['total_pay_number'], school_items[row]['total_pay_amount'],
                               school_items[row]['bind_ratio']])

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def channel_daily_sheet(self, items=[], area_map={}, channel_map={}, begin_time=BaseTask().start_time.strftime("%Y-%m-%d"),
                                end_time=datetime.now().strftime("%Y-%m-%d")):
        """
        渠道每日付费excel
        :param items:
        :param area_map:
        :param channel_map:
        :return:
        """
        if not begin_time:
            begin_time = self.begin_time
        if not end_time:
            end_time = self.end_time

        # print("area_map", area_map)
        # print("channel_map", channel_map)
        with NamedTemporaryFile() as tmp:
            wb = Workbook(tmp.name)
            default = {
                "total_pay_amount": 0,
            }
            sheet_name = "sheet"
            if items:
                sheet_name = "渠道新增数据"
                wb.create_sheet(title=sheet_name)
                ws = wb[sheet_name]


                item_map_defaultdict = defaultdict(list)
                for item in items:
                    item_map_defaultdict[item['_id']['channel']].append(item)

                channel_map_only_data = []
                area_channel_stat_defaultdict = defaultdict(list)
                for channel_id, data in channel_map.items():
                    # if not item_map.get(channel_id):
                    #     default['_id'] = channel_id
                    #     items.append(default)
                    data['stat'] = item_map_defaultdict.get(channel_id, [])
                    channel_map_only_data.append(data['stat'])
                    area_channel_stat_defaultdict[data.get("area_id", "")].append(data)


                all_days = list(self.all_day_between_2_date(begin_time, end_time))
                # ws.append(['大区名称', '渠道名称'] + all_days)

                # for row in range(0, len(channel_map_only_data)):
                #     stat_list = channel_map_only_data[row]
                #     area_name = channel_map.get(stat_list[0]['_id']['channel']).get("area_name")
                #     channel_name = channel_map.get(stat_list[0]['_id']['channel']).get("name")
                #     one_row = [area_name, channel_name]
                #
                #     for s in stat_list:
                #         if s['_id']['day'] != all_days[2+row]:
                #             one_row.append(0)
                #         else:
                #             one_row.append(s['total_pay_amount_day'])
                #     ws.append(one_row)
                area_channel_stat_defaultdict["aaaaaaaaaaaaaaaaa"] = area_channel_stat_defaultdict["5c0539dbc6453208d23a8c86"]
                # print(json.dumps(area_channel_stat_defaultdict, indent=4))
                # print(len(area_channel_stat_defaultdict))
                for area_id, data in area_channel_stat_defaultdict.items():

                    # print(data)
                    for subdata in data:
                        index = 0
                        ws.append([subdata['area_name'] + "@" + subdata['name'], "新增付费金额"])
                        stat_list = subdata['stat']
                        stat_list_day_map = {}
                        for s in stat_list:
                            stat_list_day_map[s['_id']['day']] = s
                        stat_list_len = len(stat_list)
                        all_days_len = len(all_days)
                        # print(stat_list_len, all_days_len)
                        for row in range(index, len(all_days)):
                            index+=1
                            ws.append([all_days[row], stat_list_day_map.get(all_days[row], {}).get("total_pay_amount_day", 0)])
                        index += 2
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream