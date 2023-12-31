
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill,Alignment
from datetime import datetime, timedelta

class ExportBase():
    def _red_font(self):
        font = Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FFFF0000')
        return font

    def _black_font(self):
        font = Font(name='Arial',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        return font

    def _white_font(self):
        font = Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFFFF')
        return font

    def _alignment(self):
        return Alignment(horizontal='center', wrapText=True)

    def _background_header_color(self):
        return PatternFill(bgColor="FFC7CE",fgColor="000000", fill_type="solid")

    def _white_background_color(self):
        return PatternFill(fgColor="FFFFFF", fill_type="solid")

    def _border(self):
        # border = Border(left=Side(border_style=None,color = 'FFDDDDDD'),right = Side(border_style=None,color = 'FFDDDDDD'),
        #        top = Side(border_style=None,color = 'FFDDDDDD'),bottom = Side(border_style=None,color = 'FFDDDDDD'),
        #        diagonal = Side(border_style=None,color = 'FFDDDDDD'),diagonal_direction = 0,
        #        outline = Side(border_style=None,color = 'FFDDDDDD'),vertical = Side(border_style=None,color = 'FFDDDDDD'),
        #        horizontal = Side(border_style=None,color = 'FFDDDDDD')
        # )
        thin = Side(border_style="thin", color="FF000000")
        double = Side(border_style="double", color="FF000000")

        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        fill = GradientFill(stop=("000000", "FFFFFF"))
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")
        return border

    def all_day_between_2_date(self, begin, end):
        """
        获取两个日期之间的所有日期
        :param begin:
        :param end:
        :return:
        """
        end = datetime.strptime(end, "%Y-%m-%d")
        begin = datetime.strptime(begin, "%Y-%m-%d")
        delta = end - begin
        for i in range(delta.days + 1):
            yield (begin + timedelta(i)).strftime("%Y-%m-%d")
