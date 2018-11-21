#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 导出全局 大区 渠道月报表格
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params
from demo.utils import validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from auth.utils import insert_area
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor

class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime)):
            return str(obj)
        if isinstance(obj, (ObjectId)):
            return str(obj)

        return json.JSONEncoder.default(self, obj)

class ExportReport(BaseHandler):
    def __init__(self):
        self.db = 'sales'
        self.user_coll = 'sale_user'
        self.instance_coll = 'instance'
        self.class_per_day_coll = 'class_per_day'
        self.grade_per_day_coll = 'grade_per_day'
        self.channel_per_day_coll = 'channel_per_day'
        self.thread_pool = ThreadPoolExecutor(20)

    @validate_permission()
    async def month(self, request: Request):
        """
        导出表格
        :param request:
        :return:
        """
        areas = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": request['user_info']['global_id'],
                                                                          "role": Roles.AREA.value,
                                                                          "status": 1
                                                                          })
        areas = await areas.to_list(100000)
        channels = request.app['mongodb'][self.db][self.instance_coll].find({"role": Roles.CHANNEL.value,
                                                                          "status": 1})

        channels = await channels.to_list(100000)
        area_ids = [str(item['_id'] for item in areas)]
        area_users = request.app['mongodb'][self.db][self.user_coll].find({"area_id": {"$in": area_ids},
                                                                           "instance_role_id": Roles.AREA.value,
                                                                          "status": 1})
        area_users = await area_users.to_list(10000)
        area_map = {}
        for area in areas:
            area_map[str(area['_id'])] = area
        channel_map = {}
        for channel in channels:

            channel.update({"area_info": area_map.get(channel.get("parent_id", ""), {})})
            channel_map[channel['old_id']] = channel

        for user in area_users:
            user['area_info'] = area_map.get(user['area_id'], {})

        old_ids = [item['old_id'] for item in channels]

        items = await self._list_month(request, old_ids)
        template_path = os.path.dirname(__file__) + "/templates/global_month_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.sheet,
                                                       template_path,
                                                       items,
                                                       channel_map,
                                                       area_users)



        return await self.replay_stream(sheet, "总部月报-"+datetime.now().strftime("%Y-%m-%d"), request)

    @validate_permission()
    async def week(self, request: Request):
        """
        导出表格
        :param request:
        :return:
        """
        areas = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": request['user_info']['global_id'],
                                                                          "role": Roles.AREA.value,
                                                                          "status": 1
                                                                          })
        areas = await areas.to_list(100000)
        channels = request.app['mongodb'][self.db][self.instance_coll].find({"role": Roles.CHANNEL.value,
                                                                          "status": 1})

        channels = await channels.to_list(100000)
        area_ids = [str(item['_id'] for item in areas)]
        area_users = request.app['mongodb'][self.db][self.user_coll].find({"area_id": {"$in": area_ids},
                                                                           "instance_role_id": Roles.AREA.value,
                                                                          "status": 1})
        area_users = await area_users.to_list(10000)
        area_map = {}
        for area in areas:
            area_map[str(area['_id'])] = area
        channel_map = {}
        for channel in channels:

            channel.update({"area_info": area_map.get(channel.get("parent_id", ""), {})})
            channel_map[channel['old_id']] = channel

        for user in area_users:
            user['area_info'] = area_map.get(user['area_id'], {})

        old_ids = [item['old_id'] for item in channels]

        items = await self._list_week(request, old_ids)

        template_path = os.path.dirname(__file__) + "/templates/global_week_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.sheet,
                                                       template_path,
                                                       items,
                                                       channel_map,
                                                       area_users)


        return await self.replay_stream(sheet, "总部周报-"+datetime.now().strftime("%Y-%m-%d"), request)


    def sheet(self, template, items, channel_map, users):
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]

        area_dimesion_items = {}
        for item in items:
            #学校
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")), {}).setdefault('total_school_number', []).append(item['total_school_number'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('school_number_last_month', []).append(item['school_number_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('school_number_curr_month', []).append(
                item['school_number_curr_month'])

            #老师
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('total_teacher_number', []).append(item['total_teacher_number'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('teacher_number_last_month', []).append(
                item['teacher_number_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('teacher_number_curr_month', []).append(
                item['teacher_number_curr_month'])

            # 学生
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('total_student_number', []).append(
                item['total_student_number'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('student_number_last_month', []).append(
                item['student_number_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('student_number_curr_month', []).append(
                item['student_number_curr_month'])

            # 新增考试数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_exercise_count', []).append(
                item['valid_exercise_count'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_exercise_count_last_month', []).append(
                item['valid_exercise_count_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_exercise_count_curr_month', []).append(
                item['valid_exercise_count_curr_month'])

            # 新增考试图片数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('e_image_c_last_month', []).append(
                item['e_image_c_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('e_image_c_curr_month', []).append(
                item['e_image_c_curr_month'])

            # 新增单词数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_word_count', []).append(
                item['valid_word_count'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_word_count_last_month', []).append(
                item['valid_word_count_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('valid_word_count_curr_month', []).append(
                item['valid_word_count_curr_month'])

            # 新增单词图片数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('w_image_c', []).append(
                item['w_image_c'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('w_image_c_last_month', []).append(
                item['w_image_c_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('w_image_c_curr_month', []).append(
                item['w_image_c_curr_month'])

            # 新增家长数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('total_guardian_number', []).append(
                item['total_guardian_number'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('guardian_number_last_month', []).append(
                item['guardian_number_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('guardian_number_curr_month', []).append(
                item['guardian_number_curr_month'])

            # 新增付费数
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('total_pay_amount', []).append(
                item['total_pay_amount'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('pay_amount_last_month', []).append(
                item['pay_amount_last_month'])
            area_dimesion_items.setdefault(channel_map.get(item["_id"], {}).get("area_info", {}).get("name", "") + '@' + str(channel_map.get(item["_id"], {}).get("area_info", {}).get("_id", "")),
                                           {}).setdefault('pay_amount_curr_month', []).append(
                item['pay_amount_curr_month'])
        # print(json.dumps(area_dimesion_items, indent=4, cls=CustomEncoder))
        area_number = len(area_dimesion_items)
        total_cell_begin_position = area_number
        spare = 1
        summary_map = defaultdict(list)

        for area_name, area_data in area_dimesion_items.items():
            row = sheet[area_number+4]
            #大区名字
            row[0].value = area_name.split('@')[0]
            #新增地市
            row[1].value = "暂无功能"
            row[2].value = "暂无功能"
            row[3].value = "暂无功能"
            row[4].value = "暂无功能"
            summary_map[1].append(0)
            summary_map[2].append(0)
            summary_map[3].append(0)
            summary_map[4].append(0)
            # 新增学校
            mom = (sum(area_data['school_number_curr_month']) - sum(area_data['school_number_last_month']))/sum(area_data['school_number_last_month']) \
                if sum(area_data['school_number_last_month']) else 0
            row[5].value = sum(area_data['total_school_number'])
            row[6].value = sum(area_data['school_number_last_month'])
            row[7].value = sum(area_data['school_number_curr_month'])
            row[8].value = self.percentage(mom)
            summary_map[5].append(row[5].value)
            summary_map[6].append(row[6].value)
            summary_map[7].append(row[7].value)
            summary_map[8].append(mom)

            # 新增教师数量
            mom = (sum(area_data['teacher_number_curr_month']) - sum(area_data['teacher_number_last_month'])) / sum(area_data[
                'teacher_number_last_month']) \
                if sum(area_data['teacher_number_last_month']) else 0
            row[9].value = sum(area_data['total_teacher_number'])
            row[10].value = sum(area_data['teacher_number_last_month'])
            row[11].value = sum(area_data['teacher_number_curr_month'])
            row[12].value = self.percentage(mom)
            summary_map[9].append(row[9].value)
            summary_map[10].append(row[10].value)
            summary_map[11].append(row[11].value)
            summary_map[12].append(mom)

            # 新增学生数量
            mom = (sum(area_data['student_number_curr_month']) - sum(area_data['student_number_last_month'])) / sum(area_data[
                'student_number_last_month']) \
                if sum(area_data['student_number_last_month']) else 0
            row[13].value = sum(area_data['total_student_number'])
            row[14].value = sum(area_data['student_number_last_month'])
            row[15].value = sum(area_data['student_number_curr_month'])
            row[16].value = self.percentage(mom)
            summary_map[13].append(row[13].value)
            summary_map[14].append(row[14].value)
            summary_map[15].append(row[15].value)
            summary_map[16].append(mom)
            # 新增考试数量
            mom = (sum(area_data['valid_exercise_count_curr_month']) - sum(area_data['valid_exercise_count_last_month'])) / sum(area_data[
                'valid_exercise_count_last_month']) \
                if sum(area_data['valid_exercise_count_last_month']) else 0
            row[17].value = sum(area_data['valid_exercise_count'])
            row[18].value = sum(area_data['valid_exercise_count_last_month'])
            row[19].value = sum(area_data['valid_exercise_count_curr_month'])
            row[20].value = self.percentage(mom)
            summary_map[17].append(row[17].value)
            summary_map[18].append(row[18].value)
            summary_map[19].append(row[19].value)
            summary_map[20].append(mom)
            # 新增考试图片数量
            mom = (sum(area_data['e_image_c_curr_month']) - sum(area_data['e_image_c_last_month'])) / sum(area_data[
                'e_image_c_last_month']) \
                if sum(area_data['e_image_c_last_month']) else 0
            row[21].value = sum(area_data['e_image_c_last_month'])
            row[22].value = sum(area_data['e_image_c_curr_month'])
            row[23].value = self.percentage(mom)
            summary_map[21].append(row[21].value)
            summary_map[22].append(row[22].value)
            summary_map[23].append(mom)
            # 新增单词数量
            mom = (sum(area_data['valid_word_count_curr_month']) - sum(area_data['valid_word_count_last_month'])) / sum(area_data[
                'valid_word_count_last_month']) \
                if sum(area_data['valid_word_count_last_month']) else 0
            row[24].value = sum(area_data['valid_word_count'])
            row[25].value = sum(area_data['valid_word_count_last_month'])
            row[26].value = sum(area_data['valid_word_count_curr_month'])
            row[27].value = self.percentage(mom)
            summary_map[24].append(row[24].value)
            summary_map[25].append(row[25].value)
            summary_map[26].append(row[26].value)
            summary_map[27].append(mom)
            # 新增单词图像数量
            mom = (sum(area_data['w_image_c_curr_month']) - sum(area_data['w_image_c_last_month'])) / sum(area_data[
                'w_image_c_last_month']) \
                if sum(area_data['w_image_c_last_month']) else 0
            row[27].value = sum(area_data['w_image_c'])
            row[28].value = sum(area_data['w_image_c_last_month'])
            row[29].value = sum(area_data['w_image_c_curr_month'])
            row[30].value = self.percentage(mom)
            summary_map[27].append(row[27].value)
            summary_map[28].append(row[28].value)
            summary_map[29].append(row[29].value)
            summary_map[30].append(mom)
            # 新增阅读数量
            row[31].value = "暂无功能"
            row[32].value = "暂无功能"
            row[33].value = "暂无功能"
            row[34].value = "暂无功能"
            summary_map[31].append(0)
            summary_map[32].append(0)
            summary_map[33].append(0)
            summary_map[34].append(0)
            # 新增家长数量
            mom = (sum(area_data['guardian_number_curr_month']) - sum(area_data['guardian_number_last_month'])) / sum(area_data[
                'guardian_number_last_month']) if sum(area_data['guardian_number_last_month']) else 0
            avg = sum(area_data['total_guardian_number']) / sum(area_data['total_student_number'])
            row[35].value = self.percentage(avg)
            row[36].value = sum(area_data['guardian_number_last_month'])
            row[37].value = sum(area_data['guardian_number_curr_month'])
            row[38].value = self.percentage(mom)
            summary_map[35].append(avg)
            summary_map[36].append(row[36].value)
            summary_map[37].append(row[37].value)
            summary_map[38].append(mom)
            # 新增付费
            mom = (sum(area_data['pay_amount_curr_month']) - sum(area_data['pay_amount_last_month'])) / sum(area_data[
                                                                                                              'pay_amount_last_month']) \
                if sum(area_data['pay_amount_last_month']) else 0
            row[39].value = sum(area_data['total_pay_amount'])
            row[40].value = sum(area_data['pay_amount_last_month'])
            row[41].value = sum(area_data['pay_amount_curr_month'])
            row[42].value = self.percentage(mom)
            summary_map[39].append(row[39].value)
            summary_map[40].append(row[40].value)
            summary_map[41].append(row[41].value)
            summary_map[42].append(mom)
            for one in row:
                if isinstance(one.value, (int, float)):
                    if one.value == 0:
                        one.font = self._red_font()




        print(summary_map)
        for index, cell in enumerate(sheet[total_cell_begin_position+spare+4]):
            if index == 0:
                continue
            if index in (4, 8, 12, 16, 20, 23, 27, 30, 34, 35, 38, 42): #平均值
                cell.value = self.percentage(sum((summary_map.get(index, [0]))) / area_number)
            else:
                cell.value = self.rounding(sum(summary_map.get(index,[0])))
        notifications = self._analyze(area_dimesion_items, users)
        for index, notify in enumerate(notifications):
            row = sheet[total_cell_begin_position+spare + 4 + 3]
            row[0].value = notify['user'].get("area_info", {}).get("name")
            row[1].value = notify['info']


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def _analyze(self, area_dimesion_items, users):
        """
        分析
        :return:
        """
        user_area_map = {}
        for user in users:
            user_area_map[user['area_id']] = user
        notification = []
        for area_name, area_data in area_dimesion_items.items():
            #1.若大区经理本月无任何新增情况，则输出：大区经理无新增情况，数据异常，建议重点关注
            if  sum(  area_data['school_number_last_month'] + area_data['teacher_number_last_month'] +
                      area_data['student_number_last_month'] + area_data['valid_exercise_count_last_month'] +
                      area_data['e_image_c_last_month'] + area_data['valid_word_count_last_month'] +
                      area_data['w_image_c_last_month'] + area_data['guardian_number_last_month'] + area_data['pay_amount_last_month']  ) == 0:
                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user, "info": "本月无任何新增情况，数据异常，建议重点关注"})

            a = (sum(area_data['school_number_curr_month']) - sum(area_data['school_number_last_month']))/sum(area_data['school_number_last_month']) \
                if sum(area_data['school_number_last_month']) else 0
            b = (sum(area_data['teacher_number_curr_month']) - sum(area_data['teacher_number_last_month'])) / sum(area_data[
                'teacher_number_last_month']) \
                if sum(area_data['teacher_number_last_month']) else 0
            c = (sum(area_data['student_number_curr_month']) - sum(area_data['student_number_last_month'])) / sum(area_data[
                'student_number_last_month']) \
                if sum(area_data['student_number_last_month']) else 0
            d = (sum(area_data['valid_exercise_count_curr_month']) - sum(area_data['valid_exercise_count_last_month'])) / sum(area_data[
                'valid_exercise_count_last_month']) \
                if sum(area_data['valid_exercise_count_last_month']) else 0
            e = (sum(area_data['e_image_c_curr_month']) - sum(area_data['e_image_c_last_month'])) / sum(area_data[
                'e_image_c_last_month']) \
                if sum(area_data['e_image_c_last_month']) else 0
            f = (sum(area_data['valid_word_count_curr_month']) - sum(area_data['valid_word_count_last_month'])) / sum(area_data[
                'valid_word_count_last_month']) \
                if sum(area_data['valid_word_count_last_month']) else 0
            g = (sum(area_data['w_image_c_curr_month']) - sum(area_data['w_image_c_last_month'])) / sum(area_data[
                'w_image_c_last_month']) \
                if sum(area_data['w_image_c_last_month']) else 0
            h = (sum(area_data['guardian_number_curr_month']) - sum(area_data['guardian_number_last_month'])) / sum(area_data[
                'guardian_number_last_month']) \
                if sum(area_data['guardian_number_last_month']) else 0
            i = (sum(area_data['pay_amount_curr_month']) - sum(area_data['pay_amount_last_month'])) / sum(area_data[
                                                                                                              'pay_amount_last_month']) \
                if sum(area_data['pay_amount_last_month']) else 0
            if  a > 1 or b > 1 or c > 1 or d > 1 or e > 1 or f > 1 or g > 1 or h > 1 or i > 1:
                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user, "info": area_user.get("area_info",{}).get("name", "") + ',' + area_user.get("nickname", "") +"情况表现良好，有数据增长率大于100%，请继续保持"})

            if  sum( area_data['valid_exercise_count_last_month'] +
                      area_data['valid_word_count_last_month']  ) > 0 and area_data['guardian_number_last_month'] == 0:
                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user, "info": "大区经理共使用平台" + str(sum( area_data['valid_exercise_count_last_month'] + area_data['valid_word_count_last_month']  )) + "次，本月有新增使用，但无新增家长，建议推动绑定", })
            if sum( area_data['valid_exercise_count_last_month'] +
                      area_data['valid_word_count_last_month']  ) == 0 and  sum(  area_data['school_number_last_month'] + area_data['teacher_number_last_month'] +
                      area_data['student_number_last_month'] + area_data['e_image_c_last_month'] + area_data['w_image_c_last_month'] + area_data['guardian_number_last_month'] + area_data['pay_amount_last_month'], area_data['']  ) > 0:

                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user,
                                         "info": area_user.get("area_info",{}).get("name", "") + "本月无新增使用，数据异常，建议推动使用"})

        return notification

    async def _list_month(self, request: Request, channel_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []

        last_month_first_day, last_month_last_day, curr_month_first_day, curr_month_last_day = self._curra_and_last_month()

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$guardian_number", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_number", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},

                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                            "$gte": ["$day", curr_month_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                                "$gte": ["$day", curr_month_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                                "$gte": ["$day", curr_month_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", curr_month_last_day]}, {
                                "$gte": ["$day", curr_month_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,

                        "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _list_week(self, request: Request, channel_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []
        current_week = self.current_week()
        last_week = self.last_week()

        last_week_first_day, last_week_last_day, curr_week_first_day, curr_week_last_day =  last_week[0], last_week[6], current_week[0], current_week[6]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$guardian_number", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$guardian_number", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_amount", 0]},

                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                            "$gte": ["$day", curr_week_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                                "$gte": ["$day", curr_week_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                                "$gte": ["$day", curr_week_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", curr_week_last_day]}, {
                                "$gte": ["$day", curr_week_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,

                        "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    def _red_font(self):
        font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FFFF0000')
        return font
