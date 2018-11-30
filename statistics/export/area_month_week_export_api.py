#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 导出全局 大区 渠道月报表格
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin

class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime)):
            return str(obj)
        if isinstance(obj, (ObjectId)):
            return str(obj)

        return json.JSONEncoder.default(self, obj)

class AreaExportReport(BaseHandler, ExportBase, DataExcludeMixin):
    def __init__(self):
        self.db = 'sales'
        self.user_coll = 'sale_user'
        self.instance_coll = 'instance'
        self.class_per_day_coll = 'class_per_day'
        self.grade_per_day_coll = 'grade_per_day'
        self.channel_per_day_coll = 'channel_per_day'
        self.thread_pool = ThreadPoolExecutor(20)

    @validate_permission()
    async def month(self, request: Request):
        """
        导出表格
        {
            "area_id": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        area_id = request_param.get('area_id', "")
        channels = request.app['mongodb'][self.db][self.instance_coll].find({"parent_id": area_id,
                                                                             "role": Roles.CHANNEL.value,
                                                                             "status": 1})

        channels = await channels.to_list(100000)
        channel_ids = [str(item['_id'] for item in channels)]
        channel_users = request.app['mongodb'][self.db][self.user_coll].find({"channel_id": {"$in": channel_ids},
                                                                           "instance_role_id": Roles.CHANNEL.value,
                                                                          "status": 1})
        channel_users = await channel_users.to_list(10000)
        sheet = b''
        old_ids = [item['old_id'] for item in channels]
        area_name = ""
        if old_ids:
            sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
                  ','.join([str(id) for id in old_ids])
            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(sql)
                    real_channels = await cur.fetchall()
            real_channels_map = {}
            for real in real_channels:
                real_channels_map[real['id']] = real
            channel_id_map = {}
            channel_oid_map = {}
            for channel in channels:
                channel_id_map[str(channel['_id'])] = real_channels_map.get(channel['old_id'])
                channel_oid_map[channel['old_id']] = real_channels_map.get(channel['old_id'])

            for user in channel_users:
                user['channel_info'] = channel_id_map.get(user['channel_id'], {})

            area_info = await request.app['mongodb'][self.db][self.instance_coll].find_one({"_id": ObjectId(area_id), "status": 1})
            area_name = area_info.get("name", "")
            exclude_channels = await self.exclude_channel(request.app['mysql'])
            old_ids = list(set(old_ids).difference(set(exclude_channels)))
            items = await self._list_month(request, old_ids)
            template_path = os.path.dirname(__file__) + "/templates/area_month_template.xlsx"
            sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                           self.sheet,
                                                           template_path,
                                                           items,
                                                           real_channels_map,
                                                           channel_users,
                                                           "month",
                                                           area_name)



        return await self.replay_stream(sheet, area_name+"大区月报-"+datetime.now().strftime("%Y-%m-%d"), request)

    @validate_permission()
    async def week(self, request: Request):
        """
        导出表格
        {
            "area_id": ""
        }
        :param request:
        :return:
        """
        request_param = await get_params(request)
        area_id = request_param.get('area_id', "")
        channels = request.app['mongodb'][self.db][self.instance_coll].find(
            {"parent_id": area_id,
             "role": Roles.CHANNEL.value,
             "status": 1})

        channels = await channels.to_list(100000)
        channel_ids = [str(item['_id'] for item in channels)]
        channel_users = request.app['mongodb'][self.db][self.user_coll].find({"channel_id": {"$in": channel_ids},
                                                                              "instance_role_id": Roles.CHANNEL.value,
                                                                              "status": 1})
        channel_users = await channel_users.to_list(10000)

        old_ids = [item['old_id'] for item in channels]
        sheet = b''
        area_name = ""
        if old_ids:
            sql = "select id, name from sigma_account_us_user where available = 1 and id in (%s) " % \
                  ','.join([str(id) for id in old_ids])
            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(sql)
                    real_channels = await cur.fetchall()
            real_channels_map = {}
            for real in real_channels:
                real_channels_map[real['id']] = real
            channel_id_map = {}
            channel_oid_map = {}
            for channel in channels:
                channel_id_map[str(channel['_id'])] = real_channels_map.get(channel['old_id'])
                channel_oid_map[channel['old_id']] = real_channels_map.get(channel['old_id'])

            for user in channel_users:
                user['channel_info'] = channel_id_map.get(user['channel_id'], {})
            area_info = await request.app['mongodb'][self.db][self.instance_coll].find_one(
                {"_id": ObjectId(area_id), "status": 1})
            area_name = area_info.get("name", "")
            exclude_channels = await self.exclude_channel(request.app['mysql'])
            old_ids = list(set(old_ids).difference(set(exclude_channels)))
            items = await self._list_week(request, old_ids)
            template_path = os.path.dirname(__file__) + "/templates/area_week_template.xlsx"
            sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                           self.sheet,
                                                           template_path,
                                                           items,
                                                           real_channels_map,
                                                           channel_users,
                                                           "month",
                                                           area_name)

        return await self.replay_stream(sheet, "大区周报-" + datetime.now().strftime("%Y-%m-%d"), request)



    def sheet(self, template, items, channel_map, users, report_type, area_name):
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]

        area_dimesion_items = {}

        summary_map = defaultdict(list)
        row1 = sheet[1]

        #todo
        if report_type == 'week':
            last_week = self.last_week()
            row1[0].value =area_name+ "大区_" + last_week[0] + "-" + last_week[6] + "周报数据"
        elif report_type == 'month':
            _, _, last_month, _, _, _ = self._curr_and_last_and_last_last_month()
            month = datetime.strptime(last_month, "%Y-%m-%d").timetuple()[1]
            row1[0].value = area_name+" 大区" + str(month) + "月报数据"


        for index, item in enumerate(items):
            row = sheet[index+4]
            #大区名字
            print(item['_id'], channel_map, 'channel_mapchannel_mapchannel_mapchannel_mapchannel_mapchannel_map')
            row[0].value = channel_map.get(item['_id'], {}).get('name')
            # 新增学校
            mom = (item['school_number_curr_month'] - item['school_number_last_month'])/item['school_number_last_month'] \
                if item['school_number_last_month'] else 0
            row[1].value = item['total_school_number']
            row[2].value = item['school_number_last_month']
            row[3].value = item['school_number_curr_month']
            row[4].value = self.percentage(mom)
            summary_map[1].append(row[1].value)
            summary_map[2].append(row[2].value)
            summary_map[3].append(row[3].value)
            summary_map[4].append(mom)

            # 新增教师数量
            mom = (item['teacher_number_curr_month'] - item['teacher_number_last_month']) / item[
                'teacher_number_last_month'] \
                if item['teacher_number_last_month'] else 0
            row[5].value = item['total_teacher_number']
            row[6].value = item['teacher_number_last_month']
            row[7].value = item['teacher_number_curr_month']
            row[8].value = self.percentage(mom)
            summary_map[5].append(row[5].value)
            summary_map[6].append(row[6].value)
            summary_map[7].append(row[7].value)
            summary_map[8].append(mom)

            # 新增学生数量
            mom = (item['student_number_curr_month'] - item['student_number_last_month']) / item[
                'student_number_last_month'] \
                if item['student_number_last_month'] else 0
            row[9].value = item['total_student_number']
            row[10].value = item['student_number_last_month']
            row[11].value = item['student_number_curr_month']
            row[12].value = self.percentage(mom)
            summary_map[9].append(row[9].value)
            summary_map[10].append(row[10].value)
            summary_map[11].append(row[11].value)
            summary_map[12].append(mom)
            # 新增考试数量
            mom = (item['valid_exercise_count_curr_month'] - item['valid_exercise_count_last_month']) / item[
                'valid_exercise_count_last_month'] \
                if item['valid_exercise_count_last_month'] else 0
            row[13].value = item['valid_exercise_count']
            row[14].value = item['valid_exercise_count_last_month']
            row[15].value = item['valid_exercise_count_curr_month']
            row[16].value = self.percentage(mom)
            summary_map[13].append(row[13].value)
            summary_map[14].append(row[14].value)
            summary_map[15].append(row[15].value)
            summary_map[16].append(mom)
            # 新增考试图片数量
            mom = (item['e_image_c_curr_month'] - item['e_image_c_last_month']) / item[
                'e_image_c_last_month'] \
                if item['e_image_c_last_month'] else 0
            row[17].value = item['e_image_c_last_month']
            row[18].value = item['e_image_c_curr_month']
            row[19].value = self.percentage(mom)
            summary_map[17].append(row[17].value)
            summary_map[18].append(row[18].value)
            summary_map[19].append(mom)
            # 新增单词数量
            mom = (item['valid_word_count_curr_month'] - item['valid_word_count_last_month']) / item[
                'valid_word_count_last_month'] \
                if item['valid_word_count_last_month'] else 0
            row[20].value = item['valid_word_count']
            row[21].value = item['valid_word_count_last_month']
            row[22].value = item['valid_word_count_curr_month']
            row[23].value = self.percentage(mom)
            summary_map[20].append(row[20].value)
            summary_map[21].append(row[21].value)
            summary_map[22].append(row[22].value)
            summary_map[23].append(mom)
            # 新增单词图像数量
            mom = (item['w_image_c_curr_month'] - item['w_image_c_last_month']) / item[
                'w_image_c_last_month'] \
                if item['w_image_c_last_month'] else 0
            # row[24].value = sum(channel_data['w_image_c'])
            row[24].value = item['w_image_c_last_month']
            row[25].value = item['w_image_c_curr_month']
            row[26].value = self.percentage(mom)
            summary_map[24].append(row[24].value)
            summary_map[25].append(row[25].value)
            summary_map[26].append(mom)

            # 新增阅读数量
            mom = (item['valid_reading_count_curr_month'] - item['valid_reading_count_last_month']) / item[
                'valid_reading_count_last_month'] \
                if item['valid_reading_count_last_month'] else 0
            row[27].value = item['valid_reading_count']
            row[28].value = item['valid_reading_count_last_month']
            row[29].value = item['valid_reading_count_curr_month']
            row[30].value = self.percentage(mom)
            summary_map[27].append(row[27].value)
            summary_map[28].append(row[28].value)
            summary_map[29].append(row[29].value)
            summary_map[30].append(mom)
            # 新增家长数量
            mom = (item['guardian_number_curr_month'] - item['guardian_number_last_month']) / item[
                'guardian_number_last_month'] if item['guardian_number_last_month'] else 0
            avg = item['total_guardian_number'] / item['total_student_number'] if item['total_student_number'] >0 else 0
            row[31].value = self.percentage(avg)
            row[32].value = item['guardian_number_last_month']
            row[33].value = item['guardian_number_curr_month']
            row[34].value = self.percentage(mom)
            summary_map[31].append(avg)
            summary_map[32].append(row[32].value)
            summary_map[33].append(row[33].value)
            summary_map[34].append(mom)
            # 新增付费
            mom = (item['pay_amount_curr_month'] - item['pay_amount_last_month']) / item[
                                                                                                              'pay_amount_last_month']\
                if item['pay_amount_last_month'] else 0
            row[35].value = item['total_pay_amount']
            row[36].value = item['pay_amount_last_month']
            row[37].value = item['pay_amount_curr_month']
            row[38].value = self.percentage(mom)
            summary_map[35].append(row[35].value)
            summary_map[36].append(row[36].value)
            summary_map[37].append(row[37].value)
            summary_map[38].append(mom)
            for one in row:
                if isinstance(one.value, (int, float)):
                    if one.value == 0:
                        one.font = self._red_font()




        total_offset = len(items) + 4
        divider = len(items)
        for index, cell in enumerate(sheet[total_offset]):
            if index == 0:
                cell.value = "总计"
                continue
            if index in (4, 8, 12, 16, 23, 30, 34,  38): #平均值
                cell.value = self.percentage(sum((summary_map.get(index, [0]))) / divider if divider > 0 else 0)
            else:
                cell.value = self.rounding(sum(summary_map.get(index,[0])))
        row = sheet[total_offset + 2]
        row[0].value = "分析"

        # top = Border(top=self._border().top)
        # sheet[total_cell_begin_position + spare + 3 + 1 +1][0].value = "分析啊啊啊"
        # sheet[total_cell_begin_position + spare + 3 + 1 + 1][0].fill = self._fill()
        # sheet[total_cell_begin_position + spare + 3 + 1 + 1][0].border =sheet[total_cell_begin_position + spare + 3 + 1 + 1][0].border + top
        #分析
        # notifications = self._analyze(area_dimesion_items, users)
        # for index, notify in enumerate(notifications):
        #     row = sheet[total_offset + 3]
        #     row[0].value = notify['user'].get("area_info", {}).get("name")
        #     row[1].value = notify['info']


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def _analyze(self, area_dimesion_items, users):
        """
        分析
        :return:
        """
        user_area_map = {}
        for user in users:
            user_area_map[user['area_id']] = user
        notification = []
        for area_name, channel_data in area_dimesion_items.items():
            #1.若大区经理本月无任何新增情况，则输出：大区经理无新增情况，数据异常，建议重点关注
            if  sum(  channel_data['school_number_last_month'] + channel_data['teacher_number_last_month'] +
                      channel_data['student_number_last_month'] + channel_data['valid_exercise_count_last_month'] +
                      channel_data['e_image_c_last_month'] + channel_data['valid_word_count_last_month'] +
                      channel_data['w_image_c_last_month'] + channel_data['guardian_number_last_month'] + channel_data['pay_amount_last_month']  ) == 0:
                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user, "info": "本月无任何新增情况，数据异常，建议重点关注"})

            a = (sum(channel_data['school_number_curr_month']) - sum(channel_data['school_number_last_month']))/sum(channel_data['school_number_last_month']) \
                if sum(channel_data['school_number_last_month']) else 0
            b = (sum(channel_data['teacher_number_curr_month']) - sum(channel_data['teacher_number_last_month'])) / sum(channel_data[
                'teacher_number_last_month']) \
                if sum(channel_data['teacher_number_last_month']) else 0
            c = (sum(channel_data['student_number_curr_month']) - sum(channel_data['student_number_last_month'])) / sum(channel_data[
                'student_number_last_month']) \
                if sum(channel_data['student_number_last_month']) else 0
            d = (sum(channel_data['valid_exercise_count_curr_month']) - sum(channel_data['valid_exercise_count_last_month'])) / sum(channel_data[
                'valid_exercise_count_last_month']) \
                if sum(channel_data['valid_exercise_count_last_month']) else 0
            e = (sum(channel_data['e_image_c_curr_month']) - sum(channel_data['e_image_c_last_month'])) / sum(channel_data[
                'e_image_c_last_month']) \
                if sum(channel_data['e_image_c_last_month']) else 0
            f = (sum(channel_data['valid_word_count_curr_month']) - sum(channel_data['valid_word_count_last_month'])) / sum(channel_data[
                'valid_word_count_last_month']) \
                if sum(channel_data['valid_word_count_last_month']) else 0
            g = (sum(channel_data['w_image_c_curr_month']) - sum(channel_data['w_image_c_last_month'])) / sum(channel_data[
                'w_image_c_last_month']) \
                if sum(channel_data['w_image_c_last_month']) else 0
            h = (sum(channel_data['guardian_number_curr_month']) - sum(channel_data['guardian_number_last_month'])) / sum(channel_data[
                'guardian_number_last_month']) \
                if sum(channel_data['guardian_number_last_month']) else 0
            i = (sum(channel_data['pay_amount_curr_month']) - sum(channel_data['pay_amount_last_month'])) / sum(channel_data[
                                                                                                              'pay_amount_last_month']) \
                if sum(channel_data['pay_amount_last_month']) else 0
            if  a > 1 or b > 1 or c > 1 or d > 1 or e > 1 or f > 1 or g > 1 or h > 1 or i > 1:
                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user, "info": area_user.get("area_info",{}).get("name", "") + ',' + area_user.get("nickname", "") +"情况表现良好，有数据增长率大于100%，请继续保持"})

            if  sum( channel_data['valid_exercise_count_last_month'] +
                      channel_data['valid_word_count_last_month']  ) > 0 and channel_data['guardian_number_last_month'] == 0:
                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user, "info": "大区经理共使用平台" + str(sum( channel_data['valid_exercise_count_last_month'] + channel_data['valid_word_count_last_month']  )) + "次，本月有新增使用，但无新增家长，建议推动绑定", })
            if sum( channel_data['valid_exercise_count_last_month'] +
                      channel_data['valid_word_count_last_month']  ) == 0 and  sum(  channel_data['school_number_last_month'] + channel_data['teacher_number_last_month'] +
                      channel_data['student_number_last_month'] + channel_data['e_image_c_last_month'] + channel_data['w_image_c_last_month'] + channel_data['guardian_number_last_month'] + channel_data['pay_amount_last_month'], channel_data['']  ) > 0:

                area_users = user_area_map.get(area_name.split('@')[1], "")
                for area_user in area_users:
                    notification.append({"user": area_user,
                                         "info": area_user.get("area_info",{}).get("name", "") + "本月无新增使用，数据异常，建议推动使用"})

        return notification

    async def _list_month(self, request: Request, channel_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []

        last_last_month_first_day, last_last_month_last_day, last_month_first_day, last_month_last_day,\
        curr_month_first_day, curr_month_last_day = self._curr_and_last_and_last_last_month()

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$guardian_number", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$guardian_number", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                            "$gte": ["$day", last_month_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                            "$gte": ["$day", last_last_month_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_month_last_day]}, {
                                "$gte": ["$day", last_month_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_month_last_day]}, {
                                "$gte": ["$day", last_last_month_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,

                        # "pay_ratio": {"$divide": ["$total_pay_number", "$total_student_number"]},
                        # "bind_ratio": {"$divide": ["$total_guardian_number", "$total_student_number"]},
                        # "school_MoM":
                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    async def _list_week(self, request: Request, channel_ids: list):
        """
        学校数
        :param request:
        :param channel_ids:
        :return:
        """
        coll = request.app['mongodb'][self.db][self.channel_per_day_coll]
        items = []
        current_week = self.current_week()
        last_week = self.last_week()
        last_last_week = self.last_last_week()
        last_week_first_day, last_week_last_day, last_last_week_first_day, last_last_week_last_day =  last_week[0], last_week[6], last_last_week[0], last_last_week[6]

        item_count = coll.aggregate(
            [
                {
                    "$match": {
                        "channel": {"$in": channel_ids}
                    }
                },
                {
                    "$project": {
                        "channel": 1,
                        "school_number": 1,
                        "teacher_number": 1,
                        "student_number": 1,
                        "guardian_count": 1,
                        "pay_number": 1,
                        "pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count":1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "total_images": {"$sum": ["$e_image_c", "$w_image_c"]},

                        "school_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$school_number", 0]},
                        "school_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$school_number", 0]},

                        "teacher_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$teacher_number", 0]},
                        "teacher_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$teacher_number", 0]},

                        "student_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$student_number", 0]},
                        "student_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$student_number", 0]},

                        "guardian_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$guardian_number", 0]},
                        "guardian_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$guardian_number", 0]},

                        "pay_number_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_number", 0]},
                        "pay_number_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_number", 0]},

                        "pay_amount_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$pay_amount", 0]},
                        "pay_amount_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$pay_amount", 0]},

                        "valid_reading_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_reading_count", 0]},
                        "valid_reading_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_reading_count", 0]},

                        "valid_exercise_count_curr_month": {"$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                            "$gte": ["$day", last_week_first_day]}]}, "$valid_exercise_count", 0]},
                        "valid_exercise_count_last_month": {"$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                            "$gte": ["$day", last_last_week_first_day]}]}, "$valid_exercise_count", 0]},

                        "valid_word_count_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$valid_word_count", 0]},
                        "valid_word_count_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$valid_word_count", 0]},

                        "e_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$e_image_c", 0]},
                        "e_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$e_image_c", 0]},

                        "w_image_c_curr_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_week_last_day]}, {
                                "$gte": ["$day", last_week_first_day]}]}, "$w_image_c", 0]},
                        "w_image_c_last_month": {
                            "$cond": [{"$and": [{"$lte": ["$day", last_last_week_last_day]}, {
                                "$gte": ["$day", last_last_week_first_day]}]}, "$w_image_c", 0]},

                        "day": 1
                    }
                },

                {"$group": {"_id": "$channel",
                            "valid_reading_count": {"$sum": "$valid_reading_count"},
                            "valid_exercise_count": {"$sum": "$valid_exercise_count"},
                            "valid_word_count": {"$sum": "$valid_word_count"},
                            "e_image_c": {"$sum": "$e_image_c"},
                            "w_image_c": {"$sum": "$w_image_c"},
                            "total_school_number": {"$sum": "$school_number"},
                            "total_teacher_number": {"$sum": "$teacher_number"},
                            "total_student_number": {"$sum": "$student_number"},
                            "total_guardian_number": {"$sum": "$guardian_count"},
                            "total_pay_number": {"$sum": "$pay_number"},
                            "total_pay_amount": {"$sum": "$pay_amount"},
                            "school_number_curr_month": {"$sum": "$school_number_curr_month"},
                            "school_number_last_month": {"$sum": "$school_number_last_month"},
                            "teacher_number_curr_month": {"$sum": "$teacher_number_curr_month"},
                            "teacher_number_last_month": {"$sum": "$teacher_number_last_month"},
                            "student_number_curr_month": {"$sum": "$student_number_curr_month"},
                            "student_number_last_month": {"$sum": "$student_number_last_month"},
                            "guardian_number_curr_month": {"$sum": "$guardian_number_curr_month"},
                            "guardian_number_last_month": {"$sum": "$guardian_number_last_month"},
                            "pay_number_curr_month": {"$sum": "$pay_number_curr_month"},
                            "pay_number_last_month": {"$sum": "$pay_number_last_month"},
                            "pay_amount_curr_month": {"$sum": "$pay_amount_curr_month"},
                            "pay_amount_last_month": {"$sum": "$pay_amount_last_month"},
                            "valid_reading_count_curr_month": {"$sum": "$valid_reading_count_curr_month"},
                            "valid_reading_count_last_month": {"$sum": "$valid_reading_count_last_month"},
                            "valid_exercise_count_curr_month": {"$sum": "$valid_exercise_count_curr_month"},
                            "valid_exercise_count_last_month": {"$sum": "$valid_exercise_count_last_month"},
                            "valid_word_count_curr_month": {"$sum": "$valid_word_count_curr_month"},
                            "valid_word_count_last_month": {"$sum": "$valid_word_count_last_month"},
                            "e_image_c_curr_month": {"$sum": "$e_image_c_curr_month"},
                            "e_image_c_last_month": {"$sum": "$e_image_c_last_month"},
                            "w_image_c_curr_month": {"$sum": "$w_image_c_curr_month"},
                            "w_image_c_last_month": {"$sum": "$w_image_c_last_month"}
                            }
                 },
                {
                    "$project": {
                        "_id": 1,
                        "total_school_number": 1,
                        "total_teacher_number": 1,
                        "total_student_number": 1,
                        "total_guardian_number": 1,
                        "total_pay_number": 1,
                        "total_pay_amount": 1,
                        "valid_reading_count": 1,
                        "valid_exercise_count": 1,
                        "valid_word_count": 1,
                        "e_image_c": 1,
                        "w_image_c": 1,
                        "school_number_curr_month": 1,
                        "school_number_last_month": 1,
                        "teacher_number_curr_month": 1,
                        "teacher_number_last_month": 1,
                        "student_number_curr_month": 1,
                        "student_number_last_month": 1,
                        "guardian_number_curr_month": 1,
                        "guardian_number_last_month": 1,
                        "pay_number_curr_month": 1,
                        "pay_number_last_month": 1,
                        "pay_amount_curr_month": 1,
                        "pay_amount_last_month": 1,
                        "valid_reading_count_curr_month": 1,
                        "valid_reading_count_last_month": 1,
                        "valid_exercise_count_curr_month": 1,
                        "valid_exercise_count_last_month": 1,
                        "valid_word_count_curr_month": 1,
                        "valid_word_count_last_month": 1,
                        "e_image_c_curr_month": 1,
                        "e_image_c_last_month": 1,
                        "w_image_c_curr_month": 1,
                        "w_image_c_last_month": 1,

                    }

                }

            ])

        async for item in item_count:
            items.append(item)

        return items

    def _red_font(self):
        font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FFFF0000')
        return font

    def _fill(self):
        fill = PatternFill("solid", fgColor="DDDDDD")
        return fill

    def _border(self):
        # border = Border(left=Side(border_style=None,color = 'FFDDDDDD'),right = Side(border_style=None,color = 'FFDDDDDD'),
        #        top = Side(border_style=None,color = 'FFDDDDDD'),bottom = Side(border_style=None,color = 'FFDDDDDD'),
        #        diagonal = Side(border_style=None,color = 'FFDDDDDD'),diagonal_direction = 0,
        #        outline = Side(border_style=None,color = 'FFDDDDDD'),vertical = Side(border_style=None,color = 'FFDDDDDD'),
        #        horizontal = Side(border_style=None,color = 'FFDDDDDD')
        # )
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ffdddddd")

        border = Border(top=double, left=thin, right=thin, bottom=double)
        fill = PatternFill("solid", fgColor="DDDDDD")
        fill = GradientFill(stop=("000000", "FFFFFF"))
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")
        return border