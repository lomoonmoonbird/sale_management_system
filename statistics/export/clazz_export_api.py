#! python3.6
# --*-- coding: utf-8 --*--

"""
数据统计 ： 导出全局 大区 渠道月报表格
"""

from datetime import datetime, timedelta, date
import asyncio
import mimetypes
import os
from tempfile import NamedTemporaryFile
import json
import time
from aiohttp.web import Request
from configs import UC_SYSTEM_API_ADMIN_URL, THEMIS_SYSTEM_ADMIN_URL, ucAppKey, ucAppSecret, permissionAppKey
import aiohttp
import ujson
from utils import get_json, get_params, validate_permission
from basehandler import BaseHandler
from exceptions import InternalError, UserExistError, CreateUserError, DELETEERROR, RequestError
from menu.menu import Menu
from motor.core import Collection
from enum import Enum
from aiomysql.cursors import DictCursor
from pymongo import UpdateOne, DeleteMany
from bson import ObjectId
from collections import defaultdict
from enumconstant import Roles, PermissionRole
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bson import ObjectId
from concurrent.futures import ThreadPoolExecutor
from statistics.export.export_base import ExportBase
from mixins import DataExcludeMixin


class ClazzExportReport(BaseHandler, ExportBase ):
    """
    班级导出
    """

    def __init__(self):
        self.thread_pool = ThreadPoolExecutor(10)

    async def clazz_pay_export(self, request: Request):
        """
        班级付费导出
        {
            "group_id": ""
        }
        :return:
        """

        request_param = await get_params(request)
        group_id = request_param.get("group_id")
        class_data= []
        if not group_id:
            raise RequestError("group_id must not be empty")

        user_sql = "select u.id, u.name, u.student_vip_expire " \
                   "from sigma_account_us_user as u " \
                   "join sigma_account_re_groupuser as gu " \
                   "on u.available = 1 " \
                   "and gu.available = 1 " \
                   "and u.role_id = 2 " \
                   "and gu.user_id = u.id and gu.group_id = %s" % group_id
        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(user_sql)
                users = await cur.fetchall()
        if users:
            user_ids = [item['id'] for item in users]
            user_wechat_sql = "select user_id from sigma_account_re_userwechat where user_id in (%s)" % (
                ','.join([str(id) for id in user_ids]))

            print(user_wechat_sql)
            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(user_wechat_sql)
                    userwechat = await cur.fetchall()

            userwechat_ids = [item['user_id'] for item in userwechat]
            pay_sql = "select user_id, coupon_amount " \
                      "from sigma_pay_ob_order " \
                      "where available = 1 " \
                      "and status = 3 " \
                      "and user_id in (%s)" % (','.join([str(id) for id in user_ids]))

            async with request.app['mysql'].acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    await cur.execute(pay_sql)
                    pay = await cur.fetchall()

            pay_map = defaultdict(list)
            for p in pay:
                pay_map[p['user_id']].append(p['coupon_amount'])

            current_timestamp = time.time()
            for u in users:
                class_data.append({
                    "name": u["name"],
                    "student_id": u['id'],
                    "is_bind": 1 if u['id'] in userwechat_ids else 0,
                    "is_paid": 1 if sum(pay_map.get(u['id'], [0])) > 0 else 0,
                    "pay_amount": sum(pay_map.get(u['id'], [0])),
                    "duration": 0 if current_timestamp > u['student_vip_expire'] else (
                            datetime.fromtimestamp(u['student_vip_expire']) - datetime.fromtimestamp(
                        current_timestamp)).days
                })

        group_info_sql = "select g.name, g.grade, g.school_id, s.full_name " \
                         "from sigma_account_ob_group as g " \
                         "join sigma_account_ob_school as s " \
                         "on g.available = 1 and g.school_id = s.id"
        async with request.app['mysql'].acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                await cur.execute(group_info_sql)
                group_info = await cur.fetchone()

        # sql = "select u.id, u.name, u.student_vip_expire, sum(o.coupon_amount) as total_amount, count(uw.wechat_id) as total_wechat " \
        #       "from sigma_account_re_groupuser as gu " \
        #       "join sigma_account_us_user as u " \
        #       "on u.available = 1 and gu.available = 1 and u.role_id = 2  and gu.user_id = u.id " \
        #       "left join sigma_pay_ob_order as o " \
        #       "on o.available = 1 and o.status = 3 and u.id = o.user_id " \
        #       "left join sigma_account_re_userwechat as uw  " \
        #       "on uw.available = 1 and uw.user_id = o.user_id " \
        #       "where gu.group_id = %s  " \
        #       "group by u.id,uw.user_id;" % (group_id)
        #
        # group_info_sql = "select g.name, g.grade, g.school_id, s.full_name " \
        #                  "from sigma_account_ob_group as g " \
        #                  "join sigma_account_ob_school as s " \
        #                  "on g.available = 1 and g.school_id = s.id"
        # async with request.app['mysql'].acquire() as conn:
        #     async with conn.cursor(DictCursor) as cur:
        #         await cur.execute(sql)
        #         clazz = await cur.fetchall()
        #         await cur.execute(group_info_sql)
        #         group_info = await cur.fetchone()
        # class_data = []
        #
        # for cla in clazz:
        #     current_timestamp = time.time()
        #     class_data.append({
        #         "name": cla["name"],
        #         "student_id": cla['id'],
        #         "is_bind": 1 if cla['total_wechat'] is not None and cla['total_wechat'] > 0 else 0,
        #         "is_paid": 1 if cla['total_amount'] is not None and cla['total_amount'] > 0 else 0,
        #         "pay_amount": cla['total_amount'] if cla['total_amount'] is not None else 0,
        #         "duration": 0 if current_timestamp > cla['student_vip_expire'] else (
        #                     datetime.fromtimestamp(cla['student_vip_expire']) - datetime.fromtimestamp(
        #                 current_timestamp)).days
        #     })

        title = group_info['full_name'] + "-" + group_info['grade'] + "级" +group_info['name'] + " 付费-"
        template_path = os.path.dirname(__file__) + "/templates/clazz_pay_template.xlsx"
        sheet = await request.app.loop.run_in_executor(self.thread_pool,
                                                       self.sheet,
                                                       template_path,
                                                       class_data,
                                                       title)
        return await self.replay_stream(sheet,
                                        group_info['full_name'] + "-" +
                                        group_info['grade'] + "级" +
                                        group_info['name'] + " 付费" +
                                        datetime.now().strftime("%Y-%m-%d"), request)



    def sheet(self, template, items, title):
        file = load_workbook(template)
        sheet_names = file.sheetnames
        sheet = file[sheet_names[0]]
        row_title = sheet[1]
        row_title[0].value = title
        row_title[0].alignment = self._alignment()
        row_title[0].font = self._black_font()
        row_title[0].border = self._border()
        row_title[0].fill = self._white_background_color()
        row_header = sheet[2]
        for cell in row_header:
            cell.alignment = self._alignment()
            cell.font = self._white_font()
            cell.border = self._border()
            cell.fill = self._background_header_color()
        for index, item in enumerate(items):
            row = sheet[index+3]
            row[0].value = item['name']
            row[1].value = "是" if item['is_bind'] else "否"
            row[2].value = "是" if item['is_paid'] else "否"
            row[3].value = item['pay_amount']
            row[4].value = "剩余" + str(item['duration']) + "天"

            for cell in row:
                cell.alignment = self._alignment()
                cell.font = self._black_font()
                cell.border = self._border()
                cell.fill = self._white_background_color()


        with NamedTemporaryFile() as tmp:
            file.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream
